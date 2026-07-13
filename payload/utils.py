import os
import glob
import base64
import pandas as pd
from datetime import datetime
from pathlib import Path
from io import BytesIO
import shutil
from copy import copy



ROOT_DIR = Path(__file__).parent
TEMPLATE_DIR = ROOT_DIR / "templates"
UPLOAD_DIR = ROOT_DIR / "uploads"
RESULT_DIR = ROOT_DIR / "results"
HISTORY_DIR = ROOT_DIR / "history"
USER_DATA_DIR = ROOT_DIR / "user_data"

for folder in [TEMPLATE_DIR, UPLOAD_DIR, RESULT_DIR, HISTORY_DIR, USER_DATA_DIR]:
    folder.mkdir(exist_ok=True)


def image_to_base64(path):
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


def get_logo_path():
    for path in [
        ROOT_DIR / "logo.png",
        ROOT_DIR / "logo.jpg",
        ROOT_DIR / "logo.jpeg",
        ROOT_DIR / "oasis_logo_transparent_1200.png",
        ROOT_DIR / "oasis_logo_transparent_800.png",
    ]:
        if path.exists():
            return path
    return None


def logo_html(width=240):
    logo_path = get_logo_path()
    if logo_path:
        return (
            f"<img src='data:image/png;base64,{image_to_base64(logo_path)}' "
            f"style='width:{width}px; height:auto; object-fit:contain;'/>"
        )
    return "<div style='font-size:26px; font-weight:800; color:#0b2d66;'>OASIS</div>"


def make_upload_filename(original_name):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = original_name.replace(" ", "_").replace("(", "").replace(")", "")
    return f"업로드고객DB_{ts}_{safe_name}"


def find_customer_template():
    """기존 고객DB 양식 파일을 찾는다.

    v2.3.2부터 누적 고객DB는 새 파일을 만드는 방식이 아니라,
    기존 고객DB 양식/서식/시트 구조를 템플릿으로 복사해서 사용한다.
    """
    candidates = [
        TEMPLATE_DIR / "고객DB_양식v2.xlsx",
        ROOT_DIR / "고객DB_양식v2.xlsx",
        ROOT_DIR / "고객DB.xlsx",
        ROOT_DIR / "고객DB(16).xlsx",
    ]

    # 일부 환경에서 한글 파일명이 깨져 들어온 경우까지 대비해 xlsx 전체를 탐색한다.
    for folder in [TEMPLATE_DIR, ROOT_DIR]:
        if folder.exists():
            candidates.extend(folder.glob("*.xlsx"))

    seen = set()
    for path in candidates:
        path = Path(path)
        if path in seen or not path.exists() or path.name.startswith("~$"):
            continue
        seen.add(path)
        try:
            xls = pd.ExcelFile(path)
            if "고객DB" in xls.sheet_names:
                return path
        except Exception:
            continue

    return None


def make_basic_customer_template_bytes():
    customer_columns = [
        "업체명", "대표자명", "사업자등록번호", "업종명", "사업장 소재지",
        "설립년도", "연매출", "전년도매출", "올해예상매출", "매출감소여부",
        "상시근로자수", "고용보험가입인원", "신규채용계획", "청년채용계획",
        "희망상담주제1", "희망상담주제2", "희망상담주제3",
        "희망자금용도1", "희망자금용도2", "희망자금용도3",
        "키워드메모", "주요 사업내용", "비고",
        "벤처", "메인비즈", "이노비즈", "기업부설연구소",
        "특허보유", "R&D수행", "기술제품보유", "기술인력보유",
        "기술매출발생", "스마트공장도입", "기술보증희망", "기술성메모"
    ]

    sample = pd.DataFrame([{
        "업체명": "예시기업",
        "대표자명": "홍길동",
        "업종명": "제조업",
        "사업장 소재지": "충남 천안시",
        "전년도매출": 400000000,
        "올해예상매출": 300000000,
        "매출감소여부": "Y",
        "상시근로자수": 5,
        "고용보험가입인원": 5,
        "희망상담주제1": "정책자금",
        "희망자금용도1": "운전자금",
        "키워드메모": "제조 설비투자 운전자금",
        "주요 사업내용": "제품 제조 및 판매",
        "벤처": "N",
        "메인비즈": "N",
        "이노비즈": "N",
        "기업부설연구소": "N",
        "특허보유": "N",
        "R&D수행": "N",
        "기술제품보유": "N",
        "기술인력보유": "N",
        "기술매출발생": "N",
        "스마트공장도입": "N",
        "기술보증희망": "N"
    }], columns=customer_columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sample.to_excel(writer, sheet_name="고객DB", index=False)
        pd.DataFrame(columns=["상품명", "기관명", "대상지역", "대상업종", "추천키워드", "자금용도", "필요서류"]).to_excel(writer, sheet_name="상시정책자금DB", index=False)
        pd.DataFrame(columns=["제도명", "기관명", "대상", "최소고용인원", "추천키워드", "필요서류"]).to_excel(writer, sheet_name="고용지원금DB", index=False)
        pd.DataFrame({
            "안내": [
                "오아시스 내부용 고객DB 기본 양식입니다.",
                "파일명은 변경해도 되지만 시트명과 컬럼명은 변경하지 말아주세요.",
                "Y/N 항목은 Y 또는 N으로 입력하세요."
            ]
        }).to_excel(writer, sheet_name="사용가이드", index=False)

    output.seek(0)
    return output.getvalue()


def cleanup_old_files(folder, pattern, keep_count=30):
    files = list(Path(folder).glob(pattern))
    if len(files) <= keep_count:
        return

    files.sort(key=lambda x: x.stat().st_mtime, reverse=True)

    for old_file in files[keep_count:]:
        try:
            old_file.unlink()
        except Exception:
            pass


def run_cleanup():
    cleanup_old_files(RESULT_DIR, "매칭결과_*.xlsx", keep_count=30)
    cleanup_old_files(UPLOAD_DIR, "업로드고객DB_*.xlsx", keep_count=30)


def move_result_files_to_results(before_files):
    after_files = set(glob.glob("매칭결과_*.xlsx"))
    new_files = list(after_files - before_files)

    if not new_files:
        all_result_files = glob.glob("매칭결과_*.xlsx")
        if all_result_files:
            new_files = [max(all_result_files, key=os.path.getmtime)]

    moved_files = []
    for file in new_files:
        src = Path(file)
        dst = RESULT_DIR / src.name

        try:
            if src.exists():
                if dst.exists():
                    dst.unlink()
                src.replace(dst)
                moved_files.append(str(dst))
            elif dst.exists():
                moved_files.append(str(dst))
        except Exception:
            if src.exists():
                moved_files.append(str(src))

    return moved_files


def extract_company_previews(result_file):
    previews = {}

    try:
        xls = pd.ExcelFile(result_file)
        company_sheets = [s for s in xls.sheet_names if s.startswith("업체별_")]

        for sheet in company_sheets:
            df_raw = pd.read_excel(result_file, sheet_name=sheet, header=None)

            header_row = None
            for i in range(len(df_raw)):
                row_values = [str(v) for v in df_raw.iloc[i].tolist()]
                if "추천사업명" in row_values and "신청가능성점수" in row_values:
                    header_row = i
                    break

            if header_row is None:
                continue

            df = pd.read_excel(result_file, sheet_name=sheet, header=header_row)
            df = df.dropna(how="all")

            if "추천사업명" not in df.columns:
                continue

            cols = [
                col for col in [
                    "사업구분", "추천사업명", "기관명", "추천등급",
                    "신청가능성점수", "매칭점수", "신청판정"
                ]
                if col in df.columns
            ]

            company_name = sheet.replace("업체별_", "")
            previews[company_name] = df[cols].head(3)

    except Exception:
        pass

    return previews


# ================================
# v2.1 회원별 데이터 저장 유틸
# ================================
def safe_user_folder_name(user_id):
    import re
    user_id = str(user_id or "guest").strip().lower()
    user_id = re.sub(r"[^a-z0-9가-힣_.@-]", "_", user_id)
    return user_id[:50] or "guest"


def get_user_base_dir(user_id):
    user_dir = USER_DATA_DIR / safe_user_folder_name(user_id)
    user_dir.mkdir(parents=True, exist_ok=True)
    return user_dir


def get_user_dirs(user_id):
    base = get_user_base_dir(user_id)
    dirs = {
        "base": base,
        "uploads": base / "uploads",
        "results": base / "results",
        "history": base / "history",
    }
    for folder in dirs.values():
        folder.mkdir(parents=True, exist_ok=True)
    return dirs


def get_user_cumulative_db_path(user_id):
    return get_user_base_dir(user_id) / "고객DB누적.xlsx"


# 누적 고객DB도 기존 고객DB 양식과 동일하게 유지한다.
# 내부 관리용 메타 컬럼(누적저장일시/회원ID/담당자명)은 다운로드 파일에 넣지 않는다.
CUSTOMER_DB_SHEET_NAME = "고객DB"
LEGACY_CUMULATIVE_SHEET_NAME = "고객DB누적"
CUMULATIVE_META_COLUMNS = ["누적저장일시", "회원ID", "담당자명"]


def _normalize_customer_db_frame(df, columns=None):
    """누적 고객DB를 기존 고객DB 양식 컬럼 순서로 정리한다."""
    columns = columns or get_customer_db_columns()
    if df is None or df.empty:
        return pd.DataFrame(columns=columns)

    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    # v2.3에서 저장된 관리용 메타 컬럼은 제거한다.
    df = df.drop(columns=[c for c in CUMULATIVE_META_COLUMNS if c in df.columns], errors="ignore")

    # 기존 양식에 없는 컬럼은 매칭 안정성을 위해 제외하고, 빠진 컬럼은 공란으로 보충한다.
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    df = df[columns]
    df = df.dropna(how="all")
    return df


def _read_cumulative_customer_db(cumulative_path, columns=None):
    """신규/구버전 누적DB를 모두 읽어 기존 고객DB 양식으로 반환한다."""
    columns = columns or get_customer_db_columns()
    if not Path(cumulative_path).exists():
        return pd.DataFrame(columns=columns)

    for sheet_name in [CUSTOMER_DB_SHEET_NAME, LEGACY_CUMULATIVE_SHEET_NAME]:
        try:
            df = pd.read_excel(cumulative_path, sheet_name=sheet_name)
            return _normalize_customer_db_frame(df, columns)
        except Exception:
            continue
    return pd.DataFrame(columns=columns)


def _load_customer_template_workbook():
    """고객DB 양식 워크북을 openpyxl로 불러온다."""
    from openpyxl import load_workbook, Workbook

    template = find_customer_template()
    if template and Path(template).exists():
        return load_workbook(template)

    # 템플릿이 없는 예외 상황에서도 앱이 멈추지 않도록 최소 구조를 생성한다.
    wb = Workbook()
    ws = wb.active
    ws.title = CUSTOMER_DB_SHEET_NAME
    for idx, col in enumerate(get_customer_db_columns(), start=1):
        ws.cell(row=1, column=idx, value=col)
    wb.create_sheet("상시정책자금DB")
    wb.create_sheet("고용지원금DB")
    wb.create_sheet("코드표")
    wb.create_sheet("사용가이드")
    return wb


def _copy_row_style(ws, source_row, target_row, max_col):
    """서식이 있는 템플릿 행을 새 데이터 행에 복사한다."""
    if source_row < 1 or source_row == target_row:
        return
    try:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    except Exception:
        pass
    for col in range(1, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.border:
            dst.border = copy(src.border)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.font:
            dst.font = copy(src.font)
        if src.protection:
            dst.protection = copy(src.protection)


def _find_header_row(ws, required_header="업체명", scan_rows=10):
    for row in range(1, min(ws.max_row, scan_rows) + 1):
        values = [str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
        if required_header in values:
            return row
    return 1


def _prepare_template_customer_sheet(wb, columns):
    """템플릿의 고객DB 시트를 유지하되 기존 샘플/데이터 행은 비운다."""
    if CUSTOMER_DB_SHEET_NAME not in wb.sheetnames:
        ws = wb.active
        ws.title = CUSTOMER_DB_SHEET_NAME
    else:
        ws = wb[CUSTOMER_DB_SHEET_NAME]

    header_row = _find_header_row(ws)

    existing_headers = [str(ws.cell(row=header_row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    if "업체명" not in existing_headers:
        header_row = 1
        for idx, col in enumerate(columns, start=1):
            ws.cell(row=header_row, column=idx, value=col)
    else:
        # 템플릿 헤더를 기준으로 저장하되, 신규 컬럼이 누락되어 있으면 오른쪽에 보충한다.
        last_col = max(ws.max_column, len(existing_headers))
        for col in columns:
            if col not in existing_headers:
                last_col += 1
                ws.cell(row=header_row, column=last_col, value=col)
                existing_headers.append(col)

    # 헤더 아래 기존 샘플/데이터 행 제거. 서식은 첫 데이터 행 스타일 복사용으로 보존했다가 새 행에 다시 적용한다.
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    return ws, header_row


def _write_dataframe_to_customer_sheet(ws, header_row, df):
    headers = [str(ws.cell(row=header_row, column=col).value or "").strip() for col in range(1, ws.max_column + 1)]
    max_col = len(headers)
    style_source_row = header_row + 1

    for r_idx, (_, record) in enumerate(df.iterrows(), start=header_row + 1):
        _copy_row_style(ws, style_source_row, r_idx, max_col)
        for c_idx, header in enumerate(headers, start=1):
            if not header:
                continue
            value = record.get(header, "")
            if pd.isna(value):
                value = ""
            ws.cell(row=r_idx, column=c_idx, value=value)


def _write_cumulative_customer_db(cumulative_path, df, columns=None):
    """
    누적 고객DB를 기존 고객DB 양식과 동일하게 저장한다.

    - 고객DB 시트 컬럼/서식 유지
    - 상시정책자금DB/고용지원금DB/코드표/사용가이드 등 나머지 시트 유지
    - 기존 v2.3~v2.3.1 누적DB의 관리용 컬럼은 제거
    """
    columns = columns or get_customer_db_columns()
    df = _normalize_customer_db_frame(df, columns)

    wb = _load_customer_template_workbook()
    ws, header_row = _prepare_template_customer_sheet(wb, columns)
    _write_dataframe_to_customer_sheet(ws, header_row, df)

    cumulative_path = Path(cumulative_path)
    cumulative_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cumulative_path)
    return df




def ensure_user_cumulative_db_format(user_id):
    """
    기존 v2.3~v2.3.1에서 생성된 고객DB누적.xlsx를
    현재 고객DB 템플릿 형식으로 강제 변환한다.

    - 고객DB 시트명으로 저장
    - 상시정책자금DB/고용지원금DB/코드표/사용가이드 유지
    - 누적저장일시/회원ID/담당자명 제거
    - 이미 최신 형식이어도 템플릿 기반으로 다시 저장해 서식/시트 구조를 보정
    """
    cumulative_path = get_user_cumulative_db_path(user_id)
    if not cumulative_path.exists():
        return cumulative_path, 0, False

    columns = get_customer_db_columns()
    df = _read_cumulative_customer_db(cumulative_path, columns)
    _write_cumulative_customer_db(cumulative_path, df, columns)
    return cumulative_path, len(df), True


def count_user_cumulative_rows(user_id):
    """누적 고객DB의 현재 고객 행 수를 반환한다."""
    cumulative_path = get_user_cumulative_db_path(user_id)
    if not cumulative_path.exists():
        return 0
    try:
        df = _read_cumulative_customer_db(cumulative_path)
        return int(len(df))
    except Exception:
        return 0

def append_user_customer_db(uploaded_excel_path, user_id, manager_name=""):
    """
    v3.5.2:
    고객DB 업로드 시 사업자등록번호가 이미 누적DB에 존재하는 행은 추가하지 않는다.
    같은 업로드 파일 안에서 반복된 사업자등록번호도 첫 번째 행만 저장한다.
    사업자등록번호가 비어 있는 행은 자동 중복판단이 불가능하므로 기존처럼 저장한다.
    """
    cumulative_path = get_user_cumulative_db_path(user_id)
    try:
        columns = get_customer_db_columns()
        df_new = pd.read_excel(uploaded_excel_path, sheet_name=CUSTOMER_DB_SHEET_NAME)
        df_new = _normalize_customer_db_frame(df_new, columns)
        if df_new.empty:
            return cumulative_path, 0

        df_old = _read_cumulative_customer_db(cumulative_path, columns)

        if "사업자등록번호" in df_new.columns:
            normalized_new = df_new["사업자등록번호"].apply(normalize_business_no)
            has_business_no = normalized_new.astype(str).str.replace("-", "", regex=False).str.len().eq(10)

            # 업로드 파일 내부 중복 제거: 사업자번호가 있는 행만 적용
            duplicated_in_upload = has_business_no & normalized_new.duplicated(keep="first")
            df_new = df_new.loc[~duplicated_in_upload].copy()
            normalized_new = df_new["사업자등록번호"].apply(normalize_business_no)
            has_business_no = normalized_new.astype(str).str.replace("-", "", regex=False).str.len().eq(10)

            # 기존 누적DB와 비교하여 이미 등록된 사업자번호 제외
            if not df_old.empty and "사업자등록번호" in df_old.columns:
                old_numbers = {
                    normalize_business_no(value)
                    for value in df_old["사업자등록번호"].tolist()
                    if len(normalize_business_no(value).replace("-", "")) == 10
                }
                already_exists = has_business_no & normalized_new.isin(old_numbers)
                df_new = df_new.loc[~already_exists].copy()

        if df_new.empty:
            return cumulative_path, 0

        df_all = (
            pd.concat([df_old, df_new], ignore_index=True, sort=False)
            if not df_old.empty
            else df_new
        )
        _write_cumulative_customer_db(cumulative_path, df_all, columns)
        return cumulative_path, len(df_new)

    except Exception:
        return cumulative_path, 0


# ================================
# v2.3 크레탑 PDF 자동 고객등록 유틸
# ================================
def _clean_text(value):
    return str(value or "").replace("\xa0", " ").strip()


def _compact_text(value):
    return "".join(_clean_text(value).split())


def extract_pdf_text(pdf_path):
    """텍스트형 PDF에서 텍스트를 추출한다. 스캔본 PDF는 추출이 제한될 수 있다."""
    pdf_path = Path(pdf_path)
    errors = []

    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        text = "\n".join(pages)
        if text.strip():
            return text, ""
    except Exception as e:
        errors.append(f"pypdf: {e}")

    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(str(pdf_path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        text = "\n".join(pages)
        if text.strip():
            return text, ""
    except Exception as e:
        errors.append(f"PyPDF2: {e}")

    return "", "PDF 텍스트 추출에 실패했습니다. 텍스트형 크레탑 PDF인지 확인해주세요. " + " / ".join(errors[:2])


def _regex_first(pattern, text, default="", flags=0):
    import re
    m = re.search(pattern, text, flags)
    if not m:
        return default
    return _clean_text(m.group(1))


def _parse_latest_number_from_line(label, block):
    """표 안에서 지정 계정명의 가장 오른쪽 숫자를 반환한다."""
    import re

    escaped = re.escape(label)
    patterns = [
        rf"(?m)^\s*{escaped}(?:\(손실\)|\(순손실\)|\(\*\)|\(.*?\))?\s+((?:-?\s*[0-9][0-9,]*(?:\.[0-9]+)?\s+)+)",
        rf"{escaped}(?:\(손실\)|\(순손실\)|\(\*\)|\(.*?\))?\s+((?:-?\s*[0-9][0-9,]*(?:\.[0-9]+)?\s+)+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, block)
        if not match:
            continue

        raw_values = re.findall(r"-?\s*[0-9][0-9,]*(?:\.[0-9]+)?", match.group(1))
        if not raw_values:
            continue

        raw = raw_values[-1].replace(" ", "").replace(",", "")
        try:
            return float(raw)
        except ValueError:
            continue

    return None


def _extract_financial_block(text, start_heading, end_headings):
    """특정 재무표 제목부터 다음 표 제목 전까지의 텍스트를 잘라낸다."""
    import re

    start = re.search(start_heading, text)
    if not start:
        return ""

    tail = text[start.end():]
    end_positions = []
    for heading in end_headings:
        end_match = re.search(heading, tail)
        if end_match:
            end_positions.append(end_match.start())

    end = min(end_positions) if end_positions else min(len(tail), 6000)
    return tail[:end]


def _latest_summary_amount_million(label, text):
    """
    크레탑 요약 재무표의 최신년도 값을 원 단위로 변환한다.

    우선순위:
    1. 요약 손익계산서/요약 재무상태표(백만원)
    2. 상세 손익계산서/재무상태표(천원)
    3. MY 재무 Data(백만원)
    """
    income_labels = {"매출액", "영업이익", "당기순이익"}
    balance_labels = {"자산총계", "부채총계", "자본총계"}

    # 1) 요약표: 단위 백만원
    if label in income_labels:
        summary_block = _extract_financial_block(
            text,
            r"요약\s*손익계산서",
            [r"요약\s*현금흐름", r"요약\s*재무비율", r"연혁"],
        )
    elif label in balance_labels:
        summary_block = _extract_financial_block(
            text,
            r"요약\s*재무상태표",
            [r"요약\s*손익계산서", r"요약\s*현금흐름", r"요약\s*재무비율"],
        )
    else:
        summary_block = ""

    value = _parse_latest_number_from_line(label, summary_block)
    if value is not None:
        return int(round(value * 1_000_000))

    # 2) 상세표: 단위 천원
    if label in income_labels:
        detail_label = {
            "영업이익": "영업이익",
            "당기순이익": "당기순이익",
        }.get(label, label)
        detail_block = _extract_financial_block(
            text,
            r"손익계산서\s+단위\s*:?\s*천원",
            [r"현금흐름표", r"자본변동표", r"이익잉여금처분계산서"],
        )
        value = _parse_latest_number_from_line(detail_label, detail_block)
        if value is None and label == "당기순이익":
            value = _parse_latest_number_from_line("당기순이익(순손실)", detail_block)
    else:
        detail_name = {
            "자산총계": "자산",
            "부채총계": "부채",
            "자본총계": "자본",
        }.get(label, label)
        detail_block = _extract_financial_block(
            text,
            r"재무상태표\s+단위\s*:?\s*천원",
            [r"손익계산서", r"현금흐름표"],
        )
        value = _parse_latest_number_from_line(detail_name, detail_block)

    if value is not None:
        return int(round(value * 1_000))

    # 3) MY 재무 Data: 단위 백만원
    my_data_block = _extract_financial_block(
        text,
        r"MY\s*재무\s*Data",
        [r"기술력", r"기업인증", r"주요\s*주주"],
    )
    fallback_names = {
        "자산총계": "자산",
        "부채총계": "부채",
        "자본총계": "자본",
    }
    value = _parse_latest_number_from_line(
        fallback_names.get(label, label),
        my_data_block,
    )
    if value is not None:
        return int(round(value * 1_000_000))

    return ""


def _latest_amount_million(label, text):
    """기존 호출부 호환용."""
    return _latest_summary_amount_million(label, text)


def _yn_from_status(status):
    status = _compact_text(status)
    if not status or status in ["-", "없음", "해당사항없음", "조회된자료가없습니다."]:
        return "N"
    if "미인증" in status or "미보유" in status or "없음" in status:
        return "N"
    return "Y"


def _extract_certifications(text):
    import re
    cert = {
        "벤처": "N", "이노비즈": "N", "메인비즈": "N",
        "연구개발전담부서": "N", "기업부설연구소": "N",
        "특허보유": "N", "실용신안": "N", "디자인": "N", "상표": "N",
    }

    # 크레탑 4페이지의 대표 형태:
    # 기업인증 벤처 이노비즈 메인비즈 연구개발전담부서 부설연구소 미인증 ...
    m = re.search(
        r"기업인증\s+벤처\s+이노비즈\s+메인비즈\s+연구개발전담부서\s+부설연구소\s+(.+?)\s+산업재산권",
        text,
        re.S,
    )
    if m:
        statuses = [_clean_text(x) for x in m.group(1).split() if _clean_text(x)]
        labels = ["벤처", "이노비즈", "메인비즈", "연구개발전담부서", "기업부설연구소"]
        for label, status in zip(labels, statuses[:5]):
            cert[label] = _yn_from_status(status)

    m = re.search(
        r"산업재산권\s+특허\s+실용신안\s+디자인\s+상표권\s+(.+?)\s+주요\s*주주",
        text,
        re.S,
    )
    if m:
        statuses = [_clean_text(x) for x in m.group(1).split() if _clean_text(x)]
        labels = ["특허보유", "실용신안", "디자인", "상표"]
        for label, status in zip(labels, statuses[:4]):
            cert[label] = _yn_from_status(status)

    return cert


def extract_cretop_identity(pdf_path, max_pages=2):
    """PDF 앞부분만 읽어 업체명과 사업자등록번호를 빠르게 추출한다."""
    pdf_path = Path(pdf_path)
    errors = []
    text = ""

    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        text = "\n".join((page.extract_text() or "") for page in reader.pages[:max_pages])
    except Exception as exc:
        errors.append(f"pypdf: {exc}")

    if not text.strip():
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(pdf_path))
            text = "\n".join((page.extract_text() or "") for page in reader.pages[:max_pages])
        except Exception as exc:
            errors.append(f"PyPDF2: {exc}")

    if not text.strip():
        return {}, "PDF 앞부분에서 사업자정보를 읽지 못했습니다. " + " / ".join(errors[:2])

    company_name = _regex_first(r"기업명\s+(.+?)\s+영문기업명", text)
    if not company_name:
        company_name = _regex_first(r"기업명\s*[:：]\s*(.+?)\s+사업자번호", text)

    business_no = _regex_first(
        r"사업자번호\s+([0-9]{3}-[0-9]{2}-[0-9]{5})",
        text,
    )
    if not business_no:
        business_no = _regex_first(r"([0-9]{3}-[0-9]{2}-[0-9]{5})", text)

    return {
        "업체명": company_name,
        "사업자등록번호": normalize_business_no(business_no),
    }, ""


def _read_business_numbers_only(cumulative_path):
    """누적DB 전체가 아닌 사업자등록번호 열만 읽는다."""
    from openpyxl import load_workbook

    path = Path(cumulative_path)
    if not path.exists():
        return set()

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if CUSTOMER_DB_SHEET_NAME not in workbook.sheetnames:
            return set()

        worksheet = workbook[CUSTOMER_DB_SHEET_NAME]
        header_row = None
        business_col = None

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
            start=1,
        ):
            for col_index, value in enumerate(row, start=1):
                if str(value or "").strip() == "사업자등록번호":
                    header_row = row_index
                    business_col = col_index
                    break
            if business_col:
                break

        if not business_col:
            return set()

        numbers = set()
        for row in worksheet.iter_rows(
            min_row=(header_row or 1) + 1,
            min_col=business_col,
            max_col=business_col,
            values_only=True,
        ):
            number = normalize_business_no(row[0])
            if len(number.replace("-", "")) == 10:
                numbers.add(number)
        return numbers
    finally:
        workbook.close()


def extract_cretop_pdf_data(pdf_path):
    """크레탑 기업종합보고서 PDF에서 고객DB 자동입력용 값을 추출한다."""
    text, error = extract_pdf_text(pdf_path)
    if error:
        return {}, error

    compact = _compact_text(text)
    data = {}

    data["업체명"] = _regex_first(r"기업명\s+(.+?)\s+영문기업명", text)
    if not data["업체명"]:
        data["업체명"] = _regex_first(r"기업명\s*[:：]\s*(.+?)\s+사업자번호", text, flags=0)
    data["사업자등록번호"] = _regex_first(r"사업자번호\s+([0-9]{3}-[0-9]{2}-[0-9]{5})", text)
    if not data["사업자등록번호"]:
        data["사업자등록번호"] = _regex_first(r"([0-9]{3}-[0-9]{2}-[0-9]{5})", text)
    data["법인등록번호"] = _regex_first(r"법인\(주민\)번호\s+([0-9\-]+)", text)
    data["대표자명"] = _regex_first(r"대표자명\s+(.+?)\s+종업원수", text)
    data["종업원수"] = _regex_first(r"종업원수\s+([0-9,]+)\s*명", text)
    data["설립일"] = _regex_first(r"설립년월\s+([0-9]{4}-[0-9]{2}-[0-9]{2})", text)
    data["설립년도"] = data["설립일"][:4] if data.get("설립일") else ""
    data["기업유형"] = _regex_first(r"기업유형\s+(.+?)\s+기업규모", text)
    data["기업규모"] = _regex_first(r"기업규모\s+(.+?)(?:\n|전화번호|팩스번호)", text)
    data["사업장 소재지"] = _regex_first(r"주소\s+(.+?)\s+표준산업분류\(10차\)", text, flags=re.S) if False else ""

    # re.S를 함수 인자로 넘기기 위해 별도 처리
    import re
    data["사업장 소재지"] = _regex_first(r"주소\s+(.+?)\s+표준산업분류\(10차\)", text, flags=re.S)
    data["사업장 소재지"] = " ".join(data["사업장 소재지"].split())
    data["업종명"] = _regex_first(r"표준산업분류\(10차\)\s+\([A-Z0-9]+\)\s*(.+?)\s+표준산업분류\(11차\)", text, flags=re.S)
    data["업종명"] = " ".join(data["업종명"].split())

    data["매출액"] = _latest_summary_amount_million("매출액", text)
    data["연매출"] = data["매출액"]
    data["전년도매출"] = data["매출액"]
    data["영업이익"] = _latest_summary_amount_million("영업이익", text)
    data["당기순이익"] = _latest_summary_amount_million("당기순이익", text)
    data["자산총계"] = _latest_summary_amount_million("자산총계", text)
    data["부채총계"] = _latest_summary_amount_million("부채총계", text)
    data["자본총계"] = _latest_summary_amount_million("자본총계", text)

    cert = _extract_certifications(text)
    data.update(cert)

    business_purpose = _regex_first(r"사업목적\s+내용\s+(.+?)\s+종합의견", text, flags=re.S)
    if business_purpose:
        lines = [x.strip() for x in business_purpose.split("\n") if x.strip()]
        data["주요 사업내용"] = " / ".join(lines[:8])[:500]
    else:
        data["주요 사업내용"] = data.get("업종명", "")

    keyword_parts = [data.get("업종명", "")]
    if data.get("벤처") == "Y":
        keyword_parts.append("벤처")
    if data.get("이노비즈") == "Y":
        keyword_parts.append("이노비즈")
    if data.get("메인비즈") == "Y":
        keyword_parts.append("메인비즈")
    if data.get("기업부설연구소") == "Y" or data.get("연구개발전담부서") == "Y":
        keyword_parts.append("연구소")
    if data.get("특허보유") == "Y":
        keyword_parts.append("특허")
    data["키워드메모"] = " / ".join([x for x in keyword_parts if x])

    data["PDF추출일시"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return data, ""


def get_customer_db_columns():
    template = find_customer_template()
    if template and Path(template).exists():
        try:
            df = pd.read_excel(template, sheet_name="고객DB", nrows=0)
            cols = [str(c).strip() for c in df.columns if str(c).strip()]
            if cols:
                if "사업자등록번호" not in cols:
                    cols.append("사업자등록번호")
                return cols
        except Exception:
            pass

    return [
        "업체명", "대표자명", "사업자등록번호", "업종명", "사업장 소재지",
        "설립년도", "연매출", "전년도매출", "올해예상매출", "매출감소여부",
        "상시근로자수", "고용보험가입인원", "신규채용계획", "청년채용계획",
        "희망상담주제1", "희망상담주제2", "희망상담주제3",
        "희망자금용도1", "희망자금용도2", "희망자금용도3",
        "키워드메모", "주요 사업내용", "비고",
        "벤처", "메인비즈", "이노비즈", "기업부설연구소",
        "특허보유", "R&D수행", "기술제품보유", "기술인력보유",
        "기술매출발생", "스마트공장도입", "기술보증희망", "기술성메모"
    ]


def build_customer_row_from_cretop(data, columns=None):
    columns = columns or get_customer_db_columns()
    row = {col: "" for col in columns}

    alias_map = {
        "업체명": ["업체명", "기업명"],
        "대표자명": ["대표자명", "대표자"],
        "사업자등록번호": ["사업자등록번호", "사업자번호"],
        "법인등록번호": ["법인등록번호", "법인번호", "법인(주민)번호"],
        "업종명": ["업종명", "표준산업분류"],
        "사업장 소재지": ["사업장 소재지", "주소", "소재지"],
        "설립년도": ["설립년도"],
        "설립일": ["설립일", "설립년월"],
        "연매출": ["연매출", "매출액"],
        "전년도매출": ["전년도매출", "매출액"],
        "매출액": ["매출액"],
        "영업이익": ["영업이익"],
        "당기순이익": ["당기순이익"],
        "자산총계": ["자산총계"],
        "부채총계": ["부채총계"],
        "자본총계": ["자본총계"],
        "상시근로자수": ["상시근로자수", "종업원수"],
        "종업원수": ["종업원수"],
        "키워드메모": ["키워드메모"],
        "주요 사업내용": ["주요 사업내용"],
        "벤처": ["벤처"],
        "메인비즈": ["메인비즈"],
        "이노비즈": ["이노비즈"],
        "기업부설연구소": ["기업부설연구소"],
        "부설연구소": ["기업부설연구소"],
        "연구소": ["기업부설연구소"],
        "연구개발전담부서": ["연구개발전담부서"],
        "특허보유": ["특허보유"],
        "특허": ["특허보유"],
        "상표": ["상표"],
        "상표권": ["상표"],
        "R&D수행": ["R&D수행"],
    }

    # 기본 복사
    for col in columns:
        candidates = alias_map.get(col, [col])
        for key in candidates:
            if key in data and data.get(key) not in [None, ""]:
                row[col] = data.get(key)
                break

    # R&D수행은 연구소/전담부서가 있으면 Y로 보조 판정
    if "R&D수행" in row and not row.get("R&D수행"):
        row["R&D수행"] = "Y" if data.get("기업부설연구소") == "Y" or data.get("연구개발전담부서") == "Y" else "N"

    # 기술성 기본값 보정
    for col in ["벤처", "메인비즈", "이노비즈", "기업부설연구소", "특허보유", "R&D수행"]:
        if col in row and row[col] == "":
            row[col] = "N"

    if "비고" in row:
        row["비고"] = f"크레탑 PDF 자동등록({data.get('PDF추출일시', '')})"

    return row


def normalize_business_no(value):
    import re
    digits = re.sub(r"[^0-9]", "", str(value or ""))
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return str(value or "").strip()


def normalize_company_name(value):
    import re
    text = str(value or "").strip().lower()
    text = re.sub(r"주식회사|\(주\)|㈜|（주）", "", text)
    return re.sub(r"[^0-9a-z가-힣]", "", text)


def normalize_person_name(value):
    import re
    return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").strip().lower())


def _find_legacy_customer_matches(
    cumulative_path,
    company_name="",
    representative_name="",
):
    columns = get_customer_db_columns()
    df = _read_cumulative_customer_db(cumulative_path, columns)
    if df.empty or "업체명" not in df.columns:
        return pd.DataFrame(columns=columns)

    company_key = normalize_company_name(company_name)
    representative_key = normalize_person_name(representative_name)
    if not company_key:
        return pd.DataFrame(columns=columns)

    mask = df["업체명"].apply(normalize_company_name).eq(company_key)
    if representative_key and "대표자명" in df.columns:
        mask = mask & df["대표자명"].apply(normalize_person_name).eq(
            representative_key
        )
    return df.loc[mask].copy()


def check_user_customer_duplicate(
    user_id,
    business_no,
    company_name="",
    representative_name="",
):
    cumulative_path = get_user_cumulative_db_path(user_id)
    target = normalize_business_no(business_no)

    if not cumulative_path.exists():
        return False

    try:
        if len(target.replace("-", "")) == 10:
            if target in _read_business_numbers_only(cumulative_path):
                return True

        legacy_matches = _find_legacy_customer_matches(
            cumulative_path,
            company_name=company_name,
            representative_name=representative_name,
        )
        return not legacy_matches.empty
    except Exception:
        return False


def link_business_no_to_legacy_customer(
    user_id,
    business_no,
    company_name="",
    representative_name="",
):
    cumulative_path = get_user_cumulative_db_path(user_id)
    target = normalize_business_no(business_no)

    if not cumulative_path.exists() or len(target.replace("-", "")) != 10:
        return False

    columns = get_customer_db_columns()
    df = _read_cumulative_customer_db(cumulative_path, columns)
    if df.empty or "업체명" not in df.columns:
        return False

    company_key = normalize_company_name(company_name)
    representative_key = normalize_person_name(representative_name)

    mask = df["업체명"].apply(normalize_company_name).eq(company_key)
    if representative_key and "대표자명" in df.columns:
        mask = mask & df["대표자명"].apply(normalize_person_name).eq(
            representative_key
        )

    indexes = list(df.index[mask])
    if not indexes:
        return False

    index = indexes[0]
    current = normalize_business_no(df.at[index, "사업자등록번호"])
    if len(current.replace("-", "")) != 10:
        df.at[index, "사업자등록번호"] = target
        _write_cumulative_customer_db(cumulative_path, df, columns)
    return True


def append_cretop_to_user_customer_db(pdf_path, user_id, manager_name="", duplicate_action="skip", extracted_data=None):
    """
    v3.5.2:
    크레탑 자동등록 시 동일 사업자등록번호가 누적DB에 있으면
    사용자의 선택과 관계없이 새 행을 추가하지 않는다.
    """
    if extracted_data is None:
        data, error = extract_cretop_pdf_data(pdf_path)
    else:
        data = dict(extracted_data)
        error = ""

    cumulative_path = get_user_cumulative_db_path(user_id)
    if error:
        return cumulative_path, 0, error, data, pd.DataFrame()

    columns = get_customer_db_columns()
    row = build_customer_row_from_cretop(data, columns)
    df_new = pd.DataFrame([row], columns=columns)
    df_new = _normalize_customer_db_frame(df_new, columns)

    business_no = normalize_business_no(row.get("사업자등록번호", ""))
    company_name = row.get("업체명", data.get("업체명", ""))
    representative_name = row.get("대표자명", data.get("대표자명", ""))

    is_dup = check_user_customer_duplicate(
        user_id,
        business_no,
        company_name=company_name,
        representative_name=representative_name,
    )

    if is_dup:
        link_business_no_to_legacy_customer(
            user_id,
            business_no,
            company_name=company_name,
            representative_name=representative_name,
        )
        return (
            cumulative_path,
            0,
            f"사업자등록번호 {business_no}는 이미 누적 고객DB에 등록되어 있어 추가하지 않았습니다.",
            data,
            df_new,
        )

    # 신규 고객일 때만 누적DB 전체를 읽는다.
    df_old = _read_cumulative_customer_db(cumulative_path, columns)
    df_all = (
        pd.concat([df_old, df_new], ignore_index=True, sort=False)
        if not df_old.empty
        else df_new
    )
    _write_cumulative_customer_db(cumulative_path, df_all, columns)

    return (
        cumulative_path,
        1,
        "크레탑 PDF 추출값을 내 누적 고객DB에 추가했습니다.",
        data,
        df_new,
    )
