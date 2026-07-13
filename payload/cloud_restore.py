from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from cloud_db import (
    CloudDatabase,
    TABLE_CUSTOMERS,
    cloud_is_configured,
)
from utils import (
    find_customer_template,
    get_user_cumulative_db_path,
)


def _is_blank(value: Any) -> bool:
    return value is None or str(value).strip().lower() in {
        "",
        "nan",
        "none",
        "nat",
    }


def _local_customer_count(path: Path) -> int:
    if not path.exists():
        return 0

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "고객DB" not in workbook.sheetnames:
            return 0

        worksheet = workbook["고객DB"]
        header_row = None
        company_column = None

        for row_number in range(1, min(worksheet.max_row, 50) + 1):
            for column_number in range(1, worksheet.max_column + 1):
                value = str(
                    worksheet.cell(row_number, column_number).value or ""
                ).strip()
                if value == "업체명":
                    header_row = row_number
                    company_column = column_number
                    break
            if header_row:
                break

        if not header_row or not company_column:
            return 0

        return sum(
            1
            for row_number in range(header_row + 1, worksheet.max_row + 1)
            if not _is_blank(
                worksheet.cell(row_number, company_column).value
            )
        )
    except Exception:
        return 0


def _make_blank_workbook(path: Path, columns: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "고객DB"

    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)

    for column_number, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(1, column_number)
        cell.value = column_name
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _prepare_local_workbook(
    destination: Path,
    required_columns: list[str],
):
    destination.parent.mkdir(parents=True, exist_ok=True)

    template = find_customer_template()
    if template and Path(template).exists():
        shutil.copy2(template, destination)
    else:
        _make_blank_workbook(destination, required_columns)

    workbook = load_workbook(destination)
    worksheet = (
        workbook["고객DB"]
        if "고객DB" in workbook.sheetnames
        else workbook.create_sheet("고객DB", 0)
    )

    header_row = None
    columns: dict[str, int] = {}

    for row_number in range(1, min(worksheet.max_row, 50) + 1):
        row_columns: dict[str, int] = {}
        for column_number in range(1, worksheet.max_column + 1):
            value = str(
                worksheet.cell(row_number, column_number).value or ""
            ).strip()
            if value:
                row_columns[value] = column_number

        if "업체명" in row_columns:
            header_row = row_number
            columns = row_columns
            break

    if header_row is None:
        header_row = 1
        for column_number, column_name in enumerate(
            required_columns,
            start=1,
        ):
            worksheet.cell(header_row, column_number).value = column_name
            columns[column_name] = column_number

    for column_name in required_columns:
        if column_name not in columns:
            column_number = worksheet.max_column + 1
            worksheet.cell(header_row, column_number).value = column_name
            columns[column_name] = column_number

    if worksheet.max_row > header_row:
        worksheet.delete_rows(
            header_row + 1,
            worksheet.max_row - header_row,
        )

    return workbook, worksheet, header_row, columns


def restore_customer_db_if_needed(user_id: str) -> dict[str, Any]:
    """
    로컬 고객DB가 비어 있을 때만 Supabase 자료로 복원한다.
    기존 로컬 고객이 있으면 절대 덮어쓰지 않는다.
    """
    result = {
        "restored": False,
        "count": 0,
        "message": "",
    }

    if not user_id:
        result["message"] = "사용자 ID가 없습니다."
        return result

    local_path = get_user_cumulative_db_path(user_id)
    local_count = _local_customer_count(local_path)

    if local_count > 0:
        result["count"] = local_count
        result["message"] = "기존 로컬 고객DB를 사용합니다."
        return result

    if not cloud_is_configured():
        result["message"] = "Supabase가 설정되지 않아 자동복원을 건너뛰었습니다."
        return result

    try:
        records = CloudDatabase().select(
            TABLE_CUSTOMERS,
            filters={"owner_user_id": user_id},
            columns=(
                "business_no,company_name,representative_name,"
                "industry_name,address,manager_name,customer_data,"
                "created_at,updated_at"
            ),
            order="created_at.asc",
        )
    except Exception as exc:
        result["message"] = f"Supabase 고객조회 실패: {exc}"
        return result

    if not records:
        result["message"] = "Supabase에 복원할 고객자료가 없습니다."
        return result

    required_columns = [
        "업체명",
        "대표자명",
        "사업자등록번호",
        "업종명",
        "사업장 소재지",
        "담당자",
    ]
    merged_rows = []

    for record in records:
        customer_data = record.get("customer_data", {})
        if not isinstance(customer_data, dict):
            customer_data = {}

        row = dict(customer_data)
        fallback_values = {
            "업체명": record.get("company_name"),
            "대표자명": record.get("representative_name"),
            "사업자등록번호": record.get("business_no"),
            "업종명": record.get("industry_name"),
            "사업장 소재지": record.get("address"),
            "담당자": record.get("manager_name"),
        }

        for key, value in fallback_values.items():
            if _is_blank(row.get(key)) and not _is_blank(value):
                row[key] = value

        for key in row.keys():
            key_text = str(key).strip()
            if key_text and key_text not in required_columns:
                required_columns.append(key_text)

        merged_rows.append(row)

    try:
        workbook, worksheet, header_row, columns = _prepare_local_workbook(
            local_path,
            required_columns,
        )

        for row_offset, row_data in enumerate(merged_rows, start=1):
            row_number = header_row + row_offset
            for column_name, column_number in columns.items():
                value = row_data.get(column_name, "")
                if isinstance(value, (dict, list)):
                    value = str(value)
                worksheet.cell(row_number, column_number).value = value

        workbook.save(local_path)

        restored_count = _local_customer_count(local_path)
        result["restored"] = restored_count > 0
        result["count"] = restored_count
        result["message"] = (
            f"Supabase에서 고객 {restored_count}건을 자동 복원했습니다."
        )
        return result
    except Exception as exc:
        result["message"] = f"고객DB 복원 실패: {exc}"
        return result
