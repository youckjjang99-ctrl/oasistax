from __future__ import annotations

import inspect
import unittest
from datetime import datetime, timedelta, timezone
from io import BytesIO
from unittest.mock import patch
from zipfile import ZipFile

from openpyxl import load_workbook

from claim_correction_center import (
    _build_claim_documents_zip,
    _build_claim_results_excel,
    _claim_download_cache_fingerprint,
    _claim_downloadable_documents,
    _claim_document_download_name,
    _claim_document_scope_label,
    _format_claim_datetime,
    _plan_claim_document_zip_parts,
    _claim_result_case_view,
    _render_auto_claim_monitor,
    _render_personal_request,
    _render_results_tab,
    _render_status_tab,
    render_claim_correction_center,
)


class _ZipResponse:
    def __init__(self, content: bytes):
        self.content = content

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False

    def raise_for_status(self):
        return None

    def iter_content(self, chunk_size: int):
        del chunk_size
        yield self.content


class _BatchZipRepository:
    def __init__(self):
        self.batch_calls = []

    def document_download_urls(self, case_id, document_ids):
        self.batch_calls.append((case_id, list(document_ids)))
        return [
            f"https://example.invalid/{document_id}"
            for document_id in document_ids
        ]

    def document_download_url(self, *_args, **_kwargs):
        raise AssertionError("ZIP creation must use the batch link method")


class ClaimResultsViewTests(unittest.TestCase):
    def setUp(self) -> None:
        self.now = datetime(2026, 7, 31, 12, 0, tzinfo=timezone.utc)

    def _document(self, document_id: str, **overrides):
        document = {
            "id": document_id,
            "status": "ready",
            "storage_bucket": "oasis-claim-documents",
            "storage_path": f"owner/case/{document_id}.pdf",
            "content_type": "application/pdf",
            "retention_until": (
                self.now + timedelta(days=1)
            ).isoformat(),
            "facts": {},
            "deleted_at": None,
        }
        document.update(overrides)
        return document

    def test_claim_datetime_is_rendered_in_korea_time(self) -> None:
        self.assertEqual(
            _format_claim_datetime("2026-08-01T00:17:53+00:00"),
            "2026-08-01 09:17",
        )
        self.assertEqual(
            _format_claim_datetime(
                "2026-07-31T15:30:00Z",
                include_time=False,
            ),
            "2026-08-01",
        )

    def test_result_case_view_converts_utc_request_time(self) -> None:
        view = _claim_result_case_view(
            {
                "company_name": "오아시스 세무",
                "business_type": "individual",
                "overall_status": "ready",
                "requested_at": "2026-08-01T00:17:53+00:00",
            }
        )

        self.assertEqual(view["requested_at"], "2026-08-01 09:17")

    def test_downloadable_documents_keeps_only_valid_private_ready_files(
        self,
    ) -> None:
        ready_pdf = self._document("ready-pdf")
        ready_excel = self._document(
            "ready-excel",
            storage_path="owner/case/ready-excel.xlsx",
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )
        documents = [
            ready_pdf,
            self._document("no-data", facts={"no_data": True}),
            self._document("failed", status="failed"),
            self._document(
                "deleted",
                deleted_at="2026-07-31T00:00:00+00:00",
            ),
            self._document(
                "expired",
                retention_until=(
                    self.now - timedelta(seconds=1)
                ).isoformat(),
            ),
            self._document(
                "public-bucket",
                storage_bucket="public-documents",
            ),
            self._document(
                "unsupported",
                storage_path="owner/case/unsupported.zip",
                content_type="application/zip",
            ),
            self._document(
                "missing-path",
                storage_path="",
            ),
            self._document(
                "bad-retention",
                retention_until="not-a-date",
            ),
            ready_excel,
        ]

        selected = _claim_downloadable_documents(
            documents,
            now=self.now,
        )

        self.assertEqual(
            [document["id"] for document in selected],
            ["ready-pdf", "ready-excel"],
        )

    def test_result_case_view_is_one_compact_masked_customer_row(self) -> None:
        case = {
            "company_name": "오아시스 세무",
            "representative_name_masked": "홍*동",
            "business_no_masked": "120-**-*****",
            "phone_masked": "010-****-5678",
            "business_type": "individual",
            "hometax_status": "auth_complete",
            "comwel_status": "auth_complete",
            "overall_status": "ready",
            "requested_at": "2026-07-31T15:42:59+09:00",
            "requested_by": "담당자",
        }
        documents = [
            self._document("ready"),
            self._document("failed", status="failed"),
        ]

        view = _claim_result_case_view(
            case,
            documents,
            now=self.now,
        )

        self.assertEqual(
            view,
            {
                "customer_name": "오아시스 세무",
                "business_no": "120-**-*****",
                "phone": "010-****-5678",
                "business_type": "개인",
                "hometax_status": "인증 완료",
                "comwel_status": "인증 완료",
                "overall_status": "수집 완료",
                "requested_at": "2026-07-31 15:42",
                "requested_by": "담당자",
                "downloadable_document_count": 1,
            },
        )
        self.assertNotIn("1208800767", repr(view))

    def test_result_case_view_falls_back_to_masked_customer_name(self) -> None:
        view = _claim_result_case_view(
            {
                "company_name": "상호명 미입력",
                "representative_name_masked": "김*수",
                "business_type": "individual",
                "overall_status": "auth_complete_collection_pending",
                "last_safe_error_code": "COLLECTION_PARTIAL",
            },
            [],
            now=self.now,
        )

        self.assertEqual(view["customer_name"], "김*수")
        self.assertEqual(view["business_no"], "-")
        self.assertEqual(view["phone"], "-")
        self.assertEqual(view["requested_by"], "-")
        self.assertEqual(view["overall_status"], "일부 수집 실패")
        self.assertEqual(view["downloadable_document_count"], 0)

    def test_results_excel_contains_only_accurate_masked_case_fields(
        self,
    ) -> None:
        case = {
            "company_name": "오아시스 세무",
            "representative_name_masked": "김*수",
            "business_no_masked": "120-**-***67",
            "phone_masked": "010-****-5678",
            "business_type": "individual",
            "hometax_status": "auth_complete",
            "comwel_status": "auth_complete",
            "overall_status": "ready",
            "requested_at": "2026-07-31T15:42:59+09:00",
            "requested_by": "임주형",
        }
        case_view = _claim_result_case_view(case)

        workbook_bytes = _build_claim_results_excel([case])
        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        worksheet = workbook["수집결과"]

        self.assertIn("CRM", worksheet["A1"].value)
        self.assertIn("서류조회", worksheet["A1"].value)
        self.assertEqual(
            [worksheet.cell(row=3, column=index).value for index in range(1, 11)],
            [
                "번호",
                "고객 / 상호",
                "구분",
                "사업자번호",
                "휴대전화",
                "홈택스",
                "근로복지공단",
                "수집 상태",
                "등록일",
                "담당자",
            ],
        )
        self.assertEqual(
            [worksheet.cell(row=4, column=index).value for index in range(1, 11)],
            [
                1,
                case_view["customer_name"],
                case_view["business_type"],
                case_view["business_no"],
                case_view["phone"],
                case_view["hometax_status"],
                case_view["comwel_status"],
                case_view["overall_status"],
                case_view["requested_at"],
                case_view["requested_by"],
            ],
        )
        cell_values = repr(
            [
                cell.value
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            ]
        )
        self.assertNotIn("1208800767", cell_values)
        self.assertNotIn("01012345678", cell_values)
        self.assertNotIn("결정세액", cell_values)
        self.assertNotIn("환급액", cell_values)
        self.assertEqual(worksheet.freeze_panes, "A4")
        self.assertEqual(worksheet.auto_filter.ref, "A3:J4")
        self.assertTrue(worksheet["A3"].font.bold)
        self.assertEqual(worksheet["A3"].fill.fill_type, "solid")
        self.assertFalse(worksheet.sheet_view.showGridLines)

    def test_results_tab_keeps_document_lookup_and_excel_download(self) -> None:
        source = inspect.getsource(_render_results_tab)

        self.assertIn('"서류조회"', source)
        self.assertIn('"Excel 다운로드"', source)
        self.assertIn("_build_claim_results_excel(filtered_cases)", source)
        self.assertNotIn("결정세액", source)
        self.assertNotIn("환급액", source)

    def test_variant_scope_and_download_name_never_expose_raw_number(
        self,
    ) -> None:
        document = self._document(
            "scoped",
            document_code="hometax_income_tax_return",
            facts={
                "collection_scope_label": (
                    "오아시스 본점 · 120-88-00767 <대표>"
                ),
                "download_file_name": (
                    "../2025_1208800767_종합소득세신고서.pdf"
                ),
            },
        )

        scope_label = _claim_document_scope_label(document)
        download_name = _claim_document_download_name(document)

        self.assertIn("120-**-***67", scope_label)
        self.assertNotIn("1208800767", scope_label)
        self.assertNotIn("120-88-00767", scope_label)
        self.assertNotIn("<", scope_label)
        self.assertNotIn("1208800767", download_name)
        self.assertNotIn("120-88-00767", download_name)
        self.assertNotIn("..", download_name)
        self.assertTrue(download_name.endswith(".pdf"))

    def test_download_cache_fingerprint_changes_for_replaced_document(
        self,
    ) -> None:
        original = self._document(
            "versioned",
            storage_path="owner/case/versioned-aaaaaaaaaaaaaaaa.pdf",
            content_sha256="a" * 64,
        )
        original_fingerprint = _claim_download_cache_fingerprint(
            [original]
        )

        for field, replacement in (
            ("id", "replacement-id"),
            (
                "storage_path",
                "owner/case/versioned-bbbbbbbbbbbbbbbb.pdf",
            ),
            ("content_sha256", "b" * 64),
            ("retention_until", "2099-01-01T00:00:00+00:00"),
        ):
            changed = dict(original, **{field: replacement})
            self.assertNotEqual(
                original_fingerprint,
                _claim_download_cache_fingerprint([changed]),
                field,
            )

        self.assertEqual(
            _claim_download_cache_fingerprint(
                [original, self._document("second")]
            ),
            _claim_download_cache_fingerprint(
                [self._document("second"), original]
            ),
        )

    def test_long_zip_download_name_preserves_extension(self) -> None:
        document = self._document(
            "long-name",
            document_code="hometax_income_tax_return",
            facts={
                "collection_scope_label": "가" * 100,
                "download_file_name": f"{'나' * 170}.pdf",
            },
        )

        download_name = _claim_document_download_name(document)

        self.assertLessEqual(len(download_name), 180)
        self.assertTrue(download_name.endswith(".pdf"))

    def test_zip_plan_assigns_105_documents_to_exactly_one_part(self):
        documents = [
            self._document(
                f"document-{index:03d}",
                size_bytes=1024,
                document_code="hometax_income_tax_return",
            )
            for index in range(105)
        ]

        parts = _plan_claim_document_zip_parts(documents)
        reversed_parts = _plan_claim_document_zip_parts(
            list(reversed(documents))
        )

        self.assertEqual([len(part) for part in parts], [75, 30])
        flattened_ids = [
            document["id"]
            for part in parts
            for document in part
        ]
        self.assertEqual(len(flattened_ids), 105)
        self.assertEqual(len(set(flattened_ids)), 105)
        self.assertEqual(
            flattened_ids,
            [
                document["id"]
                for part in reversed_parts
                for document in part
            ],
        )

    def test_zip_plan_splits_by_size_and_estimates_unknown_sizes(self):
        ten_megabytes = 10 * 1024 * 1024
        known_size_documents = [
            self._document(
                f"known-{index:02d}",
                size_bytes=ten_megabytes,
            )
            for index in range(7)
        ]
        unknown_size_documents = [
            self._document(f"unknown-{index:02d}")
            for index in range(13)
        ]

        self.assertEqual(
            [
                len(part)
                for part in _plan_claim_document_zip_parts(
                    known_size_documents
                )
            ],
            [6, 1],
        )
        self.assertEqual(
            [
                len(part)
                for part in _plan_claim_document_zip_parts(
                    unknown_size_documents
                )
            ],
            [12, 1],
        )

    @patch("requests.get")
    def test_zip_build_uses_one_batch_link_request(self, requests_get):
        documents = [
            self._document(
                "first-document",
                facts={"download_file_name": "first.pdf"},
            ),
            self._document(
                "second-document",
                facts={"download_file_name": "second.pdf"},
            ),
        ]
        repository = _BatchZipRepository()
        requests_get.side_effect = [
            _ZipResponse(b"first"),
            _ZipResponse(b"second"),
        ]

        archive = _build_claim_documents_zip(
            repository,
            "case-id",
            documents,
        )

        self.assertEqual(
            repository.batch_calls,
            [
                (
                    "case-id",
                    ["first-document", "second-document"],
                )
            ],
        )
        with ZipFile(BytesIO(archive)) as zip_file:
            self.assertEqual(
                set(zip_file.namelist()),
                {"first.pdf", "second.pdf"},
            )
            self.assertEqual(zip_file.read("first.pdf"), b"first")
            self.assertEqual(zip_file.read("second.pdf"), b"second")

    def test_collection_monitor_has_only_one_render_path(self) -> None:
        status_source = inspect.getsource(_render_status_tab)
        center_source = inspect.getsource(render_claim_correction_center)

        self.assertNotIn("_render_auto_claim_monitor(", status_source)
        self.assertEqual(
            center_source.count("_render_auto_claim_monitor("),
            1,
        )

    def test_manual_request_does_not_render_optional_business_fields(
        self,
    ) -> None:
        source = inspect.getsource(_render_personal_request)

        self.assertNotIn("상호명 (선택)", source)
        self.assertNotIn("사업자등록번호 (선택)", source)
        self.assertNotIn("claim_company_", source)
        self.assertNotIn("claim_business_no_", source)

    def test_active_monitor_never_prompts_for_one_business_selection(self):
        source = inspect.getsource(_render_auto_claim_monitor)

        self.assertNotIn('"사업자 선택"', source)
        self.assertNotIn('"사업장관리번호"', source)
        self.assertNotIn("_select_claim_business_number(", source)
        self.assertNotIn("_select_claim_management_number(", source)


if __name__ == "__main__":
    unittest.main()
