import requests
import pandas as pd
from datetime import datetime
import os
import re
import glob

from bizinfo_cache import load_bizinfo_programs_cached

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# 0. 기본 설정
# =========================================================

API_KEY = os.getenv("BIZINFO_API_KEY", "EScWvJ")

CUSTOMER_FILE = "고객DB.xlsx"
NOW = datetime.now().strftime("%Y%m%d_%H%M%S")
TODAY = datetime.today().strftime("%Y%m%d")

BIZINFO_URL = "https://www.bizinfo.go.kr/uss/rss/bizinfoApi.do"

# 기업마당 공고 수집 페이지 수. 너무 많으면 실행 시간이 길어질 수 있습니다.
API_PAGE_COUNT = 10

REGIONS = [
    "서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
    "경기", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주",
    "천안", "아산", "평택", "고양", "화성", "수원", "김해", "창원", "강남구", "유성구"
]

CITY_TO_PROVINCE = {
    "천안": "충남", "아산": "충남",
    "평택": "경기", "고양": "경기", "화성": "경기", "수원": "경기",
    "김해": "경남", "창원": "경남",
    "강남구": "서울", "유성구": "대전"
}

CERT_COLUMNS = ["벤처", "메인비즈", "이노비즈", "기업부설연구소", "여성기업", "장애인기업"]
RISK_COLUMNS = ["체납여부", "대표신용이슈"]

CRISIS_KEYWORDS = [
    "매출감소", "감소", "경영애로", "긴급", "위기", "피해", "회복",
    "재도약", "경영안정", "일시적경영애로", "폐업위기", "특별경영안정"
]

REGIONAL_POLICY_KEYWORDS = [
    "중소기업육성자금", "지역신보", "특례보증", "지자체", "테크노파크",
    "경제진흥원", "기업지원", "소상공인", "경영안정자금"
]

STOPWORDS = [
    "및", "또는", "그리고", "관련", "지원", "사업", "기업", "대상",
    "모집", "공고", "신청", "참여", "소상공인", "중소기업", "업체",
    "가능", "확인", "필요", "온라인", "오프라인"
]

LOGISTICS_KEYWORDS = [
    "화물", "화물운송", "운송", "물류", "택배", "운수", "운송업", "건설", "건설업"
]

COMMERCE_PLATFORM_KEYWORDS = [
    "유통플랫폼", "MD", "상담회", "온라인판로", "판로", "이커머스",
    "오픈마켓", "스마트스토어", "쇼핑몰", "라이브커머스", "입점",
    "마켓", "판매", "브랜드", "소비재", "식품", "상품"
]

TECH_BUSINESS_KEYWORDS = [
    "기술보증", "기보", "기술보증기금", "기술개발", "R&D", "연구개발",
    "기술혁신", "기술사업화", "특허", "시제품", "실증", "테스트베드",
    "AI", "인공지능", "ICT", "SW", "소프트웨어", "로봇", "스마트공장",
    "기술기업", "혁신형", "벤처", "이노비즈", "기업부설연구소"
]

TECH_CERT_COLUMNS = [
    "벤처", "이노비즈", "기업부설연구소", "특허보유", "R&D수행",
    "기술제품보유", "기술인력보유", "기술매출발생", "스마트공장도입", "기술보증희망"
]

INNOVATION_CERT_COLUMNS = [
    "벤처", "메인비즈", "이노비즈", "기업부설연구소"
]

MANUFACTURING_KEYWORDS = [
    "제조", "공장", "생산", "가공", "제품", "식품제조", "간편식"
]

MARKETING_KEYWORDS = [
    "온라인마케팅", "디지털마케팅", "퍼포먼스마케팅", "광고", "홍보",
    "브랜딩", "콘텐츠", "SNS", "인스타", "유튜브", "블로그",
    "온라인판로", "판로", "이커머스", "오픈마켓", "스마트스토어",
    "플랫폼", "라이브커머스", "쇼핑몰", "마케팅역량", "AI 활용"
]

SYNONYM_MAP = {
    "온라인마케팅": [
        "온라인마케팅", "디지털마케팅", "퍼포먼스마케팅", "광고", "홍보",
        "브랜딩", "콘텐츠", "SNS", "인스타", "유튜브", "블로그",
        "온라인판로", "판로", "이커머스", "오픈마켓", "스마트스토어",
        "플랫폼", "라이브커머스", "쇼핑몰", "마케팅역량"
    ],
    "디지털마케팅": [
        "온라인마케팅", "디지털마케팅", "퍼포먼스마케팅", "광고", "홍보",
        "브랜딩", "콘텐츠", "SNS", "온라인판로", "판로", "이커머스", "플랫폼"
    ],
    "마케팅": [
        "마케팅", "온라인마케팅", "디지털마케팅", "광고", "홍보",
        "브랜딩", "콘텐츠", "SNS", "온라인판로", "판로", "이커머스", "플랫폼"
    ],
    "광고": ["광고", "홍보", "브랜딩", "콘텐츠", "마케팅", "온라인마케팅"],
    "온라인판로": ["온라인판로", "판로", "이커머스", "오픈마켓", "스마트스토어", "플랫폼", "라이브커머스"],
    "이커머스": ["이커머스", "온라인판로", "오픈마켓", "스마트스토어", "플랫폼", "라이브커머스", "쇼핑몰"],
    "로봇제조": ["로봇", "제조", "자동화"],
    "로봇": ["로봇", "자동화", "스마트팩토리", "스마트공장"],
    "제조": ["제조", "생산", "공장"],
    "제조업": ["제조", "생산", "공장", "스마트공장"],
    "컨설팅": ["컨설팅", "자문", "멘토링", "컨설턴트"],
    "투자유치": ["투자유치", "IR", "피칭", "투자상담", "데모데이"],
    "IR": ["IR", "피칭", "투자유치", "투자상담"],
    "벤쳐": ["벤처", "스타트업", "창업"],
    "벤처": ["벤처", "스타트업", "창업"],
    "스타트업": ["스타트업", "창업", "벤처"],
    "스마트도시": ["스마트도시", "스마트시티", "도시", "솔루션"],
    "도시개발": ["도시개발", "스마트도시", "스마트시티", "도시"],
    "AI": ["AI", "인공지능", "데이터"],
    "인공지능": ["AI", "인공지능", "데이터"],
    "자동화": ["자동화", "스마트팩토리", "스마트공장"],
    "스마트공장": ["스마트공장", "스마트팩토리", "자동화"],
    "스마트팩토리": ["스마트공장", "스마트팩토리", "자동화"],
    "정책자금": ["정책자금", "자금", "융자", "대출", "보증"],
    "운전자금": ["운전자금", "경영안정", "운영자금", "자금"],
    "시설자금": ["시설자금", "설비", "설비투자", "시설", "기계"],
    "긴급경영안정": ["긴급", "경영애로", "경영안정", "매출감소", "위기"],
    "경영위기": ["경영위기", "경영애로", "매출감소", "긴급", "위기"],
    "사업화": ["사업화", "실증", "상용화", "시장진출"],
    "수출": ["수출", "해외", "글로벌", "수출바우처"],
    "고용지원금": ["고용", "채용", "장려금", "지원금"],
    "청년": ["청년", "청년채용", "청년고용"],
    "고령자": ["고령자", "60세", "계속고용"],
    "장애인": ["장애인", "장애인고용"],
    "육아휴직": ["육아휴직", "단축근무", "출산육아기"],
}


# =========================================================
# 1. 공통 유틸
# =========================================================

def safe_str(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def safe_num(value, default=0):
    try:
        if pd.isna(value) or value == "":
            return default
        return float(str(value).replace(",", "").replace("%", "").strip())
    except Exception:
        return default


def is_yes(value):
    return safe_str(value).upper() == "Y"


def split_keywords(text):
    text = safe_str(text)
    text = re.sub(r"[,/|·\n\r\t]", " ", text)
    return [word.strip() for word in text.split() if word.strip()]


def filter_meaningful_keywords(keywords):
    filtered = []

    for kw in keywords:
        kw = safe_str(kw)

        if not kw:
            continue

        if kw in STOPWORDS:
            continue

        if len(kw) <= 1:
            continue

        filtered.append(kw)

    return filtered


def expand_keywords(keywords):
    result = []
    for kw in keywords:
        kw = safe_str(kw)
        if not kw:
            continue

        result.append(kw)

        for key, values in SYNONYM_MAP.items():
            if key in kw or kw in key:
                result.extend(values)

    return list(dict.fromkeys(filter_meaningful_keywords([x for x in result if x])))


def row_text(row):
    return " ".join([safe_str(row[col]) for col in row.index if pd.notna(row[col])])


def contains_any(text, words):
    text = safe_str(text)
    return any(safe_str(w) and safe_str(w) in text for w in words)


def get_customer_topics(customer):
    topics = []
    for col in ["희망상담주제1", "희망상담주제2", "희망상담주제3"]:
        topics.extend(split_keywords(customer.get(col, "")))
    return expand_keywords(topics)


def get_customer_fund_needs(customer):
    needs = []
    for col in ["희망자금용도1", "희망자금용도2", "희망자금용도3"]:
        needs.extend(split_keywords(customer.get(col, "")))
    return expand_keywords(needs)


def get_revenue_drop_rate(customer):
    previous_sales = safe_num(customer.get("전년도연매출", 0))
    current_sales = safe_num(customer.get("올해예상매출", 0))

    if previous_sales > 0 and current_sales >= 0:
        return max(0, (previous_sales - current_sales) / previous_sales)

    # 엑셀에서 직접 계산된 매출감소율이 있으면 사용
    raw_rate = customer.get("매출감소율", 0)
    rate = safe_num(raw_rate, 0)
    if rate > 1:
        rate = rate / 100
    return max(0, rate)


def get_revenue_drop_label(customer):
    rate = get_revenue_drop_rate(customer)
    if rate >= 0.7:
        return "매출 70% 이상 급감"
    if rate >= 0.5:
        return "매출 50% 이상 감소"
    if rate >= 0.3:
        return "매출 30% 이상 감소"
    if rate > 0:
        return "매출 일부 감소"
    return "매출감소 정보 없음"


def is_revenue_declined(customer, threshold=0.3):
    return get_revenue_drop_rate(customer) >= threshold or is_yes(customer.get("매출감소여부", ""))


def make_customer_keywords(customer):
    keywords = []

    base_cols = [
        "시도", "시군구", "업종명", "표준산업분류코드", "주요 사업내용",
        "키워드메모", "비고"
    ]

    for col in base_cols:
        keywords.extend(split_keywords(customer.get(col, "")))

    keywords.extend(get_customer_topics(customer))
    keywords.extend(get_customer_fund_needs(customer))

    if is_revenue_declined(customer, 0.3):
        keywords += ["매출감소", "경영애로", "긴급", "경영안정", "위기", "회복"]

    for col in CERT_COLUMNS:
        if is_yes(customer.get(col, "")):
            keywords.append(col)

    if safe_num(customer.get("신규채용예정인원", 0)) > 0:
        keywords += ["신규채용", "채용", "고용"]

    if safe_num(customer.get("청년채용예정인원", 0)) > 0:
        keywords += ["청년", "청년채용", "고용"]

    if safe_num(customer.get("60세이상근로자수", 0)) > 0:
        keywords += ["고령자", "60세", "계속고용"]

    if safe_num(customer.get("장애인근로자수", 0)) > 0:
        keywords += ["장애인", "장애인고용"]

    if is_yes(customer.get("육아휴직/단축근무자 있음", "")):
        keywords += ["육아휴직", "단축근무", "출산육아기"]

    if is_yes(customer.get("설비투자계획", "")):
        keywords += ["설비투자", "시설자금", "기계"]

    if is_yes(customer.get("수출실적", "")):
        keywords += ["수출", "해외", "글로벌"]

    return expand_keywords(list(dict.fromkeys([kw for kw in keywords if kw])))


def get_primary_customer_text(customer):
    return " ".join([
        safe_str(customer.get("업종명", "")),
        safe_str(customer.get("키워드메모", "")),
        safe_str(customer.get("희망상담주제1", "")),
        safe_str(customer.get("희망상담주제2", "")),
        safe_str(customer.get("희망상담주제3", "")),
        safe_str(customer.get("희망자금용도1", "")),
        safe_str(customer.get("희망자금용도2", "")),
        safe_str(customer.get("희망자금용도3", ""))
    ])


def get_secondary_customer_text(customer):
    return " ".join([
        safe_str(customer.get("주요 사업내용", "")),
        safe_str(customer.get("비고", ""))
    ])


def detect_customer_profile(customer):
    primary = get_primary_customer_text(customer)
    secondary = get_secondary_customer_text(customer)
    full = primary + " " + secondary

    profile = {
        "logistics": contains_any(primary, LOGISTICS_KEYWORDS),
        "commerce": contains_any(primary, COMMERCE_PLATFORM_KEYWORDS),
        "manufacturing": contains_any(full, MANUFACTURING_KEYWORDS),
        "marketing": contains_any(full, MARKETING_KEYWORDS),
    }

    return profile


def detect_program_profile(program_text):
    return {
        "commerce": contains_any(program_text, COMMERCE_PLATFORM_KEYWORDS),
        "marketing": contains_any(program_text, MARKETING_KEYWORDS),
        "manufacturing": contains_any(program_text, MANUFACTURING_KEYWORDS),
        "crisis": contains_any(program_text, CRISIS_KEYWORDS),
    }


def industry_fit_adjustment(customer, program_text):
    profile = detect_customer_profile(customer)
    p_profile = detect_program_profile(program_text)

    score_delta = 0
    reasons = []
    checks = []

    # 핵심 업종이 화물운송/운송/건설인데, 공고가 유통플랫폼·MD·온라인판로 성격이면 강한 감점
    if profile["logistics"] and p_profile["commerce"]:
        if not profile["commerce"] and not profile["marketing"]:
            score_delta -= 35
            checks.append("핵심 업종이 운송/건설 계열인데 유통플랫폼·온라인판로 공고로 보여 업종 적합성 낮음")
        else:
            score_delta -= 12
            checks.append("운송/건설 성격과 판로지원 성격이 혼재되어 세부 사업모델 확인 필요")

    # 고객의 핵심 키워드에 마케팅/판로가 명확하고 공고도 판로/마케팅이면 가점
    if (profile["commerce"] or profile["marketing"]) and (p_profile["commerce"] or p_profile["marketing"]):
        score_delta += 25
        reasons.append("핵심 사업모델과 온라인마케팅/판로지원 공고 성격 일치")

    # 제조업/제품 기반 기업이 판로 공고와 맞는 경우
    if profile["manufacturing"] and p_profile["commerce"]:
        score_delta += 10
        reasons.append("제품/제조 기반 기업의 판로지원 가능성 반영")

    return score_delta, reasons, checks



def has_explicit_customer_marketing_or_commerce(customer):
    """
    마케팅/판로 가점은 희망자금용도의 '사업화' 확장어가 아니라,
    업종명/주요사업/키워드메모/비고에 직접 드러난 사업모델만 기준으로 판단합니다.
    """
    raw_business_text = " ".join([
        safe_str(customer.get("업종명", "")),
        safe_str(customer.get("주요 사업내용", "")),
        safe_str(customer.get("키워드메모", "")),
        safe_str(customer.get("비고", ""))
    ])

    explicit_words = [
        "마케팅", "온라인마케팅", "디지털마케팅", "광고", "홍보",
        "브랜딩", "콘텐츠", "SNS", "인스타", "유튜브", "블로그",
        "온라인판로", "판로", "이커머스", "오픈마켓", "스마트스토어",
        "플랫폼", "라이브커머스", "쇼핑몰", "MD", "유통플랫폼",
        "판매", "제조판매", "온라인 판매"
    ]

    return contains_any(raw_business_text, explicit_words)


def has_explicit_customer_product_or_manufacturing(customer):
    raw_business_text = " ".join([
        safe_str(customer.get("업종명", "")),
        safe_str(customer.get("주요 사업내용", "")),
        safe_str(customer.get("키워드메모", "")),
        safe_str(customer.get("비고", ""))
    ])

    explicit_words = [
        "제조", "생산", "가공", "식품", "제품", "소비재",
        "브랜드", "화장품", "의류", "생활용품", "건강식품",
        "간편식", "상품"
    ]

    return contains_any(raw_business_text, explicit_words)



def has_any_yes(customer, columns):
    for col in columns:
        if is_yes(customer.get(col, "")):
            return True
    return False


def has_tech_capability(customer):
    tech_text = " ".join([
        safe_str(customer.get("업종명", "")),
        safe_str(customer.get("주요 사업내용", "")),
        safe_str(customer.get("키워드메모", "")),
        safe_str(customer.get("비고", "")),
        safe_str(customer.get("기술성메모", ""))
    ])

    if has_any_yes(customer, TECH_CERT_COLUMNS):
        return True

    if contains_any(tech_text, [
        "특허", "R&D", "연구개발", "기술개발", "기술제품", "기술매출",
        "AI", "인공지능", "로봇", "소프트웨어", "플랫폼개발", "스마트공장",
        "기업부설연구소", "시제품", "실증"
    ]):
        return True

    return False


def has_innovation_cert(customer):
    return has_any_yes(customer, INNOVATION_CERT_COLUMNS)


def is_tech_program(program_text):
    return contains_any(program_text, TECH_BUSINESS_KEYWORDS)


def is_tech_guarantee_program(program_text):
    return contains_any(program_text, [
        "기술보증", "기술보증기금", "기보", "기술평가", "기술사업화보증"
    ])


def is_innovation_cert_preferred_program(program_text):
    return contains_any(program_text, [
        "벤처", "이노비즈", "메인비즈", "기업부설연구소", "혁신형기업", "인증기업"
    ])


def tech_fit_adjustment(customer, program_text):
    score_delta = 0
    reasons = []
    checks = []

    tech_program = is_tech_program(program_text)
    tech_guarantee = is_tech_guarantee_program(program_text)
    cert_preferred = is_innovation_cert_preferred_program(program_text)

    customer_has_tech = has_tech_capability(customer)
    customer_has_cert = has_innovation_cert(customer)

    # 기술보증/기보는 기술성 없으면 추천 제외에 가까운 강한 감점
    if tech_guarantee and not customer_has_tech:
        score_delta -= 60
        checks.append("기술보증 성격이나 특허/R&D/기술제품/기술인력 등 기술성 근거 부족")

    # R&D/기술개발/실증 공고도 기술성 없으면 강한 감점
    elif tech_program and not customer_has_tech:
        score_delta -= 35
        checks.append("기술/R&D/실증 성격 공고이나 기술성 근거 확인 필요")

    # 기술성이 있으면 가점
    if tech_program and customer_has_tech:
        score_delta += 25
        reasons.append("기술성 근거 보유로 기술/R&D/기보 계열 적합성 반영")

    # 벤처/이노비즈/메인비즈 우대 공고는 인증 없으면 우대가점 금지 + 확인사항
    if cert_preferred and not customer_has_cert:
        score_delta -= 15
        checks.append("벤처/이노비즈/메인비즈/기업부설연구소 인증 미보유로 우대조건 제한")

    if cert_preferred and customer_has_cert:
        score_delta += 12
        reasons.append("혁신형 인증 보유로 우대조건 반영")

    return score_delta, reasons, checks



def should_hard_exclude_by_business_model(customer, program_text):
    """
    핵심 업종이 운송/건설 계열인데 공고가 유통플랫폼·MD·온라인판로·디지털커머스 성격이고,
    고객DB에 실제 판매/마케팅/제조 사업모델이 명확하지 않으면 추천에서 제외합니다.
    """
    primary_text = " ".join([
        safe_str(customer.get("업종명", "")),
        safe_str(customer.get("키워드메모", "")),
        safe_str(customer.get("주요 사업내용", ""))
    ])

    logistics_core = contains_any(primary_text, LOGISTICS_KEYWORDS)
    commerce_program = contains_any(program_text, COMMERCE_PLATFORM_KEYWORDS) or contains_any(program_text, [
        "디지털커머스", "유통플랫폼", "MD 상담회", "온라인 판로", "온라인판로",
        "마케팅 역량", "플랫폼 입점", "라이브커머스"
    ])

    customer_has_relevant_model = (
        has_explicit_customer_marketing_or_commerce(customer)
        or has_explicit_customer_product_or_manufacturing(customer)
    )

    if logistics_core and commerce_program and not customer_has_relevant_model:
        return True

    return False


# =========================================================
# 2. 지역/마감/등급
# =========================================================

def check_region_type(sido, sigungu, target_region, program_text):
    sido = safe_str(sido)
    sigungu = safe_str(sigungu)
    target_region = safe_str(target_region)

    if target_region:
        if "전국" in target_region or "무관" in target_region:
            return True, "전국공고", 3

        if sigungu and sigungu in target_region:
            return True, "지역직접매칭", 10

        if sido and sido in target_region:
            return True, "광역지역매칭", 7

        if sigungu in CITY_TO_PROVINCE and CITY_TO_PROVINCE[sigungu] in target_region:
            return True, "광역지역매칭", 7

        return False, "타지역제외", -999

    if "전국" in program_text or "소재지 무관" in program_text or "전국 소재" in program_text:
        return True, "전국공고", 3

    if sigungu and sigungu in program_text:
        return True, "지역직접매칭", 10

    if sido and sido in program_text:
        return True, "광역지역매칭", 7

    if sigungu in CITY_TO_PROVINCE and CITY_TO_PROVINCE[sigungu] in program_text:
        return True, "광역지역매칭", 7

    for region in REGIONS:
        if region in program_text:
            if region not in [sido, sigungu, CITY_TO_PROVINCE.get(sigungu, "")]:
                return False, "타지역제외", -999

    return True, "지역언급없음", 0


def parse_deadline(period_text):
    text = safe_str(period_text)
    dates = re.findall(r"20\d{2}[-./]\d{1,2}[-./]\d{1,2}", text)

    if not dates:
        return None

    last_date = dates[-1].replace(".", "-").replace("/", "-")

    try:
        return pd.to_datetime(last_date)
    except Exception:
        return None


def deadline_status(period_text):
    deadline = parse_deadline(period_text)

    if deadline is None:
        return "마감일 확인 필요"

    today = pd.to_datetime(datetime.today().date())
    days = (deadline - today).days

    if days < 0:
        return "마감"

    if days <= 7:
        return f"마감임박 D-{days}"

    return f"접수중 D-{days}"


def grade_from_score(score):
    if score >= 80:
        return "A"
    if score >= 60:
        return "B"
    if score >= 40:
        return "C"
    return "D"


def grade_label(grade):
    return {
        "A": "적극 검토",
        "B": "가능성 있음",
        "C": "조건 확인 필요",
        "D": "참고 수준"
    }.get(grade, "참고 수준")


# =========================================================
# 3. 조건 판정
# =========================================================

def evaluate_range(value, min_val, max_val):
    value = safe_num(value)

    if safe_str(min_val) and value < safe_num(min_val):
        return False

    if safe_str(max_val) and value > safe_num(max_val):
        return False

    return True


def check_excluded(customer, exclude_text):
    target = " ".join([
        safe_str(customer.get("업종명", "")),
        safe_str(customer.get("주요 사업내용", "")),
        safe_str(customer.get("키워드메모", "")),
        safe_str(customer.get("비고", ""))
    ])

    for word in split_keywords(exclude_text):
        if word in target:
            return True

    return False


def check_common_risks(customer):
    risks = []

    if is_yes(customer.get("체납여부", "")):
        risks.append("체납 여부 확인 필요")

    if is_yes(customer.get("대표신용이슈", "")):
        risks.append("대표 신용이슈 확인 필요")

    if is_yes(customer.get("기존정책자금이용여부", "")):
        risks.append("기존 정책자금 이용 이력 확인 필요")

    return risks


def check_policy_numeric_conditions(customer, program):
    checks = []
    score = 0
    reasons = []

    rules = [
        ("업력", "최소업력", "최대업력", "업력 조건 충족", "업력 조건 확인 필요", 10),
        ("연매출", "최소매출", "최대매출", "매출 조건 충족", "매출 조건 확인 필요", 10),
        ("상시근로자수", "최소상시근로자수", "최대상시근로자수", "상시근로자 수 조건 충족", "상시근로자 수 조건 확인 필요", 10),
        ("현재고용보험가입인원", "최소고용보험가입인원", "", "고용보험 가입인원 조건 충족", "고용보험 가입인원 조건 확인 필요", 8),
        ("신규채용예정인원", "최소신규채용인원", "", "신규채용 조건 충족", "신규채용 조건 확인 필요", 8),
    ]

    for customer_col, min_col, max_col, ok_msg, fail_msg, point in rules:
        min_val = program.get(min_col, "")
        max_val = program.get(max_col, "")

        if not safe_str(min_val) and not safe_str(max_val):
            continue

        if evaluate_range(customer.get(customer_col, 0), min_val, max_val):
            score += point
            reasons.append(ok_msg)
        else:
            checks.append(fail_msg)

    if is_yes(program.get("청년채용필요", "")):
        if safe_num(customer.get("청년채용예정인원", 0)) > 0:
            score += 10
            reasons.append("청년채용 조건 충족")
        else:
            checks.append("청년채용 조건 확인 필요")

    if is_yes(program.get("매출감소필요", "")):
        required_drop = safe_num(program.get("최소매출감소율", 30)) / 100
        actual_drop = get_revenue_drop_rate(customer)
        if actual_drop >= required_drop:
            score += 30
            reasons.append(f"매출감소 조건 충족: {actual_drop:.1%} 감소")
        else:
            checks.append(f"매출감소 조건 확인 필요: 최소 {required_drop:.0%}")

    return score, reasons, checks


def check_employment_conditions(customer, emp):
    score = 0
    reasons = []
    checks = []
    hard_fails = []

    min_insured = safe_num(emp.get("최소고용보험가입인원", 0))
    if min_insured > 0:
        if safe_num(customer.get("현재고용보험가입인원", 0)) >= min_insured:
            score += 12
            reasons.append("고용보험 가입인원 조건 충족")
        else:
            hard_fails.append(f"고용보험 가입인원 {int(min_insured)}명 이상 필요")

    min_employee = safe_num(emp.get("최소상시근로자수", 0))
    if min_employee > 0:
        if safe_num(customer.get("상시근로자수", 0)) >= min_employee:
            score += 10
            reasons.append("상시근로자 수 조건 충족")
        else:
            hard_fails.append(f"상시근로자 {int(min_employee)}명 이상 필요")

    if is_yes(emp.get("신규채용필요", "")):
        min_new = max(1, safe_num(emp.get("최소신규채용인원", 1)))
        if safe_num(customer.get("신규채용예정인원", 0)) >= min_new:
            score += 15
            reasons.append("신규채용 조건 충족")
        else:
            hard_fails.append(f"신규채용 예정인원 {int(min_new)}명 이상 필요")

    min_young = safe_num(emp.get("청년채용최소", 0))
    if min_young > 0:
        if safe_num(customer.get("청년채용예정인원", 0)) >= min_young:
            score += 20
            reasons.append("청년채용 인원 조건 충족")
        else:
            hard_fails.append(f"청년채용 예정인원 {int(min_young)}명 이상 필요")

    min_senior = safe_num(emp.get("60세이상최소", 0))
    if min_senior > 0:
        if safe_num(customer.get("60세이상근로자수", 0)) >= min_senior:
            score += 20
            reasons.append("60세 이상 근로자 조건 충족")
        else:
            hard_fails.append(f"60세 이상 근로자 {int(min_senior)}명 이상 필요")

    min_disabled = safe_num(emp.get("장애인근로자최소", 0))
    if min_disabled > 0:
        if safe_num(customer.get("장애인근로자수", 0)) >= min_disabled:
            score += 20
            reasons.append("장애인 근로자 조건 충족")
        else:
            hard_fails.append(f"장애인 근로자 {int(min_disabled)}명 이상 필요")

    if is_yes(emp.get("육아휴직/단축필요", "")):
        if is_yes(customer.get("육아휴직/단축근무자 있음", "")):
            score += 20
            reasons.append("육아휴직/단축근무 조건 충족")
        else:
            hard_fails.append("육아휴직/단축근무자 필요")

    if is_yes(emp.get("주30시간이상필요", "")):
        if is_yes(customer.get("주30시간이상근로", "")):
            score += 8
            reasons.append("주30시간 이상 조건 충족")
        else:
            checks.append("주30시간 이상 근로 여부 확인 필요")

    if is_yes(emp.get("정규직필요", "")):
        if is_yes(customer.get("정규직채용예정", "")):
            score += 8
            reasons.append("정규직 채용 조건 충족")
        else:
            checks.append("정규직 채용 여부 확인 필요")

    if is_yes(emp.get("최저임금이상필요", "")):
        if is_yes(customer.get("최저임금이상", "")):
            score += 8
            reasons.append("최저임금 이상 조건 충족")
        else:
            checks.append("최저임금 이상 지급 여부 확인 필요")

    min_months = safe_num(emp.get("고용유지기간개월", 0))
    if min_months > 0:
        if safe_num(customer.get("고용유지가능개월", 0)) >= min_months:
            score += 8
            reasons.append("고용유지기간 조건 충족")
        else:
            checks.append(f"고용유지 {int(min_months)}개월 가능 여부 확인 필요")

    return score, reasons, checks, hard_fails


# =========================================================
# 4. 상담코멘트/신청가능성
# =========================================================

def success_probability(score, customer, checks):
    probability = min(95, max(10, int(score)))

    if is_revenue_declined(customer, 0.5):
        probability += 8

    if check_common_risks(customer):
        probability -= 20

    if checks:
        probability -= min(15, len(checks) * 3)

    return max(5, min(95, probability))


def make_consulting_comment(row):
    parts = []

    if safe_str(row.get("사업구분", "")) == "공고형":
        parts.append("기업마당 공고형 사업 기준으로 추천되었습니다.")

    if "매출감소" in safe_str(row.get("매칭키워드", "")) or "매출감소" in safe_str(row.get("추천사유", "")):
        parts.append("최근 매출 감소 흐름이 반영되어 경영안정·긴급자금 계열 검토 우선순위가 높습니다.")

    if safe_str(row.get("신청판정", "")) == "신청검토":
        parts.append("현재 입력값 기준으로는 상담 우선순위가 높은 편입니다.")
    elif safe_str(row.get("신청판정", "")) == "조건확인":
        parts.append("일부 세부요건 확인 후 신청 가능성을 판단하는 것이 좋습니다.")
    else:
        parts.append("현재 입력값 기준으로는 보류 또는 추가 확인이 필요합니다.")

    if safe_str(row.get("확인필요사항", "")):
        parts.append(f"우선 확인할 항목은 {row.get('확인필요사항')}입니다.")

    return " ".join(parts)


def get_customer_exclusion_keywords(customer):
    values = []

    direct = safe_str(customer.get("제외키워드", ""))
    if direct:
        values += split_keywords(direct)

    memo = safe_str(customer.get("비고", ""))
    match = re.search(
        r"제외키워드\s*[:：]\s*(.+?)(?:/|투자예정금액|투자예정시기|$)",
        memo,
    )
    if match:
        values += split_keywords(match.group(1))

    return list(dict.fromkeys(
        keyword.strip()
        for keyword in values
        if keyword and keyword.strip()
    ))


def should_exclude_by_user_keywords(customer, program_text):
    for keyword in get_customer_exclusion_keywords(customer):
        if keyword and keyword in safe_str(program_text):
            return True
    return False


# =========================================================
# 5. 점수 산정
# =========================================================

def score_api_program(customer, program):
    text = row_text(program)
    period = safe_str(program.get("reqstBeginEndDe", ""))
    status = deadline_status(period)

    if status == "마감":
        return None

    ok, region_type, region_score = check_region_type(
        customer.get("시도", ""),
        customer.get("시군구", ""),
        "",
        text
    )

    if not ok:
        return None

    # 사용자가 지정한 제외키워드가 공고문에 포함되면 추천에서 제외
    if should_exclude_by_user_keywords(customer, text):
        return None

    # 업종-공고 성격이 명확히 어긋나는 경우 추천 제외
    if should_hard_exclude_by_business_model(customer, text):
        return None

    score = max(region_score, 0)
    reasons = [f"지역: {region_type}"]
    checks = ["공고문 세부 조건 확인"]
    matched = []

    keywords = make_customer_keywords(customer)
    topic_need_keywords = get_customer_topics(customer) + get_customer_fund_needs(customer)

    for keyword in keywords:
        if keyword and keyword in text:
            point = 7 if keyword in topic_need_keywords else 2
            score += point
            matched.append(keyword)

    # 온라인마케팅/판로지원 공고 가중치
    # 희망자금용도 '사업화' 때문에 생기는 과매칭을 막기 위해
    # 실제 사업내용/키워드메모/비고에 마케팅·판로·판매·제조 성격이 있을 때만 가점
    customer_has_marketing_model = has_explicit_customer_marketing_or_commerce(customer)
    customer_has_product_model = has_explicit_customer_product_or_manufacturing(customer)
    program_marketing_count = sum(1 for kw in MARKETING_KEYWORDS if kw in text)
    program_commerce_count = sum(1 for kw in COMMERCE_PLATFORM_KEYWORDS if kw in text)

    if customer_has_marketing_model and (program_marketing_count >= 2 or program_commerce_count >= 2):
        score += 28
        matched.append("온라인마케팅")
        reasons.append("온라인마케팅/판로지원 특화 매칭")

    elif customer_has_product_model and program_commerce_count >= 2:
        score += 14
        matched.append("제품판로")
        reasons.append("제품/제조 기반 판로지원 가능성 반영")

    # 매출감소/경영위기 공고 가중치
    if is_revenue_declined(customer, 0.3) and contains_any(text, CRISIS_KEYWORDS):
        drop_rate = get_revenue_drop_rate(customer)
        if drop_rate >= 0.7:
            score += 35
        elif drop_rate >= 0.5:
            score += 28
        else:
            score += 20
        matched.append("매출감소")
        reasons.append(f"{get_revenue_drop_label(customer)}로 경영애로/긴급성 가점")

    # 지역형 정책자금 키워드
    if contains_any(text, REGIONAL_POLICY_KEYWORDS):
        score += 6
        reasons.append("지역·지자체성 정책자금 키워드 반영")

    for risk in check_common_risks(customer):
        score -= 20
        checks.append(risk)

    if score <= 0:
        return None

    grade = grade_from_score(score)
    probability = success_probability(score, customer, checks)

    return {
        "추천유형": region_type,
        "매칭점수": score,
        "추천등급": grade,
        "등급설명": grade_label(grade),
        "신청가능성점수": probability,
        "매칭키워드": ", ".join(list(dict.fromkeys(matched))),
        "추천사유": " / ".join(list(dict.fromkeys(reasons))),
        "확인필요사항": " / ".join(list(dict.fromkeys(checks))),
        "마감상태": status,
        "신청가능성": "조건확인"
    }


def score_policy_program(customer, program):
    text = row_text(program)

    ok, region_type, region_score = check_region_type(
        customer.get("시도", ""),
        customer.get("시군구", ""),
        program.get("대상지역", ""),
        text
    )

    if not ok:
        return None

    if check_excluded(customer, program.get("제외업종", "")):
        return None

    if should_exclude_by_user_keywords(customer, text):
        return None

    score = max(region_score, 0)
    reasons = [f"지역: {region_type}"]
    checks = []
    matched = []

    customer_keywords = make_customer_keywords(customer)
    program_keywords = row_text(program)

    for kw in customer_keywords:
        if kw and kw in program_keywords:
            score += 3
            matched.append(kw)

    for kw in get_customer_topics(customer) + get_customer_fund_needs(customer):
        if kw and kw in program_keywords:
            score += 8
            reasons.append(f"핵심 니즈 일치: {kw}")
            matched.append(kw)

    industry = safe_str(customer.get("업종명", ""))
    if industry and industry in safe_str(program.get("대상업종", "")) + safe_str(program.get("추천키워드", "")):
        score += 15
        reasons.append("업종 조건 일치")
        matched.append(industry)

    # 온라인마케팅/판로지원 DB 가중치
    customer_has_marketing_model = has_explicit_customer_marketing_or_commerce(customer)
    customer_has_product_model = has_explicit_customer_product_or_manufacturing(customer)
    program_marketing_count = sum(1 for kw in MARKETING_KEYWORDS if kw in program_keywords)
    program_commerce_count = sum(1 for kw in COMMERCE_PLATFORM_KEYWORDS if kw in program_keywords)

    if customer_has_marketing_model and (program_marketing_count >= 2 or program_commerce_count >= 2):
        score += 25
        reasons.append("온라인마케팅/판로지원 특화 매칭")
        matched.append("온라인마케팅")
    elif customer_has_product_model and program_commerce_count >= 2:
        score += 12
        reasons.append("제품/제조 기반 판로지원 가능성 반영")
        matched.append("제품판로")

    numeric_score, numeric_reasons, numeric_checks = check_policy_numeric_conditions(customer, program)
    score += numeric_score
    reasons += numeric_reasons
    checks += numeric_checks

    preferred_text = f"{safe_str(program.get('우대조건',''))} {safe_str(program.get('추천키워드',''))} {text}"
    for col in CERT_COLUMNS:
        if is_yes(customer.get(col, "")) and col in preferred_text:
            score += 5
            reasons.append(f"우대조건: {col}")
            matched.append(col)

    for risk in check_common_risks(customer):
        score -= 20
        checks.append(risk)

    if score <= 0:
        return None

    grade = grade_from_score(score)
    probability = success_probability(score, customer, checks)
    possibility = "신청검토" if grade in ["A", "B"] and not any("확인 필요" in c for c in checks) else "조건확인"

    return {
        "추천유형": region_type,
        "매칭점수": score,
        "추천등급": grade,
        "등급설명": grade_label(grade),
        "신청가능성점수": probability,
        "매칭키워드": ", ".join(list(dict.fromkeys(matched))),
        "추천사유": " / ".join(list(dict.fromkeys(reasons))),
        "확인필요사항": " / ".join(list(dict.fromkeys(checks))) if checks else "세부 공고 및 증빙서류 확인",
        "마감상태": "상시/제도별 확인",
        "신청가능성": possibility
    }


def score_employment_program(customer, emp):
    text = row_text(emp)

    ok, region_type, region_score = check_region_type(
        customer.get("시도", ""),
        customer.get("시군구", ""),
        emp.get("대상지역", ""),
        text
    )

    if not ok:
        return None

    if check_excluded(customer, emp.get("제외업종", "")):
        return None

    if should_exclude_by_user_keywords(customer, text):
        return None

    score = max(region_score, 0)
    reasons = [f"지역: {region_type}"]
    checks = []
    matched = []

    condition_score, condition_reasons, condition_checks, hard_fails = check_employment_conditions(customer, emp)

    if hard_fails:
        checks += hard_fails + condition_checks
        grade = "D"
        return {
            "추천유형": "고용지원",
            "매칭점수": score,
            "추천등급": grade,
            "등급설명": grade_label(grade),
            "신청가능성점수": 10,
            "매칭키워드": "",
            "추천사유": " / ".join(reasons),
            "확인필요사항": " / ".join(checks),
            "마감상태": "상시/제도별 확인",
            "신청가능성": "보류"
        }

    score += condition_score
    reasons += condition_reasons
    checks += condition_checks

    customer_keywords = make_customer_keywords(customer)
    for kw in customer_keywords:
        if kw and kw in text:
            score += 2
            matched.append(kw)

    for kw in get_customer_topics(customer):
        if kw and kw in text:
            score += 5
            matched.append(kw)

    for risk in check_common_risks(customer):
        score -= 20
        checks.append(risk)

    if score <= 0:
        return None

    grade = grade_from_score(score)
    probability = success_probability(score, customer, checks)
    possibility = "신청검토" if grade in ["A", "B"] and not hard_fails else "조건확인"

    return {
        "추천유형": "고용지원",
        "매칭점수": score,
        "추천등급": grade,
        "등급설명": grade_label(grade),
        "신청가능성점수": probability,
        "매칭키워드": ", ".join(list(dict.fromkeys(matched))),
        "추천사유": " / ".join(list(dict.fromkeys(reasons))),
        "확인필요사항": " / ".join(list(dict.fromkeys(checks))) if checks else "세부 증빙서류 확인",
        "마감상태": "상시/제도별 확인",
        "신청가능성": possibility
    }


# =========================================================
# 6. 결과/문자/엑셀
# =========================================================

def consultation_checklist(row):
    if row["사업구분"] == "고용지원금":
        return "고용보험 가입자 명부, 근로계약서, 임금대장, 신규채용자 정보, 근로시간, 최저임금, 고용유지기간 확인"
    if row["사업구분"] == "상시정책자금":
        return "사업자등록증, 최근 재무제표, 부가세과표, 국세/지방세 완납, 기존 차입금, 대표 신용이슈, 매출감소 증빙 확인"
    return "공고문, 신청자격, 마감일, 사업계획서, 필요서류, 자부담 여부, 매출감소 증빙 확인"


def make_sms(row):
    return (
        f"{row['대표자명']}님, 안녕하세요.\n"
        f"이번 달 기준으로 {row['업체명']}에 검토 가능한 지원사업을 확인해봤습니다.\n\n"
        f"[{row['사업구분']}] {row['추천사업명']}\n"
        f"- 추천등급: {row['추천등급']} ({row['등급설명']})\n"
        f"- 신청가능성: {row['신청가능성']} / {row['신청가능성점수']}점\n"
        f"- 추천사유: {row['추천사유']}\n"
        f"- 확인필요: {row['확인필요사항']}\n\n"
        f"세부 요건은 자료 확인 후 최종 판단이 필요합니다.\n"
        f"대표님께 도움 될 만한 내용만 추려서 다시 안내드리겠습니다.\n"
        f"편하실 때 확인해 주세요. 오늘도 좋은 하루 보내세요 :)"
    )


def make_decision(row):
    if row.get("신청가능성", "") == "보류":
        return "보류"
    if row["추천등급"] in ["A", "B"] and "체납" not in row["확인필요사항"] and "신용" not in row["확인필요사항"]:
        return "신청검토"
    if row["추천등급"] == "C":
        return "조건확인"
    return "보류"


def decision_reason(row):
    if row["신청판정"] == "신청검토":
        return "추천등급과 세부조건이 양호하여 상담 우선순위 높음"
    if row["신청판정"] == "조건확인":
        return "일부 조건 확인 후 신청 가능성 판단 필요"
    return "점수 낮음 또는 필수 조건 미충족/리스크 항목 확인 필요"


def make_result_id(row):
    return f"{row['업체명']}|{row['사업구분']}|{row['추천사업명']}"


def find_previous_result_file(current_file):
    candidates = []
    for pattern in ["매칭결과_기술성조건반영_*.xlsx", "매칭결과_세부조건형_*.xlsx", "매칭결과_고도화실무형_*.xlsx", "매칭결과_실무형_*.xlsx"]:
        candidates.extend(glob.glob(pattern))
    candidates = [f for f in candidates if f != current_file]
    if not candidates:
        return None
    candidates.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return candidates[0]


def add_new_flag(df, current_file):
    df["추천ID"] = df.apply(make_result_id, axis=1)
    previous_file = find_previous_result_file(current_file)

    if not previous_file:
        df["신규여부"] = "신규"
        df["비교기준파일"] = "이전 결과 없음"
        return df

    try:
        prev_df = pd.read_excel(previous_file, sheet_name="전체결과")
        if "추천ID" not in prev_df.columns:
            prev_df["추천ID"] = prev_df.apply(make_result_id, axis=1)
        prev_ids = set(prev_df["추천ID"].astype(str).tolist())
        df["신규여부"] = df["추천ID"].apply(lambda x: "신규" if str(x) not in prev_ids else "기존")
        df["비교기준파일"] = previous_file
        return df
    except Exception:
        df["신규여부"] = "확인필요"
        df["비교기준파일"] = f"{previous_file} 읽기 실패"
        return df


def check_input_errors(customer_df):
    required_cols = [
        "업체명", "대표자명", "대표 연락처", "시도", "시군구", "업종명",
        "업력", "전년도연매출", "올해예상매출", "상시근로자수", "현재고용보험가입인원"
    ]
    yn_cols = [
        "매출감소여부", "정규직채용예정", "주30시간이상근로", "최저임금이상",
        "체납여부", "대표신용이슈", "기존정책자금이용여부",
        "벤처", "메인비즈", "이노비즈", "기업부설연구소", "여성기업", "장애인기업",
        "육아휴직/단축근무자 있음", "설비투자계획", "수출실적",
        "특허보유", "상표권보유", "R&D수행", "기술제품보유", "기술인력보유",
        "기술매출발생", "스마트공장도입", "기술보증희망"
    ]
    numeric_cols = [
        "업력", "전년도연매출", "올해예상매출", "연매출", "영업이익",
        "상시근로자수", "직전년도평균고용인원",
        "현재고용보험가입인원", "신규채용예정인원", "청년채용예정인원",
        "60세이상근로자수", "장애인근로자수", "고용유지가능개월"
    ]

    errors = []

    for idx, row in customer_df.iterrows():
        company = safe_str(row.get("업체명", f"{idx+2}행"))

        for col in required_cols:
            if col in customer_df.columns and not safe_str(row.get(col, "")):
                errors.append({"행번호": idx + 2, "업체명": company, "오류항목": col, "오류내용": "필수값 누락"})

        for col in yn_cols:
            if col in customer_df.columns:
                val = safe_str(row.get(col, "")).upper()
                if val and val not in ["Y", "N"]:
                    errors.append({"행번호": idx + 2, "업체명": company, "오류항목": col, "오류내용": "Y 또는 N으로 입력 필요"})

        for col in numeric_cols:
            if col in customer_df.columns:
                val = row.get(col, "")
                if safe_str(val) and safe_num(val, None) is None:
                    errors.append({"행번호": idx + 2, "업체명": company, "오류항목": col, "오류내용": "숫자 형식 확인 필요"})

    if not errors:
        return pd.DataFrame([{"행번호": "", "업체명": "", "오류항목": "", "오류내용": "입력 오류 없음"}])

    return pd.DataFrame(errors)


def load_bizinfo_programs():
    """
    v3.5.0: 매칭 시 기업마당을 매번 실시간 호출하지 않고
    data/bizinfo_programs.json 내부 DB를 우선 사용합니다.
    내부 DB가 최초 한 번도 없는 경우에만 실시간 동기화를 한 번 시도합니다.
    """
    return load_bizinfo_programs_cached(
        api_key=API_KEY,
        page_count=API_PAGE_COUNT,
        allow_live_fallback=True,
        data_dir="data",
    )

def style_ws(ws, color="1F4E78"):
    header_fill = PatternFill("solid", fgColor=color)
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        header = col[0].value
        max_len = max([len(str(c.value)) if c.value else 0 for c in col])

        if header == "문자문안":
            ws.column_dimensions[letter].width = 35
        elif header == "상담체크리스트":
            ws.column_dimensions[letter].width = 34
        elif header in ["추천사유", "확인필요사항", "신청판정사유", "상담코멘트", "핵심 확인사항", "상담 질문"]:
            ws.column_dimensions[letter].width = 32
        else:
            ws.column_dimensions[letter].width = min(max_len + 4, 24)

        for cell in col:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 45


def save_pretty_excel(df, result_file, input_error_df, excluded_df=None):
    with pd.ExcelWriter(result_file, engine="openpyxl") as writer:

        # 1. 입력오류체크 시트만 공통 시트로 유지
        input_error_df.to_excel(writer, sheet_name="입력오류체크", index=False)

        # 2. 업체별 시트 생성
        for company in df["업체명"].unique():
            company_df = (
                df[df["업체명"] == company]
                .sort_values(
                    ["신청가능성점수", "매칭점수"],
                    ascending=[False, False]
                )
            )

            sheet_name = f"업체별_{company}"[:31]
            start_row = 0

            # 2-1. 업체별 TOP3
            top3 = company_df.head(3)

            pd.DataFrame([[f"■ {company} 우선검토 TOP3"]]).to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                index=False,
                header=False
            )
            start_row += 1

            top3_cols = [
                "사업구분", "추천사업명", "기관명", "추천등급", "신청가능성점수",
                "매칭점수", "신청판정", "추천사유", "확인필요사항", "상담코멘트",
                "필요서류", "출처URL"
            ]

            top3[[col for col in top3_cols if col in top3.columns]].to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                index=False
            )

            start_row += len(top3) + 4

            # 2-2. 카테고리별 상세
            for category in ["상시정책자금", "공고형", "고용지원금"]:
                category_df = (
                    company_df[company_df["사업구분"] == category]
                    .sort_values(
                        ["신청가능성점수", "매칭점수"],
                        ascending=[False, False]
                    )
                )

                if category_df.empty:
                    continue

                pd.DataFrame([[f"■ {category}"]]).to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=start_row,
                    index=False,
                    header=False
                )

                start_row += 1

                detail_cols = [
                    "사업구분", "추천사업명", "기관명", "추천등급", "등급설명",
                    "신청가능성점수", "매칭점수", "신청가능성", "신청판정",
                    "마감상태", "추천사유", "확인필요사항", "상담코멘트",
                    "상담체크리스트", "매칭키워드", "필요서류", "신청기간",
                    "출처URL", "문자문안"
                ]

                category_df[[col for col in detail_cols if col in category_df.columns]].to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=start_row,
                    index=False
                )

                start_row += len(category_df) + 4

            # 2-3. 제외/보류 사업
            if excluded_df is not None and not excluded_df.empty:
                ex_company_df = excluded_df[excluded_df["업체명"] == company]

                if not ex_company_df.empty:
                    pd.DataFrame([[f"■ 제외/보류 참고사업"]]).to_excel(
                        writer,
                        sheet_name=sheet_name,
                        startrow=start_row,
                        index=False,
                        header=False
                    )

                    start_row += 1

                    ex_cols = ["사업구분", "지원사업명", "제외사유", "확인포인트"]
                    ex_company_df[[col for col in ex_cols if col in ex_company_df.columns]].to_excel(
                        writer,
                        sheet_name=sheet_name,
                        startrow=start_row,
                        index=False
                    )

                    start_row += len(ex_company_df) + 4

            # 2-4. 상담질문
            consultation_questions = pd.DataFrame([
                ["상담질문", "전년도 매출과 올해 예상매출 증빙자료가 있으신가요?"],
                ["상담질문", "국세·지방세 체납은 없으신가요?"],
                ["상담질문", "기존 정책자금 또는 보증 이용 내역이 있으신가요?"],
                ["상담질문", "현재 고용보험 가입인원과 신규채용 계획은 어떻게 되시나요?"],
                ["상담질문", "온라인마케팅·판로·광고·브랜딩 관련 지원이 필요하신가요?"],
                ["상담질문", "특허, R&D 수행, 기술제품, 기업부설연구소 등 기술성 근거가 있으신가요?"],
            ], columns=["구분", "내용"])

            pd.DataFrame([[f"■ 상담질문"]]).to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                index=False,
                header=False
            )

            start_row += 1

            consultation_questions.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=start_row,
                index=False
            )

        workbook = writer.book

        category_colors = {
            "우선검토 TOP3": "C6EFCE",
            "상시정책자금": "D9EAD3",
            "공고형": "D9EAF7",
            "고용지원금": "FCE4D6",
            "제외/보류 참고사업": "F4CCCC",
            "상담질문": "FFF2CC"
        }

        cell_colors = {
            "A": "C6EFCE",
            "B": "D9EAD3",
            "C": "FFF2CC",
            "D": "F4CCCC",
            "신청검토": "C6EFCE",
            "조건확인": "FFF2CC",
            "보류": "F4CCCC",
            "신규": "BDD7EE",
            "기존": "E7E6E6"
        }

        for ws in workbook.worksheets:
            if ws.title == "입력오류체크":
                style_ws(ws)
                continue

            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("■"):
                        title = cell.value.replace("■", "").strip()
                        fill_color = "E2F0D9"

                        for key, color in category_colors.items():
                            if key in title:
                                fill_color = color
                                break

                        cell.font = Font(bold=True, size=14)
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

                    elif cell.value in cell_colors:
                        cell.fill = PatternFill("solid", fgColor=cell_colors[cell.value])
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

            # 헤더 스타일
            for row in ws.iter_rows():
                values = [cell.value for cell in row]
                if any(v in values for v in ["사업구분", "추천사업명", "지원사업명", "구분"]):
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="1F4E78")
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col in ws.columns:
                letter = get_column_letter(col[0].column)
                header_values = [safe_str(c.value) for c in col[:5]]
                header_text = " ".join(header_values)
                max_len = max([len(str(c.value)) if c.value else 0 for c in col])

                if "문자문안" in header_text:
                    ws.column_dimensions[letter].width = 35
                elif "상담체크리스트" in header_text:
                    ws.column_dimensions[letter].width = 34
                elif any(key in header_text for key in ["추천사유", "확인필요사항", "상담코멘트", "확인포인트", "내용"]):
                    ws.column_dimensions[letter].width = 34
                else:
                    ws.column_dimensions[letter].width = min(max_len + 4, 24)

            for row in ws.iter_rows():
                ws.row_dimensions[row[0].row].height = 45

            ws.freeze_panes = "A2"


# =========================================================
# 7. 실행
# =========================================================

def build_excluded_df(customer_df, policy_df, employment_df, result_df):
    excluded_rows = []

    matched_keys = set()
    if result_df is not None and not result_df.empty:
        for _, row in result_df.iterrows():
            matched_keys.add(
                (
                    safe_str(row.get("업체명", "")),
                    safe_str(row.get("사업구분", "")),
                    safe_str(row.get("추천사업명", ""))
                )
            )

    for _, customer in customer_df.iterrows():
        company_name = safe_str(customer.get("업체명", ""))

        if not company_name:
            continue

        # 상시정책자금 제외사유
        for _, policy in policy_df.iterrows():
            program_name = safe_str(policy.get("상품명", ""))

            if not program_name:
                continue

            key = (company_name, "상시정책자금", program_name)

            if key not in matched_keys:
                excluded_rows.append({
                    "업체명": company_name,
                    "사업구분": "상시정책자금",
                    "지원사업명": program_name,
                    "제외사유": "지역/업종/매출/고용/우대조건 미충족 또는 점수 부족 가능",
                    "확인포인트": "대상업종, 대상지역, 매출감소율, 상시근로자수, 고용보험 가입인원 확인"
                })

        # 고용지원금 제외사유
        for _, emp in employment_df.iterrows():
            program_name = safe_str(emp.get("제도명", ""))

            if not program_name:
                continue

            key = (company_name, "고용지원금", program_name)

            if key not in matched_keys:
                excluded_rows.append({
                    "업체명": company_name,
                    "사업구분": "고용지원금",
                    "지원사업명": program_name,
                    "제외사유": "고용보험/신규채용/청년·고령자·장애인/근로조건 미충족 가능",
                    "확인포인트": "고용보험 가입인원, 신규채용 예정인원, 청년채용, 정규직, 주30시간, 최저임금 확인"
                })

    if not excluded_rows:
        return pd.DataFrame([{
            "업체명": "",
            "사업구분": "",
            "지원사업명": "",
            "제외사유": "제외 항목 없음",
            "확인포인트": ""
        }])

    return pd.DataFrame(excluded_rows)


def main():
    if not os.path.exists(CUSTOMER_FILE):
        print("고객DB.xlsx 파일이 없습니다.")
        return

    customer_df = pd.read_excel(CUSTOMER_FILE, sheet_name="고객DB")

    # 빈 행 자동 제거: 업체명이 없는 행은 검사/매칭에서 제외
    customer_df = customer_df[
        customer_df["업체명"].notna() &
        (customer_df["업체명"].astype(str).str.strip() != "")
    ]

    policy_df = pd.read_excel(CUSTOMER_FILE, sheet_name="상시정책자금DB")
    employment_df = pd.read_excel(CUSTOMER_FILE, sheet_name="고용지원금DB")
    api_df = load_bizinfo_programs()

    input_error_df = check_input_errors(customer_df)

    results = []
    excluded_results = []

    for _, customer in customer_df.iterrows():
        company_name = safe_str(customer.get("업체명", ""))

        if not company_name:
            continue

        common = {
            "업체명": company_name,
            "대표자명": safe_str(customer.get("대표자명", "")),
            "대표 연락처": safe_str(customer.get("대표 연락처", "")),
            "시도": safe_str(customer.get("시도", "")),
            "시군구": safe_str(customer.get("시군구", "")),
            "업종명": safe_str(customer.get("업종명", "")),
            "매출감소율": get_revenue_drop_rate(customer)
        }

        for _, program in api_df.iterrows():
            scored = score_api_program(customer, program)

            if scored:
                results.append({
                    **common,
                    "사업구분": "공고형",
                    "추천사업명": safe_str(program.get("pblancNm", "")),
                    "기관명": safe_str(program.get("jrsdInsttNm", "")),
                    "신청기간": safe_str(program.get("reqstBeginEndDe", "")),
                    "공고링크": safe_str(program.get("pblancUrl", "")),
                    "출처URL": safe_str(program.get("pblancUrl", "")),
                    "필요서류": "",
                    **scored
                })

        for _, policy in policy_df.iterrows():
            scored = score_policy_program(customer, policy)

            if scored:
                results.append({
                    **common,
                    "사업구분": "상시정책자금",
                    "추천사업명": safe_str(policy.get("상품명", "")),
                    "기관명": safe_str(policy.get("기관", "")),
                    "신청기간": "상시 또는 기관별 접수기간 확인",
                    "공고링크": safe_str(policy.get("출처URL", "")),
                    "출처URL": safe_str(policy.get("출처URL", "")),
                    "필요서류": safe_str(policy.get("필요서류", "")),
                    **scored
                })

        for _, employment in employment_df.iterrows():
            scored = score_employment_program(customer, employment)

            if scored:
                results.append({
                    **common,
                    "사업구분": "고용지원금",
                    "추천사업명": safe_str(employment.get("제도명", "")),
                    "기관명": safe_str(employment.get("기관", "")),
                    "신청기간": "제도별 상시/분기/예산소진 확인",
                    "공고링크": safe_str(employment.get("출처URL", "")),
                    "출처URL": safe_str(employment.get("출처URL", "")),
                    "필요서류": safe_str(employment.get("필요자료", "")),
                    **scored
                })

    if not results:
        print("매칭 결과 없음")
        return

    df = pd.DataFrame(results)

    df["상담체크리스트"] = df.apply(consultation_checklist, axis=1)
    df["신청판정"] = df.apply(make_decision, axis=1)
    df["신청판정사유"] = df.apply(decision_reason, axis=1)
    df["상담코멘트"] = df.apply(make_consulting_comment, axis=1)
    df["문자문안"] = df.apply(make_sms, axis=1)
    df["검토상태"] = "검토필요"
    df["문자발송여부"] = "미발송"

    df = (
        df.sort_values("신청가능성점수", ascending=False)
        .drop_duplicates(subset=["업체명", "사업구분", "추천사업명"], keep="first")
    )

    result_file = f"매칭결과_기술성조건반영_{NOW}.xlsx"
    df = add_new_flag(df, result_file)

    columns = [
        "사업구분", "업체명", "대표자명", "대표 연락처", "시도", "시군구", "업종명",
        "매출감소율", "추천등급", "등급설명", "매칭점수", "신청가능성점수",
        "신청가능성", "신청판정", "신청판정사유", "신규여부", "추천유형",
        "추천사업명", "기관명", "마감상태", "추천사유", "확인필요사항",
        "상담코멘트", "상담체크리스트", "매칭키워드", "필요서류", "신청기간",
        "출처URL", "공고링크", "문자문안", "검토상태", "문자발송여부",
        "비교기준파일", "추천ID"
    ]

    df = df[[col for col in columns if col in df.columns]]

    # 전체결과/업체별 시트 모두 업체별로 점수 높은 순서가 먼저 보이도록 정렬
    df = df.sort_values(
        ["업체명", "신청가능성점수", "매칭점수", "신규여부", "신청판정"],
        ascending=[True, False, False, False, True]
    )

    excluded_df = build_excluded_df(customer_df, policy_df, employment_df, df)

    save_pretty_excel(df, result_file, input_error_df, excluded_df)

    print("===================================")
    print("매출감소 반영 고도화 통합 매칭 완료")
    print(f"저장파일: {result_file}")
    print("===================================")


if __name__ == "__main__":
    main()
