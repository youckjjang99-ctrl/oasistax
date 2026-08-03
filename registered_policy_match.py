from __future__ import annotations

import json
import re
import shutil
from collections import Counter, defaultdict, deque
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from performance_cache import cache_generation


_CLOUD_METADATA_COLUMNS = (
    "_customer_id",
    "_company_uid",
    "_lifecycle_status",
    "_cloud_updated_at",
)


def normalize_business_no(value: Any) -> str:
    raw = "" if value is None else str(value)
    digits = re.sub(r"[^0-9]", "", raw)
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return raw.strip()


def normalize_text(value: Any) -> str:
    raw = "" if value is None else str(value)
    return re.sub(r"\s+", "", raw.strip()).lower()


def _business_no_merge_key(value: Any) -> str:
    """Return an identity key only for an exact ten-digit business number."""
    raw = "" if value is None else str(value)
    digits = re.sub(r"[^0-9]", "", raw)
    return digits if len(digits) == 10 else ""


def _is_blank_customer_value(value: Any) -> bool:
    if value is None:
        return True
    try:
        missing = pd.isna(value)
        if isinstance(missing, bool) and missing:
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower() in {
        "",
        "nan",
        "none",
        "nat",
        "<na>",
    }


def _merge_registered_customer_frames(
    local_df: pd.DataFrame | None,
    cloud_df: pd.DataFrame | None,
) -> pd.DataFrame:
    """Losslessly union local and cloud rows using only strong identity.

    A key is merged only when it occurs exactly once in each source. Duplicate
    candidates remain separate so source ordering can never attach the wrong
    cloud customer UUID.  For a unique identified pair, a nonblank local value
    wins and the cloud value only fills a blank. Rows without a ten-digit
    business number never match one another.
    """
    local = (
        local_df.copy()
        if isinstance(local_df, pd.DataFrame)
        else pd.DataFrame()
    )
    cloud = (
        cloud_df.copy()
        if isinstance(cloud_df, pd.DataFrame)
        else pd.DataFrame()
    )
    local.columns = [str(column).strip() for column in local.columns]
    cloud.columns = [str(column).strip() for column in cloud.columns]
    local = local.reset_index(drop=True)
    cloud = cloud.reset_index(drop=True)

    columns = list(local.columns)
    columns.extend(column for column in cloud.columns if column not in columns)
    if not columns:
        return pd.DataFrame()

    local_rows = local.to_dict(orient="records")
    cloud_rows = cloud.to_dict(orient="records")
    local_key_counts = Counter(
        key
        for row in local_rows
        if (key := _business_no_merge_key(row.get("사업자등록번호")))
    )
    cloud_key_counts = Counter(
        key
        for row in cloud_rows
        if (key := _business_no_merge_key(row.get("사업자등록번호")))
    )
    cloud_by_business_no: dict[str, deque[int]] = defaultdict(deque)
    for index, row in enumerate(cloud_rows):
        key = _business_no_merge_key(row.get("사업자등록번호"))
        if (
            key
            and local_key_counts.get(key) == 1
            and cloud_key_counts.get(key) == 1
        ):
            cloud_by_business_no[key].append(index)

    consumed_cloud_rows: set[int] = set()
    merged_rows: list[dict[str, Any]] = []
    for local_row in local_rows:
        key = _business_no_merge_key(local_row.get("사업자등록번호"))
        cloud_index = (
            cloud_by_business_no[key].popleft()
            if key and cloud_by_business_no.get(key)
            else None
        )
        if cloud_index is None:
            merged_rows.append(
                {column: local_row.get(column, pd.NA) for column in columns}
            )
            continue

        consumed_cloud_rows.add(cloud_index)
        cloud_row = cloud_rows[cloud_index]
        merged: dict[str, Any] = {}
        for column in columns:
            local_value = local_row.get(column, pd.NA)
            cloud_value = cloud_row.get(column, pd.NA)
            if column in _CLOUD_METADATA_COLUMNS:
                merged[column] = (
                    local_value
                    if _is_blank_customer_value(cloud_value)
                    else cloud_value
                )
            else:
                merged[column] = (
                    cloud_value
                    if _is_blank_customer_value(local_value)
                    else local_value
                )
        merged_rows.append(merged)

    for index, cloud_row in enumerate(cloud_rows):
        if index in consumed_cloud_rows:
            continue
        merged_rows.append(
            {column: cloud_row.get(column, pd.NA) for column in columns}
        )

    return pd.DataFrame(merged_rows, columns=columns).reset_index(drop=True)


def _rpc_function_unavailable(exc: Exception, function_name: str) -> bool:
    message = str(exc or "").lower()
    function_name = str(function_name or "").lower()
    return (
        "pgrst202" in message
        or "could not find the function" in message
        or (
            function_name in message
            and (
                "does not exist" in message
                or "undefined function" in message
                or "42883" in message
            )
        )
    )


def _owner_user_id_from_cumulative_path(cumulative_path: Path) -> str:
    """회원별 누적DB 경로에서 로그인 회원 ID를 복원한다."""
    try:
        return str(Path(cumulative_path).parent.name or "").strip()
    except Exception:
        return ""


def _parse_customer_data(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return dict(value)

    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
            if isinstance(parsed, dict):
                return parsed
        except Exception:
            pass

    return {}


def _load_registered_customers_from_cloud(
    owner_user_id: str,
) -> pd.DataFrame | None:
    """Supabase에서 로그인 회원 소유 고객을 조회한다."""
    owner_user_id = str(owner_user_id or "").strip().lower()
    if not owner_user_id:
        return None

    try:
        from cloud_db import (
            CloudDatabase,
            TABLE_CUSTOMERS,
            cloud_is_configured,
        )

        if not cloud_is_configured():
            return None

        database = CloudDatabase()
        try:
            rows = database.rpc(
                "oasis_list_unified_customers",
                {"p_owner_user_id": owner_user_id},
            )
            if not isinstance(rows, list):
                rows = [rows] if isinstance(rows, dict) else []
        except Exception as exc:
            if not _rpc_function_unavailable(
                exc,
                "oasis_list_unified_customers",
            ):
                return None
            rows = database.select_all(
                TABLE_CUSTOMERS,
                filters={"owner_user_id": owner_user_id},
                order="company_name.asc,id.asc",
                max_rows=50000,
            )

        # Service-role 조회는 소유자 조건을 절대 제거하지 않는다.
    except Exception:
        return None

    records: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            continue

        record = _parse_customer_data(row.get("customer_data"))
        canonical_fields = {
            "업체명": row.get("company_name"),
            "대표자명": row.get("representative_name"),
            "사업자등록번호": row.get("business_no"),
            "업종명": row.get("industry_name"),
            "사업장 소재지": row.get("address"),
            "담당자": row.get("manager_name"),
        }
        for field, value in canonical_fields.items():
            if value is not None and str(value).strip():
                record[field] = value

        record["_customer_id"] = row.get("id")
        record["_company_uid"] = row.get("company_uid")
        record["_lifecycle_status"] = row.get("lifecycle_status")
        record["_cloud_updated_at"] = row.get("updated_at")
        records.append(record)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    df = df.dropna(how="all").copy()
    df.columns = [str(column).strip() for column in df.columns]
    if "사업자등록번호" in df.columns:
        df["사업자등록번호"] = df["사업자등록번호"].map(
            normalize_business_no
        )
    return df.reset_index(drop=True)

def _load_registered_customers_from_excel(
    cumulative_path: Path,
) -> pd.DataFrame:
    if not cumulative_path.exists():
        return pd.DataFrame()

    try:
        df = pd.read_excel(cumulative_path, sheet_name="고객DB")
    except Exception:
        return pd.DataFrame()

    if df.empty:
        return df

    df = df.dropna(how="all").copy()
    df.columns = [str(column).strip() for column in df.columns]
    return df.reset_index(drop=True)


@st.cache_data(
    ttl=45,
    max_entries=128,
    show_spinner=False,
    scope="session",
)
def _load_registered_customers_cached(
    cumulative_path_str: str,
    file_mtime_ns: int,
    file_size: int,
    owner_user_id: str,
    generation: int,
) -> pd.DataFrame:
    """Cache one user-scoped customer list for a short, bounded period."""
    del file_mtime_ns, file_size, generation
    local_df = _load_registered_customers_from_excel(
        Path(cumulative_path_str)
    )
    cloud_df = _load_registered_customers_from_cloud(owner_user_id)
    return _merge_registered_customer_frames(local_df, cloud_df)


def load_registered_customers(
    cumulative_path: Path,
    owner_user_id: str | None = None,
) -> pd.DataFrame:
    explicit_owner = str(owner_user_id or "").strip().lower()
    resolved_owner = (
        explicit_owner
        or _owner_user_id_from_cumulative_path(cumulative_path).lower()
    )
    cumulative_path = Path(cumulative_path)
    try:
        stat = cumulative_path.stat()
        mtime_ns = int(stat.st_mtime_ns)
        file_size = int(stat.st_size)
    except OSError:
        mtime_ns = 0
        file_size = 0
    cached = _load_registered_customers_cached(
        str(cumulative_path),
        mtime_ns,
        file_size,
        resolved_owner,
        cache_generation("registered_customers", resolved_owner),
    )
    return cached.copy()


def build_customer_labels(df: pd.DataFrame) -> tuple[list[str], dict[str, int]]:
    labels: list[str] = []
    row_map: dict[str, int] = {}
    used: dict[str, int] = {}

    for index, row in df.iterrows():
        lifecycle_value = row.get("_lifecycle_status", "")
        lifecycle_status = (
            ""
            if _is_blank_customer_value(lifecycle_value)
            else str(lifecycle_value).strip().lower()
        )
        if lifecycle_status == "archived":
            continue
        company_value = row.get("업체명", "")
        representative_value = row.get("대표자명", "")
        company = (
            ""
            if _is_blank_customer_value(company_value)
            else str(company_value).strip()
        )
        representative = (
            ""
            if _is_blank_customer_value(representative_value)
            else str(representative_value).strip()
        )
        business_no = normalize_business_no(
            ""
            if _is_blank_customer_value(row.get("사업자등록번호", ""))
            else row.get("사업자등록번호", "")
        )

        if not company:
            continue

        label_parts = [company]
        if business_no:
            label_parts.append(business_no)
        if representative:
            label_parts.append(representative)

        base_label = " · ".join(label_parts)
        used[base_label] = used.get(base_label, 0) + 1
        label = (
            base_label
            if used[base_label] == 1
            else f"{base_label} · {used[base_label]}"
        )

        labels.append(label)
        row_map[label] = int(index)

    return labels, row_map


def customer_preview(row: pd.Series) -> pd.DataFrame:
    fields = [
        "업체명",
        "대표자명",
        "사업자등록번호",
        "업종명",
        "사업장 소재지",
        "설립일",
        "종업원수",
        "상시근로자수",
        "매출액",
        "연매출",
        "영업이익",
        "당기순이익",
        "벤처",
        "이노비즈",
        "메인비즈",
        "기업부설연구소",
        "연구개발전담부서",
        "특허보유",
    ]

    rows = []
    for field in fields:
        if field not in row.index:
            continue

        value = row.get(field)
        if value is None or str(value).strip().lower() in {
            "",
            "nan",
            "none",
            "nat",
        }:
            continue

        if field in {
            "매출액",
            "연매출",
            "영업이익",
            "당기순이익",
        }:
            try:
                value = f"{int(float(str(value).replace(',', ''))):,}"
            except Exception:
                pass

        rows.append({"항목": field, "값": value})

    return pd.DataFrame(rows)


def _find_header_row_and_columns(worksheet):
    for row_number in range(1, min(worksheet.max_row, 40) + 1):
        values = {
            str(worksheet.cell(row_number, column).value or "").strip(): column
            for column in range(1, worksheet.max_column + 1)
        }
        if "업체명" in values:
            return row_number, values
    raise ValueError("고객DB 시트에서 '업체명' 헤더를 찾지 못했습니다.")


def _append_text(current: Any, addition: str) -> str:
    current_text = str(current or "").strip()
    addition = str(addition or "").strip()

    if not addition:
        return current_text
    if not current_text:
        return addition
    if addition in current_text:
        return current_text
    return f"{current_text} / {addition}"


def _ensure_column(worksheet, header_row: int, columns: dict[str, int], name: str) -> int:
    if name in columns:
        return columns[name]

    column_number = worksheet.max_column + 1
    worksheet.cell(header_row, column_number).value = name
    columns[name] = column_number
    return column_number


def create_single_customer_workbook(
    cumulative_path: Path,
    destination_dir: Path,
    selected_row: pd.Series,
    manager_name: str = "",
    matching_preferences: dict[str, Any] | None = None,
) -> Path:
    """
    기존 누적 고객DB를 복사한 뒤 고객DB 시트만 선택 고객 1행으로 필터링한다.
    원본 파일과 기존 고객리스트는 수정하지 않는다.
    다른 정책자금·고용지원금 시트와 서식은 그대로 유지한다.
    """
    if not cumulative_path.exists():
        raise FileNotFoundError("누적 고객DB를 찾지 못했습니다.")

    destination_dir.mkdir(parents=True, exist_ok=True)

    company_name = str(selected_row.get("업체명", "고객") or "고객").strip()
    safe_company = re.sub(r'[\\/:*?"<>|]', "_", company_name)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destination = destination_dir / (
        f"등록고객매칭_{safe_company}_{timestamp}.xlsx"
    )

    shutil.copy2(cumulative_path, destination)

    workbook = load_workbook(destination)
    if "고객DB" not in workbook.sheetnames:
        raise ValueError("누적 고객DB에 '고객DB' 시트가 없습니다.")

    worksheet = workbook["고객DB"]
    header_row, columns = _find_header_row_and_columns(worksheet)

    target_business_no = normalize_business_no(
        selected_row.get("사업자등록번호", "")
    )
    target_company = normalize_text(selected_row.get("업체명", ""))
    target_representative = normalize_text(
        selected_row.get("대표자명", "")
    )

    company_column = columns["업체명"]
    business_column = columns.get("사업자등록번호")
    representative_column = columns.get("대표자명")
    manager_column = columns.get("담당자")

    matching_rows: list[int] = []

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        company = normalize_text(
            worksheet.cell(row_number, company_column).value
        )
        business_no = (
            normalize_business_no(
                worksheet.cell(row_number, business_column).value
            )
            if business_column
            else ""
        )
        representative = (
            normalize_text(
                worksheet.cell(row_number, representative_column).value
            )
            if representative_column
            else ""
        )

        is_match = False
        if (
            target_business_no
            and len(target_business_no.replace("-", "")) == 10
            and business_no == target_business_no
        ):
            is_match = True
        elif company == target_company:
            if target_representative and representative_column:
                is_match = representative == target_representative
            else:
                is_match = True

        if is_match:
            matching_rows.append(row_number)

    if not matching_rows:
        raise ValueError("누적 고객DB에서 선택 고객 행을 찾지 못했습니다.")

    keep_row = matching_rows[0]

    # 선택 고객 외 데이터행 삭제. 제목·헤더·서식·다른 시트는 유지.
    for row_number in range(worksheet.max_row, header_row, -1):
        if row_number != keep_row:
            worksheet.delete_rows(row_number, 1)

    # 삭제 후 선택 고객은 header_row + 1 위치가 된다.
    if manager_column and manager_name.strip():
        worksheet.cell(header_row + 1, manager_column).value = (
            manager_name.strip()
        )

    preferences = dict(matching_preferences or {})
    if preferences:
        keyword_column = _ensure_column(
            worksheet,
            header_row,
            columns,
            "키워드메모",
        )
        memo_column = _ensure_column(
            worksheet,
            header_row,
            columns,
            "비고",
        )
        topic_columns = [
            _ensure_column(
                worksheet,
                header_row,
                columns,
                f"희망상담주제{index}",
            )
            for index in range(1, 4)
        ]
        purpose_columns = [
            _ensure_column(
                worksheet,
                header_row,
                columns,
                f"희망자금용도{index}",
            )
            for index in range(1, 4)
        ]

        matching_keywords = preferences.get("매칭키워드", []) or []
        interest_fields = preferences.get("관심지원분야", []) or []
        exclusion_keywords = preferences.get("제외키워드", []) or []
        fund_purpose = str(
            preferences.get("자금사용목적", "") or ""
        ).strip()

        keyword_text = ", ".join(
            [str(item).strip() for item in matching_keywords + interest_fields if str(item).strip()]
        )
        if keyword_text:
            current = worksheet.cell(
                header_row + 1,
                keyword_column,
            ).value
            worksheet.cell(
                header_row + 1,
                keyword_column,
            ).value = _append_text(current, keyword_text)

        for column_number, value in zip(
            topic_columns,
            list(interest_fields)[:3],
        ):
            worksheet.cell(header_row + 1, column_number).value = value

        purpose_values = []
        if fund_purpose:
            purpose_values.append(fund_purpose)
        purpose_values.extend(
            [
                str(item).strip()
                for item in interest_fields
                if str(item).strip()
            ]
        )

        for column_number, value in zip(
            purpose_columns,
            purpose_values[:3],
        ):
            worksheet.cell(header_row + 1, column_number).value = value

        memo_parts = []
        if exclusion_keywords:
            memo_parts.append(
                "제외키워드: "
                + ", ".join(
                    str(item).strip()
                    for item in exclusion_keywords
                    if str(item).strip()
                )
            )

        planned_amount = str(
            preferences.get("투자예정금액", "") or ""
        ).strip()
        planned_timing = str(
            preferences.get("투자예정시기", "") or ""
        ).strip()

        if planned_amount:
            memo_parts.append(f"투자예정금액: {planned_amount}")
        if planned_timing:
            memo_parts.append(f"투자예정시기: {planned_timing}")

        if memo_parts:
            current = worksheet.cell(
                header_row + 1,
                memo_column,
            ).value
            worksheet.cell(
                header_row + 1,
                memo_column,
            ).value = _append_text(
                current,
                " / ".join(memo_parts),
            )

    workbook.save(destination)
    return destination
