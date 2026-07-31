from __future__ import annotations

import hashlib
import hmac
import html
import json
import os
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Callable, Mapping

import pandas as pd
import streamlit as st
from cryptography.fernet import Fernet, InvalidToken

from claim_correction_catalog import (
    AUTOMATIC_COLLECTION_CODES,
    DOCUMENT_SPECS,
    automatic_collection_supported,
)
from claim_correction_repository import (
    CLAIM_DEFAULT_COLLECTION_KEY,
    CLAIM_DOWNLOAD_EXTENSION_BY_CONTENT_TYPE,
    CLAIM_DOWNLOAD_URL_TTL_SECONDS,
    CLAIM_STORAGE_BUCKET,
    ClaimRepository,
    ClaimRepositoryError,
)
from registered_policy_match import (
    build_customer_labels,
    load_registered_customers,
)
from tilko_claim_client import (
    ClaimProviderError,
    CollectedClaimDocument,
    TilkoClaimClient,
    provider_readiness,
)
from utils import get_user_cumulative_db_path


CONSENT_VERSION = "claim-collection-v2-2026-07"
RETENTION_POLICY_VERSION = "claim-document-retention-v1-2026-07"
CONSENT_NOTICE_TEXT = (
    "고객에게 수집 항목·이용 목적·보유기간·제3자 제공 내용을 안내했고 "
    "유효한 동의를 확인했습니다. 주민등록번호는 DB·로그에 저장하지 않고 "
    "인증 및 자료수집 중 암호화된 서버 메모리에 최대 45분 보관한 뒤 "
    "즉시 삭제합니다."
)
COLLECTION_AUTHORITY_TEXT = (
    "민감정보 처리가 필요한 경우 적용되는 법적 근거와 위임 범위를 "
    "확인했습니다."
)
CONSENT_TEXT_SHA256 = hashlib.sha256(
    f"{CONSENT_NOTICE_TEXT}|{COLLECTION_AUTHORITY_TEXT}".encode("utf-8")
).hexdigest()
AUTH_TTL_SECONDS = 10 * 60
COLLECTION_TTL_SECONDS = 45 * 60
AUTH_POLL_SECONDS = 1.0
AUTH_MEDIUM_POLL_SECONDS = 3.0
AUTH_SLOW_POLL_SECONDS = 10.0
AUTH_FAST_POLL_WINDOW_SECONDS = 30.0
AUTH_MEDIUM_POLL_WINDOW_SECONDS = 90.0
TRANSIENT_AUTH_RETRY_SECONDS = 3.0
COMWEL_DISPATCH_DELAY_SECONDS = 1.0
_CLAIM_JOB_CIPHER = Fernet(Fernet.generate_key())
_CLAIM_BUSINESS_TOKEN_KEY = Fernet.generate_key()
_CLAIM_JOB_LOCK = threading.RLock()
_CLAIM_JOBS: dict[str, dict[str, Any]] = {}
_CLAIM_JOB_EXECUTOR = ThreadPoolExecutor(
    max_workers=4,
    thread_name_prefix="claim-auth",
)
_CLAIM_SWEEPER_STARTED = False
STATUS_LABELS = {
    "request_ready": "발송 준비",
    "auth_preparing": "인증 준비",
    "auth_requested": "고객 인증 대기",
    "auth_pending": "고객 인증 대기",
    "auth_complete": "인증 완료",
    "auth_partial": "일부 인증 완료 · 재요청 필요",
    "certificate_required": "공동인증서 대기",
    "collection_queued": "자료수집 대기",
    "auth_complete_collection_pending": "인증 완료 · 일부 자료 재수집 필요",
    "integration_required": "자동수집 연동 예정",
    "collecting": "자료수집 중",
    "collected": "수집 완료",
    "ready": "결과 확인 가능",
    "failed": "실패",
    "not_requested": "제외",
}

_REMOTE_INVITE_NOTICE_KEY = "claim_remote_invite_notice_v1"
_REMOTE_INVITE_NAME_PATTERN = re.compile(r"^[가-힣]+(?:[ ·][가-힣]+)*$")
_REMOTE_INVITE_PHONE_PATTERN = re.compile(r"^010\d{8}$")
_KOREA_TIMEZONE = timezone(timedelta(hours=9))
_REMOTE_INVITE_ERROR_MESSAGES = {
    "PUBLIC_BASE_URL_REQUIRED": (
        "고객 인증 주소 설정이 완료되지 않았습니다. 관리자에게 문의해주세요."
    ),
    "REMOTE_STORAGE_NOT_CONFIGURED": (
        "원격 인증 저장소 연결이 준비되지 않았습니다. 관리자에게 문의해주세요."
    ),
    "REMOTE_CRYPTO_NOT_CONFIGURED": (
        "원격 인증 보안 설정이 준비되지 않았습니다. 관리자에게 문의해주세요."
    ),
    "REMOTE_REPOSITORY_UNAVAILABLE": (
        "원격 인증 저장소에 연결하지 못했습니다. 잠시 후 다시 시도해주세요."
    ),
    "REMOTE_INVITE_CREATE_FAILED": (
        "고객 인증 요청을 저장하지 못했습니다. 잠시 후 다시 시도해주세요."
    ),
    "REMOTE_INVITE_NOT_READY": (
        "카카오톡 원격 인증 발송 설정이 완료되지 않았습니다. "
        "관리자에게 문의해주세요."
    ),
    "REMOTE_OWNER_REQUIRED": "로그인 정보를 다시 확인해주세요.",
    "INVALID_CUSTOMER_NAME": "고객 이름을 한글로 정확히 입력해주세요.",
    "INVALID_CUSTOMER_PHONE": "010으로 시작하는 휴대전화번호를 확인해주세요.",
}


class RemoteInviteUIError(RuntimeError):
    """A redacted remote-invite error that is safe to render."""


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "nat"}:
        return ""
    return text


def _claim_document_is_no_data(document: dict[str, Any]) -> bool:
    facts = document.get("facts")
    return bool(isinstance(facts, dict) and facts.get("no_data") is True)


def _claim_document_is_provider_blocked(document: dict[str, Any]) -> bool:
    """Return True when no provider query produced the stored outcome."""
    facts = document.get("facts")
    if not isinstance(facts, dict):
        return False
    if facts.get("provider_query_attempted") is False:
        return True
    return bool(
        str(document.get("document_code", "")).strip()
        == "comwel_workplace_rate"
        and str(facts.get("no_data_reason", "")).strip()
        == "no_management_number"
    )


def _claim_document_needs_recollection(document: dict[str, Any]) -> bool:
    """Return True for successful-looking rows created by an obsolete query."""
    if (
        str(document.get("document_code", "")).strip()
        != "hometax_income_tax_return"
        or str(document.get("status", "") or "").strip().lower()
        != "ready"
    ):
        return False
    facts = document.get("facts")
    return not (
        isinstance(facts, dict)
        and str(facts.get("query_strategy", "")).strip()
        in {"filing_year_v2", "filing_year_taxpayer_v3"}
        and facts.get("tax_year_verified") is True
    )


def _claim_document_is_downloadable(
    document: dict[str, Any],
    *,
    now: datetime | None = None,
) -> bool:
    if (
        _claim_document_is_no_data(document)
        or _claim_document_needs_recollection(document)
    ):
        return False
    if (
        str(document.get("status", "") or "").strip().lower() != "ready"
        or not _clean(document.get("storage_path"))
        or bool(document.get("deleted_at"))
        or _clean(document.get("storage_bucket")) != CLAIM_STORAGE_BUCKET
    ):
        return False
    content_type = _clean(document.get("content_type")).lower()
    expected_extension = CLAIM_DOWNLOAD_EXTENSION_BY_CONTENT_TYPE.get(
        content_type,
        "",
    )
    if (
        not expected_extension
        or not _clean(document.get("storage_path"))
        .lower()
        .endswith(expected_extension)
    ):
        return False
    retention_until = _clean(document.get("retention_until"))
    if not retention_until:
        return False
    try:
        retention_deadline = datetime.fromisoformat(
            retention_until.replace("Z", "+00:00")
        )
        if retention_deadline.tzinfo is None:
            retention_deadline = retention_deadline.replace(
                tzinfo=timezone.utc
            )
    except ValueError:
        return False
    reference_now = now or datetime.now(timezone.utc)
    if reference_now.tzinfo is None:
        reference_now = reference_now.replace(tzinfo=timezone.utc)
    return retention_deadline > reference_now


def _digits(value: Any) -> str:
    return re.sub(r"[^0-9]", "", str(value or ""))


def _validate_remote_invite_input(
    customer_name: Any,
    customer_phone: Any,
) -> tuple[str, str, list[str]]:
    name = re.sub(r"\s+", " ", str(customer_name or "")).strip()
    phone = _digits(customer_phone)
    errors: list[str] = []
    hangul_count = len(re.sub(r"[^가-힣]", "", name))
    if (
        not _REMOTE_INVITE_NAME_PATTERN.fullmatch(name)
        or not 2 <= hangul_count <= 20
    ):
        errors.append("고객 이름은 한글 2~20자로 입력해주세요.")
    if not _REMOTE_INVITE_PHONE_PATTERN.fullmatch(phone):
        errors.append("010으로 시작하는 휴대전화번호 11자리를 입력해주세요.")
    return name, phone, errors


def _remote_invite_safe_error(exc: BaseException) -> str:
    error_code = str(
        getattr(exc, "error_code", "")
        or getattr(exc, "code", "")
        or ""
    ).strip().upper()
    return _REMOTE_INVITE_ERROR_MESSAGES.get(
        error_code,
        "인증 링크 발송 준비 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.",
    )


def _remote_invite_runtime_readiness(
    checker: Callable[[], Mapping[str, Any]] | None = None,
) -> tuple[bool, str]:
    selected_checker = checker
    if selected_checker is None:
        try:
            from claim_remote_service import (  # noqa: PLC0415
                remote_invite_environment_readiness,
            )
        except (ImportError, ModuleNotFoundError):
            return (
                False,
                "카카오톡 원격 인증 발송 기능을 준비하지 못했습니다. "
                "관리자에게 문의해주세요.",
            )
        selected_checker = remote_invite_environment_readiness
    try:
        readiness = dict(selected_checker() or {})
    except Exception:
        return (
            False,
            "카카오톡 원격 인증 발송 설정을 확인하지 못했습니다. "
            "관리자에게 문의해주세요.",
        )
    if bool(readiness.get("ready")):
        return True, ""
    return (
        False,
        "카카오톡 원격 인증 발송 설정이 완료되지 않았습니다. "
        "관리자에게 문의해주세요.",
    )


def _create_remote_claim_invite(
    *,
    owner_user_id: str,
    requested_by: str,
    customer_name: str,
    customer_phone: str,
    invite_creator: Callable[..., dict[str, Any]] | None = None,
) -> dict[str, Any]:
    name, phone, errors = _validate_remote_invite_input(
        customer_name,
        customer_phone,
    )
    if errors:
        raise RemoteInviteUIError(errors[0])

    creator = invite_creator
    if creator is None:
        try:
            from claim_remote_service import (  # noqa: PLC0415
                create_staff_claim_invite,
            )
        except (ImportError, ModuleNotFoundError) as exc:
            raise RemoteInviteUIError(
                "카카오톡 인증 발송 기능을 불러오지 못했습니다. "
                "관리자에게 문의해주세요."
            ) from exc
        creator = create_staff_claim_invite

    try:
        result = creator(
            owner_user_id=str(owner_user_id or "").strip().lower(),
            requested_by=str(requested_by or owner_user_id).strip(),
            customer_name=name,
            customer_phone=phone,
        )
    except Exception as exc:
        raise RemoteInviteUIError(_remote_invite_safe_error(exc)) from None

    if not isinstance(result, dict) or not result.get("invite_id"):
        raise RemoteInviteUIError(
            "고객 인증 요청의 저장 결과를 확인하지 못했습니다. "
            "잠시 후 다시 시도해주세요."
        )
    if result.get("message_queued") is not True:
        raise RemoteInviteUIError(
            "카카오톡 발송 대기열 등록 결과를 확인하지 못했습니다. "
            "잠시 후 다시 시도해주세요."
        )
    return dict(result)


def _format_remote_invite_expiry(value: Any) -> str:
    selected: datetime | None = None
    if isinstance(value, datetime):
        selected = value
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            selected = datetime.fromtimestamp(float(value), timezone.utc)
        except (OverflowError, OSError, ValueError):
            selected = None
    else:
        text = str(value or "").strip()
        if text:
            try:
                selected = datetime.fromisoformat(text.replace("Z", "+00:00"))
            except ValueError:
                selected = None
    if selected is None:
        return ""
    if selected.tzinfo is None:
        selected = selected.replace(tzinfo=timezone.utc)
    return selected.astimezone(_KOREA_TIMEZONE).strftime(
        "%Y년 %m월 %d일 %H:%M"
    )


def _format_business_no(value: Any) -> str:
    digits = _digits(value)
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return _clean(value)


def _format_phone(value: Any) -> str:
    digits = _digits(value)
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return _clean(value)


def _birth_date_from_identity(front: str, rear: str) -> str:
    front_digits = _digits(front)
    rear_digits = _digits(rear)
    if len(front_digits) != 6:
        return ""
    century_code = rear_digits[:1]
    if century_code in {"1", "2", "5", "6"}:
        century = "19"
    elif century_code in {"3", "4", "7", "8"}:
        century = "20"
    else:
        return ""
    birth_date = f"{century}{front_digits}"
    try:
        datetime.strptime(birth_date, "%Y%m%d")
    except ValueError:
        return ""
    return birth_date


def _is_valid_business_no(value: Any) -> bool:
    digits = _digits(value)
    if len(digits) != 10:
        return False
    weights = (1, 3, 7, 1, 3, 7, 1, 3, 5)
    checksum = sum(
        int(digit) * weight
        for digit, weight in zip(digits[:9], weights)
    )
    checksum += (int(digits[8]) * 5) // 10
    expected = (10 - (checksum % 10)) % 10
    return expected == int(digits[-1])


def _repository(user_id: str) -> tuple[ClaimRepository | None, str]:
    try:
        repository = ClaimRepository(user_id)
        status = repository.status()
        if not status.available:
            return None, status.message
        return repository, ""
    except ClaimRepositoryError as exc:
        return None, str(exc)


def _source_status(value: Any) -> str:
    key = str(value or "").strip()
    return STATUS_LABELS.get(key, key or "-")


def _claim_result_document_status(
    document: dict[str, Any],
    selected_case: dict[str, Any],
    *,
    multiple_management_numbers: bool = False,
    no_management_workplaces: bool = False,
) -> str:
    document_code = str(document.get("document_code", ""))
    status = str(document.get("status", ""))
    facts = document.get("facts")
    document_safe_error_code = (
        str(facts.get("safe_error_code", ""))
        if isinstance(facts, dict)
        else ""
    )
    case_safe_error_code = str(
        selected_case.get("last_safe_error_code", "")
    )
    business_dependent_codes = {
        "hometax_business_registration_certificate",
        "comwel_management_number_list",
        "comwel_workplace_rate",
    }

    if _claim_document_is_provider_blocked(document):
        return "API 미호출 · 재수집 필요"
    if _claim_document_is_no_data(document):
        return "조회된 신고내역 없음"
    if (
        document_code == "comwel_workplace_rate"
        and no_management_workplaces
    ):
        return "조회된 가입 사업장 없음"
    if (
        document_code == "comwel_worker_status"
        and status == "integration_required"
    ):
        return "공동인증서 필요"
    if (
        document_code == "comwel_workplace_rate"
        and status == "integration_required"
        and multiple_management_numbers
    ):
        return "사업장 선택 필요"
    if (
        document_code in business_dependent_codes
        and (
            document_safe_error_code == "BUSINESS_NUMBER_NOT_FOUND"
            or (
                status in {"auth_pending", "integration_required"}
                and case_safe_error_code == "BUSINESS_NUMBER_NOT_FOUND"
            )
        )
    ):
        return "홈택스 사업자번호 확인 필요"
    if (
        document_code == "hometax_business_registration_list"
        and status == "ready"
        and str(
            (facts or {}).get("record_count", "")
            if isinstance(facts, dict)
            else ""
        )
        == "0"
        and case_safe_error_code == "BUSINESS_NUMBER_NOT_FOUND"
    ):
        return "사업자번호 미확인"
    if (
        status == "integration_required"
        and automatic_collection_supported(document_code)
    ):
        return "재수집 필요"
    if (
        status == "auth_pending"
        and not automatic_collection_supported(document_code)
    ):
        return _source_status("integration_required")
    return _source_status(status)


def _resolve_auth_progress(
    expected_sources: list[str],
    source_statuses: dict[str, Any],
) -> tuple[str, bool, bool]:
    expected = [
        str(source_statuses.get(source, "") or "").strip()
        for source in expected_sources
        if source in {"hometax", "comwel"}
    ]
    all_completed = bool(expected) and all(
        status == "auth_complete" for status in expected
    )
    any_failed = any(status == "failed" for status in expected)
    overall_status = (
        "auth_complete_collection_pending"
        if all_completed
        else "auth_partial"
        if any_failed
        else "auth_pending"
    )
    return overall_status, all_completed, any_failed


CLAIM_AUTH_STAGE_MESSAGES = (
    "국세청 홈택스 인증요청을 보냈습니다.",
    "홈택스 인증이 완료되어 근로복지공단 인증을 발송했습니다.",
    "최종 인증이 모두 완료되었습니다.",
    "자료를 수집하겠습니다.",
)


def _claim_auth_stage(case: dict[str, Any]) -> tuple[int, str]:
    overall_status = str(case.get("overall_status", "") or "").strip()
    hometax_status = str(case.get("hometax_status", "") or "").strip()
    comwel_status = str(case.get("comwel_status", "") or "").strip()

    both_complete = (
        hometax_status == "auth_complete"
        and comwel_status == "auth_complete"
    )
    if hometax_status == "failed":
        return 1, "국세청 홈택스 인증을 완료하지 못했습니다."
    if hometax_status == "auth_complete" and comwel_status == "failed":
        return 2, "근로복지공단 인증을 완료하지 못했습니다."
    if both_complete and overall_status in {
        "auth_complete_collection_pending",
        "collecting",
        "collected",
        "ready",
    }:
        return 4, CLAIM_AUTH_STAGE_MESSAGES[3]
    if both_complete:
        return 3, CLAIM_AUTH_STAGE_MESSAGES[2]
    if (
        hometax_status == "auth_complete"
        and comwel_status in {
            "request_ready",
            "auth_requested",
            "auth_pending",
            "auth_complete",
        }
    ):
        return 2, CLAIM_AUTH_STAGE_MESSAGES[1]
    if hometax_status in {
        "auth_requested",
        "auth_pending",
        "auth_complete",
        "failed",
    }:
        return 1, CLAIM_AUTH_STAGE_MESSAGES[0]
    return 0, "국세청 홈택스 인증요청을 준비하고 있습니다."


def _claim_active_collection_documents(
    documents: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    taxpayer_returns_by_year = {
        int(document.get("period_year") or 0): document
        for document in documents
        if str(document.get("document_code", "")).strip()
        == "hometax_income_tax_return"
        and str(
            document.get(
                "collection_key",
                CLAIM_DEFAULT_COLLECTION_KEY,
            )
            or CLAIM_DEFAULT_COLLECTION_KEY
        ).strip().lower()
        == CLAIM_DEFAULT_COLLECTION_KEY
        and str(document.get("status", "")).strip().lower() == "ready"
        and isinstance(document.get("facts"), dict)
        and str(document["facts"].get("query_strategy", "")).strip()
        == "filing_year_taxpayer_v3"
    }
    variant_groups = {
        (
            str(document.get("source", "")).strip(),
            str(document.get("document_code", "")).strip(),
            int(document.get("period_year") or 0),
        )
        for document in documents
        if str(
            document.get(
                "collection_key",
                CLAIM_DEFAULT_COLLECTION_KEY,
            )
            or CLAIM_DEFAULT_COLLECTION_KEY
        ).strip().lower()
        != CLAIM_DEFAULT_COLLECTION_KEY
    }

    def is_superseded_default(document: dict[str, Any]) -> bool:
        collection_key = str(
            document.get(
                "collection_key",
                CLAIM_DEFAULT_COLLECTION_KEY,
            )
            or CLAIM_DEFAULT_COLLECTION_KEY
        ).strip().lower()
        if collection_key != CLAIM_DEFAULT_COLLECTION_KEY:
            return False
        facts = document.get("facts")
        if (
            str(document.get("document_code", "")).strip()
            == "hometax_income_tax_return"
            and isinstance(facts, dict)
            and str(facts.get("query_strategy", "")).strip()
            == "filing_year_taxpayer_v3"
        ):
            return False
        group = (
            str(document.get("source", "")).strip(),
            str(document.get("document_code", "")).strip(),
            int(document.get("period_year") or 0),
        )
        legacy_scoped_file = bool(
            isinstance(facts, dict)
            and (
                _document_scope_fingerprint(document)
                or str(facts.get("collection_scope", "")).strip().lower()
                in {"business", "management"}
            )
        )
        return group in variant_groups or legacy_scoped_file

    def is_superseded_income_tax_business_scope(
        document: dict[str, Any],
    ) -> bool:
        if (
            str(document.get("document_code", "")).strip()
            != "hometax_income_tax_return"
        ):
            return False
        collection_key = str(
            document.get(
                "collection_key",
                CLAIM_DEFAULT_COLLECTION_KEY,
            )
            or CLAIM_DEFAULT_COLLECTION_KEY
        ).strip().lower()
        if collection_key == CLAIM_DEFAULT_COLLECTION_KEY:
            return False
        taxpayer_document = taxpayer_returns_by_year.get(
            int(document.get("period_year") or 0)
        )
        if not taxpayer_document:
            return False
        return bool(
            _claim_document_is_no_data(document)
            or not _claim_document_is_no_data(taxpayer_document)
        )

    def business_identity(
        document: dict[str, Any],
    ) -> tuple[str, str]:
        facts = document.get("facts")
        if not isinstance(facts, dict):
            return "", ""
        fingerprint = str(
            facts.get("business_scope_fingerprint", "") or ""
        ).strip().lower()
        if not re.fullmatch(r"s_[0-9a-f]{32}", fingerprint):
            fingerprint = ""
        masked_number = str(
            facts.get("business_number_masked", "") or ""
        ).strip()
        business_name = str(facts.get("business_name", "") or "").strip()
        label = (
            f"{business_name}|{masked_number}"
            if masked_number or business_name
            else ""
        )
        return fingerprint, label

    management_scoped_fingerprints: dict[
        tuple[str, int], set[str]
    ] = {}
    management_scoped_labels: dict[tuple[str, int], set[str]] = {}
    for document in documents:
        document_code = str(document.get("document_code", "")).strip()
        if document_code not in {
            "comwel_total_remuneration",
            "comwel_workplace_rate",
        }:
            continue
        facts = document.get("facts")
        if not (
            isinstance(facts, dict)
            and str(facts.get("management_number_masked", "") or "").strip()
        ):
            continue
        year = int(document.get("period_year") or 0)
        group = (document_code, year)
        fingerprint, label = business_identity(document)
        if fingerprint:
            management_scoped_fingerprints.setdefault(group, set()).add(
                fingerprint
            )
        if label:
            management_scoped_labels.setdefault(group, set()).add(label)

    def is_superseded_empty_management_scope(
        document: dict[str, Any],
    ) -> bool:
        document_code = str(document.get("document_code", "")).strip()
        if document_code not in {
            "comwel_total_remuneration",
            "comwel_workplace_rate",
        }:
            return False
        facts = document.get("facts")
        if isinstance(facts, dict) and str(
            facts.get("management_number_masked", "") or ""
        ).strip():
            return False
        year = int(document.get("period_year") or 0)
        group = (document_code, year)
        fingerprint, label = business_identity(document)
        if fingerprint:
            return fingerprint in management_scoped_fingerprints.get(
                group,
                set(),
            )
        return bool(
            label and label in management_scoped_labels.get(group, set())
        )

    return [
        document
        for document in documents
        if automatic_collection_supported(
            str(document.get("document_code", "")).strip()
        )
        and not is_superseded_default(document)
        and not is_superseded_income_tax_business_scope(document)
        and not is_superseded_empty_management_scope(document)
    ]


def _claim_collection_progress(
    documents: list[dict[str, Any]],
) -> tuple[int, str, int, int]:
    targets = _claim_active_collection_documents(documents)
    target_count = len(targets)
    no_data_count = sum(
        1
        for document in targets
        if str(document.get("status", "")).strip().lower() == "ready"
        and _claim_document_is_no_data(document)
        and not _claim_document_is_provider_blocked(document)
        and not _claim_document_needs_recollection(document)
    )
    collected_count = sum(
        1
        for document in targets
        if str(document.get("status", "")).strip().lower() == "ready"
        and not _claim_document_is_no_data(document)
        and not _claim_document_is_provider_blocked(document)
        and not _claim_document_needs_recollection(document)
    )
    blocked_count = sum(
        1
        for document in targets
        if str(document.get("status", "")).strip().lower() == "ready"
        and _claim_document_is_provider_blocked(document)
    )
    processed_count = collected_count + no_data_count
    ready_count = processed_count
    unresolved_count = max(
        0,
        target_count - processed_count - blocked_count,
    )
    if target_count:
        calculated = round((processed_count / target_count) * 100)
        percentage = (
            100
            if processed_count >= target_count
            else min(99, calculated)
        )
    else:
        percentage = 0
    if target_count:
        progress_detail = (
            f"다운로드 파일 {collected_count}건"
            + (
                f" · 기관 조회 결과 없음 {no_data_count}건"
                if no_data_count
                else ""
            )
            + (
                f" · API 미호출/선행정보 부족 {blocked_count}건"
                if blocked_count
                else ""
            )
            + (
                f" · 실패·대기 {unresolved_count}건"
                if unresolved_count
                else ""
            )
        )
        if processed_count >= target_count:
            progress_text = (
                f"{percentage}% · 자동수집 대상 {target_count}건 처리 완료 "
                f"({progress_detail})"
            )
        else:
            progress_text = (
                f"{percentage}% · 자동수집 대상 {target_count}건 중 "
                f"{processed_count}건 처리 완료 ({progress_detail})"
            )
    else:
        progress_text = "0% · 수집할 자동연동 자료를 확인하고 있습니다."
    return (
        percentage,
        progress_text,
        ready_count,
        target_count,
    )


def _claim_collection_progress_from_repository(
    repository: Any,
    case_id: str,
) -> tuple[int, str, int, int, bool]:
    try:
        documents = repository.list_documents(case_id)
        if isinstance(documents, list):
            percentage, text, ready_count, target_count = (
                _claim_collection_progress(documents)
            )
            return (
                percentage,
                text,
                ready_count,
                target_count,
                True,
            )
    except Exception:
        # 진행률 조회 실패가 실제 서류 수집을 중단시키면 안 된다.
        pass
    return (
        0,
        "0% · Supabase의 실제 수집 자료를 확인하고 있습니다.",
        0,
        0,
        False,
    )


def _claim_progress(
    case: dict[str, Any],
    documents: list[dict[str, Any]] | None = None,
) -> tuple[int, str]:
    if documents is not None:
        percentage, text, _, _ = _claim_collection_progress(documents)
        return percentage, text
    return 0, "0% · 실제 수집 자료를 확인한 뒤 진행률을 표시합니다."


def _render_claim_auth_stage(case: dict[str, Any]) -> None:
    stage, _ = _claim_auth_stage(case)
    overall_status = str(case.get("overall_status", "") or "").strip()
    hometax_status = str(case.get("hometax_status", "") or "").strip()
    comwel_status = str(case.get("comwel_status", "") or "").strip()
    failed_stage = (
        1
        if hometax_status == "failed"
        else 2
        if hometax_status == "auth_complete" and comwel_status == "failed"
        else 0
    )
    collection_finished = overall_status in {"ready", "collected"}
    rows: list[str] = []
    for index, message in enumerate(CLAIM_AUTH_STAGE_MESSAGES, start=1):
        if index == failed_stage:
            state = "실패"
            background = "#fff1f1"
            border = "#f2b8b8"
            color = "#b42318"
        elif index < stage or (
            index == 4 and stage == 4 and collection_finished
        ):
            state = "완료"
            background = "#eef8f2"
            border = "#b9e2c8"
            color = "#18733d"
        elif index == stage:
            state = "진행"
            background = "#eef5ff"
            border = "#b9d3ff"
            color = "#155dcc"
        else:
            state = "대기"
            background = "#f7f8fa"
            border = "#e1e5eb"
            color = "#758096"
        rows.append(
            (
                f'<div style="display:flex;align-items:center;gap:.7rem;'
                f'padding:.62rem .75rem;border:1px solid {border};'
                f'border-radius:10px;background:{background};">'
                f'<span style="font-weight:800;color:{color};'
                f'min-width:1.8rem;">{index:02d}</span>'
                f'<span style="flex:1;color:#17335f;">'
                f'{html.escape(message)}</span>'
                f'<span style="font-size:.78rem;font-weight:700;'
                f'color:{color};">{state}</span></div>'
            )
        )
    st.markdown(
        '<div style="display:grid;gap:.45rem;margin:.2rem 0 1rem;">'
        + "".join(rows)
        + "</div>",
        unsafe_allow_html=True,
    )


def _next_auth_action(
    case: dict[str, Any],
    transient: dict[str, Any],
) -> tuple[str, str]:
    hometax_status = str(case.get("hometax_status", "") or "")
    comwel_status = str(case.get("comwel_status", "") or "")
    has_hometax_session = bool(transient.get("hometax"))
    has_comwel_session = bool(transient.get("comwel"))

    if hometax_status != "auth_complete" and has_hometax_session:
        return (
            "check_hometax",
            "홈택스 인증 확인 후 근로복지공단 발송",
        )
    if hometax_status == "auth_complete" and comwel_status != "auth_complete":
        if has_comwel_session:
            return (
                "check_comwel",
                "근로복지공단 인증 확인 및 자료수집",
            )
        return "request_comwel", "근로복지공단 카카오 인증 발송"
    if (
        hometax_status == "auth_complete"
        and comwel_status == "auth_complete"
        and has_hometax_session
    ):
        return "collect", "홈택스 서류 다시 수집"
    return "", ""


def _safe_provider_error_code(
    exc: Exception,
    source: str = "HOMETAX",
) -> str:
    prefix = re.sub(r"[^A-Z0-9_]", "", str(source or "").upper())
    if not prefix:
        prefix = "DOCUMENT"
    match = re.search(
        r"(?:오류코드|TargetCode|ErrorCode)\s*[:：]\s*([A-Za-z0-9_-]+)",
        str(exc),
        flags=re.IGNORECASE,
    )
    if match:
        return f"{prefix}_{match.group(1).upper()}"[:80]
    return f"{prefix}_DOCUMENT_COLLECTION_FAILED"[:80]


def _provider_error_code(exc: Exception) -> str:
    explicit = str(getattr(exc, "error_code", "") or "").strip().upper()
    if explicit:
        return explicit
    match = re.search(
        r"(?:오류코드|TargetCode|ErrorCode)\s*[:：]\s*([A-Za-z0-9_-]+)",
        str(exc),
        flags=re.IGNORECASE,
    )
    return match.group(1).upper() if match else ""


def _is_transient_auth_error(exc: Exception) -> bool:
    return _provider_error_code(exc) in {"OACX_NO_USER"}


def _document_label(document_code: str) -> str:
    code = str(document_code or "").strip()
    for spec in DOCUMENT_SPECS:
        if spec.code == code:
            return spec.name
    return code or "서류"


def _auth_poll_delay(transient: dict[str, Any]) -> float:
    stage_started_at = float(
        transient.get("stage_started_at", time.time()) or time.time()
    )
    elapsed = time.time() - stage_started_at
    if elapsed <= AUTH_FAST_POLL_WINDOW_SECONDS:
        return AUTH_POLL_SECONDS
    if elapsed <= AUTH_MEDIUM_POLL_WINDOW_SECONDS:
        return AUTH_MEDIUM_POLL_SECONDS
    return AUTH_SLOW_POLL_SECONDS


def _claim_absolute_expiry(transient: dict[str, Any]) -> float:
    absolute_expires_at = float(
        transient.get("absolute_expires_at", 0) or 0
    )
    if absolute_expires_at > 0:
        return absolute_expires_at
    request_started_at = float(
        transient.get("request_started_at", time.time()) or time.time()
    )
    absolute_expires_at = request_started_at + COLLECTION_TTL_SECONDS
    transient["request_started_at"] = request_started_at
    transient["absolute_expires_at"] = absolute_expires_at
    return absolute_expires_at


def _set_claim_expiry(
    transient: dict[str, Any],
    ttl_seconds: float,
) -> float:
    expires_at = min(
        time.time() + max(float(ttl_seconds), 0.0),
        _claim_absolute_expiry(transient),
    )
    transient["expires_at"] = expires_at
    return expires_at


def _ensure_claim_operation_active(
    should_continue: Any | None,
) -> None:
    if callable(should_continue) and not bool(should_continue()):
        raise ClaimProviderError(
            "인증 유효시간이 지나 임시 인증정보를 삭제했습니다.",
            error_code="AUTH_SESSION_EXPIRED",
        )


def _masked_business_no(value: Any) -> str:
    digits = _digits(value)
    if len(digits) != 10:
        return ""
    return f"{digits[:3]}-**-*****"


def _masked_business_choice_no(value: Any) -> str:
    digits = _digits(value)
    if len(digits) != 10:
        return ""
    return f"{digits[:3]}-**-***{digits[-2:]}"


def _business_candidate_token(case_id: str, business_number: str) -> str:
    payload = (
        f"{str(case_id or '').strip()}|{_digits(business_number)}"
    ).encode("utf-8")
    return hmac.new(
        _CLAIM_BUSINESS_TOKEN_KEY,
        payload,
        hashlib.sha256,
    ).hexdigest()[:24]


def _claim_collection_variant_key(
    case_id: str,
    scope: str,
    *values: Any,
) -> str:
    secret = _claim_collection_hmac_secret()
    payload = "|".join(
        (
            str(case_id or "").strip(),
            str(scope or "").strip().lower(),
            *(_digits(value) for value in values),
        )
    ).encode("utf-8")
    return f"v_{hmac.new(secret, payload, hashlib.sha256).hexdigest()[:32]}"


def _claim_collection_hmac_secret() -> bytes:
    for name in (
        "CLAIM_DOCUMENT_VARIANT_KEY",
        "CLAIM_JOB_ENCRYPTION_KEY",
    ):
        value = str(os.environ.get(name, "") or "").strip()
        if not value:
            try:
                value = str(st.secrets.get(name, "") or "").strip()
            except Exception:
                value = ""
        if not value:
            continue
        if len(value) < 32:
            raise ClaimRepositoryError(
                f"{name} must contain at least 32 characters."
            )
        return value.encode("utf-8")
    raise ClaimRepositoryError(
        "A durable claim document variant secret is required."
    )


def _claim_collection_scope_fingerprint(
    case_id: str,
    scope: str,
    *values: Any,
) -> str:
    payload = "|".join(
        (
            str(case_id or "").strip(),
            str(scope or "").strip().lower(),
            *(_digits(value) for value in values),
        )
    ).encode("utf-8")
    digest = hmac.new(
        _claim_collection_hmac_secret(),
        payload,
        hashlib.sha256,
    ).hexdigest()
    return f"s_{digest[:32]}"


def _document_scope_fingerprint(document: Any) -> str:
    facts = document.get("facts") if isinstance(document, dict) else {}
    if not isinstance(facts, dict):
        return ""
    value = str(
        facts.get("collection_scope_fingerprint", "") or ""
    ).strip().lower()
    return value if re.fullmatch(r"s_[0-9a-f]{32}", value) else ""


def _scope_fingerprint_matches(
    document: Any,
    expected_fingerprint: Any,
) -> bool:
    expected = str(expected_fingerprint or "").strip().lower()
    if not re.fullmatch(r"s_[0-9a-f]{32}", expected):
        return False
    stored = _document_scope_fingerprint(document)
    return bool(stored) and hmac.compare_digest(stored, expected)


def _masked_management_number(value: Any) -> str:
    digits = _digits(value)
    if len(digits) < 6:
        return ""
    return f"{digits[:3]}-{'*' * max(3, len(digits) - 3)}"


def _collection_scope_label(
    *,
    business_name: Any = "",
    business_number: Any = "",
    management_number: Any = "",
) -> str:
    parts = [
        part
        for part in (
            _clean(business_name)[:120],
            _masked_business_no(business_number),
            (
                f"관리번호 {_masked_management_number(management_number)}"
                if _masked_management_number(management_number)
                else ""
            ),
        )
        if part
    ]
    return " · ".join(parts)


def _scoped_claim_document(
    document: CollectedClaimDocument,
    *,
    scope_index: int,
    scope_count: int,
    collection_scope: str,
    business_name: str = "",
    business_number: str = "",
    management_number: str = "",
    scope_fingerprint: str = "",
    business_scope_fingerprint: str = "",
) -> CollectedClaimDocument:
    facts = dict(document.facts or {})
    facts.update(
        {
            "collection_scope": str(collection_scope or "").strip(),
            "scope_index": max(1, int(scope_index)),
            "scope_count": max(1, int(scope_count)),
            "collection_scope_fingerprint": str(
                scope_fingerprint or ""
            ).strip(),
        }
    )
    clean_business_name = _clean(business_name)
    if clean_business_name:
        facts["business_name"] = clean_business_name[:120]
    masked_business_number = _masked_business_no(business_number)
    if masked_business_number:
        facts["business_number_masked"] = masked_business_number
    safe_business_scope_fingerprint = str(
        business_scope_fingerprint or ""
    ).strip().lower()
    if re.fullmatch(r"s_[0-9a-f]{32}", safe_business_scope_fingerprint):
        facts["business_scope_fingerprint"] = (
            safe_business_scope_fingerprint
        )
    masked_management_number = _masked_management_number(
        management_number
    )
    if masked_management_number:
        facts["management_number_masked"] = masked_management_number
    scope_label = _collection_scope_label(
        business_name=clean_business_name,
        business_number=business_number,
        management_number=management_number,
    )
    if scope_label:
        facts["collection_scope_label"] = scope_label

    original_name = str(document.file_name or "")
    stem, extension = os.path.splitext(original_name)
    scoped_name = (
        original_name
        if int(scope_count) <= 1
        else (
            f"{stem or 'claim-document'}-scope-"
            f"{max(1, int(scope_index)):02d}{extension}"
        )
    )
    return CollectedClaimDocument(
        content=document.content,
        file_name=scoped_name,
        content_type=document.content_type,
        provider_reference=document.provider_reference,
        facts=facts,
        transient_facts=dict(document.transient_facts or {}),
    )


def _claim_no_data_document(
    *,
    document_code: str,
    year: int | None = None,
    reason: str = "provider_no_records",
) -> CollectedClaimDocument:
    """Build a non-downloadable, PII-free result for a verified empty scope."""
    facts: dict[str, Any] = {
        "no_data": True,
        "no_data_reason": str(reason or "provider_no_records")[:80],
        "record_count": 0,
    }
    if year:
        facts["year"] = str(int(year))
    content = json.dumps(
        facts,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    safe_code = re.sub(
        r"[^a-z0-9_-]+",
        "-",
        str(document_code or "claim-document").strip().lower(),
    ).strip("-")
    suffix = f"-{int(year)}" if year else ""
    return CollectedClaimDocument(
        content=content,
        file_name=f"{safe_code or 'claim-document'}{suffix}-no-data.json",
        content_type="application/json",
        provider_reference="",
        facts=facts,
    )


def _business_collection_scopes(
    case_id: str,
    candidates: list[dict[str, Any]],
    fallback_business_number: str = "",
) -> list[dict[str, str]]:
    by_number: dict[str, dict[str, str]] = {}
    for candidate in candidates:
        if not isinstance(candidate, dict):
            continue
        business_number = _digits(candidate.get("business_number"))
        if not _is_valid_business_no(business_number):
            continue
        current = by_number.get(business_number, {})
        by_number[business_number] = {
            "business_number": business_number,
            "business_name": (
                current.get("business_name")
                or _clean(candidate.get("business_name"))[:120]
            ),
            "business_status": (
                current.get("business_status")
                or _clean(candidate.get("business_status"))[:120]
            ),
        }
    fallback = _digits(fallback_business_number)
    if _is_valid_business_no(fallback) and fallback not in by_number:
        by_number[fallback] = {
            "business_number": fallback,
            "business_name": "",
            "business_status": "",
        }

    scopes: list[dict[str, str]] = []
    for business_number in sorted(by_number):
        candidate = by_number[business_number]
        collection_key = _claim_collection_variant_key(
            case_id,
            "business",
            business_number,
        )
        scopes.append(
            {
                **candidate,
                "collection_key": collection_key,
                "collection_scope_fingerprint": (
                    _claim_collection_scope_fingerprint(
                        case_id,
                        "business",
                        business_number,
                    )
                ),
            }
        )
    return scopes


def _discover_hometax_business_number(
    repository: ClaimRepository,
    client: TilkoClaimClient,
    *,
    case_id: str,
    birth_date: str,
    representative: str,
    cellphone: str,
    session: dict[str, str],
    transient: dict[str, Any],
    on_progress: Any | None = None,
    should_continue: Any | None = None,
) -> dict[str, Any]:
    documents = [
        document
        for document in repository.list_documents(case_id)
        if str(document.get("source", "")) == "hometax"
        and str(document.get("document_code", ""))
        == "hometax_business_registration_list"
    ]
    if not documents:
        return {
            "target": 0,
            "ready": 0,
            "failed": 0,
            "errors": [],
            "business_number": _digits(transient.get("business_number")),
            "candidates": [],
            "selection_required": False,
        }

    existing_ready = any(
        str(document.get("status", "")) == "ready" for document in documents
    )
    raw_candidates = transient.get("business_candidates")
    candidates = (
        [dict(candidate) for candidate in raw_candidates]
        if isinstance(raw_candidates, list)
        else []
    )
    errors: list[dict[str, str]] = []
    ready_count = 1 if existing_ready else 0
    failed_count = 0
    discovery_already_attempted = bool(
        transient.get("hometax_business_discovery_attempted")
    )

    if not candidates and not discovery_already_attempted:
        if on_progress:
            on_progress(0, 1, "hometax_business_registration_list")
        try:
            _ensure_claim_operation_active(should_continue)
            discovery = client.discover_hometax_businesses(
                birth_date=birth_date,
                user_name=representative,
                cellphone=cellphone,
                session=session,
            )
            _ensure_claim_operation_active(should_continue)
            candidates = [
                {
                    "business_number": _digits(candidate.business_number),
                    "business_name": _clean(candidate.business_name),
                    "business_status": _clean(candidate.business_status),
                }
                for candidate in discovery.candidates
                if _is_valid_business_no(candidate.business_number)
            ]
            transient["business_candidates"] = candidates
            if not existing_ready:
                repository.store_collected_document(
                    case_id,
                    document_code="hometax_business_registration_list",
                    document=discovery.document,
                )
                ready_count = 1
            # An empty successful response is also cached for the lifetime of
            # this authenticated job. Otherwise every retry would purchase the
            # same MyBizInfo lookup again.
            transient["hometax_business_discovery_attempted"] = True
        except (ClaimProviderError, ClaimRepositoryError) as exc:
            if isinstance(exc, ClaimProviderError) and (
                _is_transient_auth_error(exc)
                or _provider_error_code(exc) == "AUTH_SESSION_EXPIRED"
            ):
                raise
            # Keep a previous downloadable snapshot intact, but do not count
            # it as a successful discovery for this authenticated run. Without
            # the current raw business set we cannot prove that every business
            # scope was collected.
            failed_count = 1
            ready_count = 0
            safe_error_code = _safe_provider_error_code(
                exc,
                "HOMETAX_BUSINESS_DISCOVERY",
            )
            errors.append(
                {
                    "document_code": "hometax_business_registration_list",
                    "safe_error_code": safe_error_code,
                    "message": str(exc)[:240],
                }
            )
            if not existing_ready:
                try:
                    repository.fail_document(
                        case_id,
                        document_code="hometax_business_registration_list",
                        safe_error_code=safe_error_code,
                    )
                except ClaimRepositoryError:
                    pass
        if on_progress:
            on_progress(1, 1, "hometax_business_registration_list")

    allowed_numbers = {
        _digits(candidate.get("business_number"))
        for candidate in candidates
        if _is_valid_business_no(candidate.get("business_number"))
    }
    requested_number = _digits(
        transient.get("selected_business_number")
        or transient.get("business_number")
    )
    business_number = (
        requested_number
        if _is_valid_business_no(requested_number)
        and (
            not allowed_numbers
            or requested_number in allowed_numbers
        )
        else sorted(allowed_numbers)[0]
        if allowed_numbers
        else ""
    )
    if business_number:
        transient["business_number"] = business_number

    return {
        "target": 1,
        "ready": ready_count,
        "failed": failed_count,
        "errors": errors,
        "business_number": business_number,
        "candidates": candidates,
        "selection_required": False,
    }


def _collect_supported_hometax_documents(
    repository: ClaimRepository,
    client: TilkoClaimClient,
    *,
    case_id: str,
    birth_date: str,
    representative: str,
    cellphone: str,
    identity_number: str = "",
    business_number: str,
    session: dict[str, str],
    businesses: list[dict[str, Any]] | None = None,
    force_tax_number_discovery: bool = False,
    known_ready_keys: set[tuple[Any, ...]] | None = None,
    on_progress: Any | None = None,
    should_continue: Any | None = None,
) -> dict[str, Any]:
    _ensure_claim_operation_active(should_continue)
    hometax_documents = [
        document
        for document in repository.list_documents(case_id)
        if str(document.get("source", "")) == "hometax"
    ]
    planned_hometax_documents = [
        document
        for document in hometax_documents
        if str(
            document.get(
                "collection_key",
                CLAIM_DEFAULT_COLLECTION_KEY,
            )
            or CLAIM_DEFAULT_COLLECTION_KEY
        )
        == CLAIM_DEFAULT_COLLECTION_KEY
    ]
    existing = {
        (
            str(document.get("document_code", "")),
            int(document.get("period_year") or 0),
            str(
                document.get(
                    "collection_key",
                    CLAIM_DEFAULT_COLLECTION_KEY,
                )
                or CLAIM_DEFAULT_COLLECTION_KEY
            ),
        ): document
        for document in hometax_documents
    }
    jobs: list[
        tuple[str, int, str, dict[str, Any], Any]
    ] = []
    skipped_codes: list[str] = [
        (
            f"{document.get('document_code')}:{document.get('period_year')}"
            if document.get("period_year")
            else str(document.get("document_code", ""))
        )
        for document in planned_hometax_documents
        if not automatic_collection_supported(
            str(document.get("document_code", ""))
        )
    ]
    business_scopes = _business_collection_scopes(
        case_id,
        list(businesses or []),
        business_number,
    )
    scope_count = len(business_scopes)

    def add_business_jobs(
        document_code: str,
        period_year: int,
        collector_factory: Any,
    ) -> None:
        for scope_index, scope in enumerate(
            business_scopes,
            start=1,
        ):
            selected_business_number = str(
                scope.get("business_number", "")
            )
            collection_key = str(
                scope.get(
                    "collection_key",
                    CLAIM_DEFAULT_COLLECTION_KEY,
                )
            )
            scope_fingerprint = str(
                scope.get("collection_scope_fingerprint", "")
            )
            scope_facts = {
                "collection_scope": "business",
                "scope_index": scope_index,
                "scope_count": scope_count,
                "collection_scope_fingerprint": scope_fingerprint,
                "business_name": str(
                    scope.get("business_name", "")
                )[:120],
                "business_number_masked": _masked_business_no(
                    selected_business_number
                ),
                "collection_scope_label": _collection_scope_label(
                    business_name=scope.get("business_name", ""),
                    business_number=selected_business_number,
                ),
            }

            def collect_scoped(
                *,
                factory: Any = collector_factory,
                selected_scope: dict[str, str] = scope,
                index: int = scope_index,
            ) -> CollectedClaimDocument:
                return _scoped_claim_document(
                    factory(
                        str(selected_scope.get("business_number", ""))
                    ),
                    scope_index=index,
                    scope_count=scope_count,
                    collection_scope="business",
                    business_name=str(
                        selected_scope.get("business_name", "")
                    ),
                    business_number=str(
                        selected_scope.get("business_number", "")
                    ),
                    scope_fingerprint=str(
                        selected_scope.get(
                            "collection_scope_fingerprint",
                            "",
                        )
                    ),
                )

            jobs.append(
                (
                    document_code,
                    period_year,
                    collection_key,
                    scope_facts,
                    collect_scoped,
                )
            )

    if (
        "hometax_business_registration_certificate",
        0,
        CLAIM_DEFAULT_COLLECTION_KEY,
    ) in existing and business_scopes:
        add_business_jobs(
            "hometax_business_registration_certificate",
            0,
            lambda selected_business_number: (
                client.collect_hometax_business_registration_certificate(
                    birth_date=birth_date,
                    user_name=representative,
                    cellphone=cellphone,
                    business_number=selected_business_number,
                    session=session,
                )
            ),
        )
    elif (
        "hometax_business_registration_certificate",
        0,
        CLAIM_DEFAULT_COLLECTION_KEY,
    ) in existing:
        skipped_codes.append("hometax_business_registration_certificate")

    if (
        "hometax_tax_payment_certificate",
        0,
        CLAIM_DEFAULT_COLLECTION_KEY,
    ) in existing:
        jobs.append(
            (
                "hometax_tax_payment_certificate",
                0,
                CLAIM_DEFAULT_COLLECTION_KEY,
                {},
                lambda: client.collect_hometax_tax_payment_certificate(
                    birth_date=birth_date,
                    user_name=representative,
                    cellphone=cellphone,
                    session=session,
                ),
            )
        )

    for document in planned_hometax_documents:
        document_code = str(document.get("document_code", ""))
        period_year = int(document.get("period_year") or 0)
        if document_code == "hometax_income_tax_help" and period_year:
            jobs.append(
                (
                    document_code,
                    period_year,
                    CLAIM_DEFAULT_COLLECTION_KEY,
                    {},
                    lambda year=period_year: client.collect_hometax_income_tax_help(
                        year=year,
                        birth_date=birth_date,
                        user_name=representative,
                        cellphone=cellphone,
                        session=session,
                    ),
                )
            )
        elif document_code == "hometax_income_tax_return" and period_year:
            taxpayer_number = _digits(identity_number)
            if len(taxpayer_number) == 13:
                jobs.append(
                    (
                        document_code,
                        period_year,
                        CLAIM_DEFAULT_COLLECTION_KEY,
                        {
                            "collection_scope": "taxpayer",
                            "query_strategy": "filing_year_taxpayer_v3",
                        },
                        lambda year=period_year, taxpayer_number=taxpayer_number: client.collect_hometax_income_tax_return(
                            year=year,
                            birth_date=birth_date,
                            user_name=representative,
                            cellphone=cellphone,
                            business_number=taxpayer_number,
                            session=session,
                        ),
                    )
                )
            elif business_scopes:
                add_business_jobs(
                    document_code,
                    period_year,
                    lambda selected_business_number, year=period_year: (
                        client.collect_hometax_income_tax_return(
                            year=year,
                            birth_date=birth_date,
                            user_name=representative,
                            cellphone=cellphone,
                            business_number=selected_business_number,
                            session=session,
                        )
                    ),
                )
            else:
                skipped_codes.append(f"{document_code}:{period_year}")

    if (
        "hometax_closure_certificate",
        0,
        CLAIM_DEFAULT_COLLECTION_KEY,
    ) in existing and business_scopes:
        add_business_jobs(
            "hometax_closure_certificate",
            0,
            lambda selected_business_number: (
                client.collect_hometax_closure_certificate(
                    birth_date=birth_date,
                    user_name=representative,
                    cellphone=cellphone,
                    business_number=selected_business_number,
                    session=session,
                )
            ),
        )
    elif (
        "hometax_closure_certificate",
        0,
        CLAIM_DEFAULT_COLLECTION_KEY,
    ) in existing:
        skipped_codes.append("hometax_closure_certificate")

    jobs.sort(key=lambda item: (item[0], -item[1], item[2]))
    completed_count = 0
    failed_count = 0
    errors: list[dict[str, str]] = []
    transient_business_numbers: list[str] = []
    ready_keys: set[tuple[str, int, str]] = set()
    for raw_key in known_ready_keys or set():
        if len(raw_key) < 2:
            continue
        ready_keys.add(
            (
                str(raw_key[0]),
                int(raw_key[1] or 0),
                (
                    str(raw_key[2])
                    if len(raw_key) >= 3 and raw_key[2]
                    else CLAIM_DEFAULT_COLLECTION_KEY
                ),
            )
        )
    tax_number_discovery_attempted = False
    target_count = len(jobs)
    for (
        document_code,
        period_year,
        collection_key,
        scope_facts,
        collector,
    ) in jobs:
        key = (document_code, period_year, collection_key)
        current = existing.get(key)
        current_ready = bool(
            current and str(current.get("status", "")) == "ready"
        )
        expected_scope_fingerprint = str(
            scope_facts.get("collection_scope_fingerprint", "")
        )
        scope_matches = bool(
            expected_scope_fingerprint
            and _scope_fingerprint_matches(
                current,
                expected_scope_fingerprint,
            )
        )
        scoped_job = bool(expected_scope_fingerprint)
        preexisting_ready = bool(
            current_ready
            if scoped_job
            else (key in ready_keys or current_ready)
        )
        number_discovery_refresh = bool(
            force_tax_number_discovery
            and document_code == "hometax_tax_payment_certificate"
        )
        stale_income_return_refresh = bool(
            current_ready
            and document_code == "hometax_income_tax_return"
            and current
            and _claim_document_needs_recollection(current)
        )
        force_refresh = bool(
            number_discovery_refresh or stale_income_return_refresh
        )
        if (
            preexisting_ready
            and (not scoped_job or scope_matches)
            and not force_refresh
        ):
            completed_count += 1
            ready_keys.add(key)
            if on_progress:
                on_progress(completed_count, target_count, document_code)
            continue
        if on_progress:
            on_progress(
                completed_count + failed_count,
                target_count,
                document_code,
            )
        try:
            _ensure_claim_operation_active(should_continue)
            collected = collector()
            _ensure_claim_operation_active(should_continue)
            if document_code == "hometax_tax_payment_certificate":
                business_numbers = collected.transient_facts.get(
                    "business_numbers"
                )
                if isinstance(business_numbers, list):
                    for candidate in business_numbers:
                        digits = _digits(candidate)
                        if (
                            _is_valid_business_no(digits)
                            and digits not in transient_business_numbers
                        ):
                            transient_business_numbers.append(digits)
                tax_number_discovery_attempted = True
            # A forced number-only refresh must never overwrite a previously
            # downloadable certificate. It is used only for transient facts.
            if not (number_discovery_refresh and preexisting_ready):
                store_kwargs: dict[str, Any] = {
                    "document_code": document_code,
                    "document": collected,
                }
                if period_year:
                    store_kwargs["period_year"] = period_year
                if collection_key != CLAIM_DEFAULT_COLLECTION_KEY:
                    store_kwargs["collection_key"] = collection_key
                repository.store_collected_document(case_id, **store_kwargs)
            ready_keys.add(key)
            completed_count += 1
        except (ClaimProviderError, ClaimRepositoryError) as exc:
            if isinstance(exc, ClaimProviderError):
                if (
                    _is_transient_auth_error(exc)
                    or _provider_error_code(exc) == "AUTH_SESSION_EXPIRED"
                ):
                    raise
            safe_error_code = _safe_provider_error_code(exc)
            errors.append(
                {
                    "document_code": document_code,
                    "period_year": str(period_year or ""),
                    "safe_error_code": safe_error_code,
                    "message": str(exc)[:240],
                }
            )
            if preexisting_ready and (
                force_refresh
                or (scoped_job and not scope_matches)
            ):
                # The previous file remains downloadable, but it does not
                # satisfy the current scope. Count this run as failed without
                # downgrading or deleting the preserved file.
                failed_count += 1
            else:
                failed_count += 1
                try:
                    fail_kwargs: dict[str, Any] = {
                        "document_code": document_code,
                        "safe_error_code": safe_error_code,
                    }
                    if period_year:
                        fail_kwargs["period_year"] = period_year
                    if collection_key != CLAIM_DEFAULT_COLLECTION_KEY:
                        fail_kwargs["collection_key"] = collection_key
                        fail_kwargs["facts"] = scope_facts
                    repository.fail_document(case_id, **fail_kwargs)
                except ClaimRepositoryError:
                    pass
        if on_progress:
            on_progress(completed_count + failed_count, target_count, document_code)
    return {
        "target": target_count,
        "ready": completed_count,
        "failed": failed_count,
        "skipped": skipped_codes,
        "errors": errors,
        "business_numbers": transient_business_numbers,
        "ready_codes": sorted({code for code, _, _ in ready_keys}),
        "ready_keys": sorted(ready_keys),
        "tax_number_discovery_attempted": tax_number_discovery_attempted,
    }


def _collect_supported_comwel_documents(
    repository: ClaimRepository,
    client: TilkoClaimClient,
    *,
    case_id: str,
    identity_number: str,
    representative: str,
    cellphone: str,
    business_number: str,
    session: dict[str, str],
    businesses: list[dict[str, Any]] | None = None,
    management_cache: dict[str, list[str]] | None = None,
    selected_management_number: str = "",
    on_progress: Any | None = None,
    should_continue: Any | None = None,
) -> dict[str, Any]:
    del selected_management_number
    _ensure_claim_operation_active(should_continue)
    documents = [
        document
        for document in repository.list_documents(case_id)
        if str(document.get("source", "")) == "comwel"
    ]
    planned_documents = [
        document
        for document in documents
        if str(
            document.get(
                "collection_key",
                CLAIM_DEFAULT_COLLECTION_KEY,
            )
            or CLAIM_DEFAULT_COLLECTION_KEY
        )
        == CLAIM_DEFAULT_COLLECTION_KEY
    ]
    existing = {
        (
            str(document.get("document_code", "")),
            int(document.get("period_year") or 0),
            str(
                document.get(
                    "collection_key",
                    CLAIM_DEFAULT_COLLECTION_KEY,
                )
                or CLAIM_DEFAULT_COLLECTION_KEY
            ),
        ): document
        for document in documents
    }
    remuneration_years = sorted(
        {
            int(document.get("period_year") or 0)
            for document in planned_documents
            if str(document.get("document_code", ""))
            == "comwel_total_remuneration"
            and int(document.get("period_year") or 0) > 0
        },
        reverse=True,
    )
    rate_years = sorted(
        {
            int(document.get("period_year") or 0)
            for document in planned_documents
            if str(document.get("document_code", ""))
            == "comwel_workplace_rate"
            and int(document.get("period_year") or 0) > 0
        },
        reverse=True,
    )
    business_scopes = _business_collection_scopes(
        case_id,
        list(businesses or []),
        business_number,
    )
    has_management_document = (
        "comwel_management_number_list",
        0,
        CLAIM_DEFAULT_COLLECTION_KEY,
    ) in existing
    safe_cache = (
        management_cache
        if isinstance(management_cache, dict)
        else {}
    )
    management_number_count = 0
    completed_count = 0
    failed_count = 0
    errors: list[dict[str, str]] = []
    skipped_codes: list[str] = []
    target_count = 0
    if not business_scopes:
        target_count = len(remuneration_years)
        skipped_codes.extend(
            (
                ["comwel_management_number_list"]
                if has_management_document
                else []
            )
            + [
                f"comwel_workplace_rate:{year}"
                for year in rate_years
            ]
        )

    def report(document_code: str) -> None:
        if on_progress:
            on_progress(
                completed_count + failed_count,
                max(target_count, 1),
                document_code,
            )

    def store_one(
        document_code: str,
        collector: Any,
        *,
        period_year: int | None = None,
        collection_key: str = CLAIM_DEFAULT_COLLECTION_KEY,
        scope_facts: dict[str, Any] | None = None,
        refresh_ready: bool = False,
    ) -> Any | None:
        nonlocal completed_count, failed_count
        key = (
            document_code,
            int(period_year or 0),
            collection_key,
        )
        current = existing.get(key)
        preexisting_ready = bool(
            current and str(current.get("status", "")) == "ready"
        )
        expected_scope_fingerprint = str(
            (scope_facts or {}).get(
                "collection_scope_fingerprint",
                "",
            )
        )
        scoped_job = bool(expected_scope_fingerprint)
        scope_matches = bool(
            scoped_job
            and _scope_fingerprint_matches(
                current,
                expected_scope_fingerprint,
            )
        )
        scope_refresh = bool(
            preexisting_ready and scoped_job and not scope_matches
        )
        if (
            preexisting_ready
            and (not scoped_job or scope_matches)
            and not refresh_ready
        ):
            completed_count += 1
            report(document_code)
            return current
        report(document_code)
        try:
            _ensure_claim_operation_active(should_continue)
            collected = collector()
            _ensure_claim_operation_active(should_continue)
            if (
                preexisting_ready
                and refresh_ready
                and not scope_refresh
                and not _claim_document_is_no_data(current)
            ):
                stored = current
            else:
                store_kwargs: dict[str, Any] = {
                    "document_code": document_code,
                    "document": collected,
                    "period_year": period_year,
                }
                if collection_key != CLAIM_DEFAULT_COLLECTION_KEY:
                    store_kwargs["collection_key"] = collection_key
                stored = repository.store_collected_document(
                    case_id,
                    **store_kwargs,
                )
            completed_count += 1
            report(document_code)
            return {
                **dict(stored or {}),
                "facts": dict(collected.facts or {}),
            }
        except (ClaimProviderError, ClaimRepositoryError) as exc:
            if isinstance(exc, ClaimProviderError):
                if (
                    _is_transient_auth_error(exc)
                    or _provider_error_code(exc) == "AUTH_SESSION_EXPIRED"
                ):
                    raise
            safe_error_code = _safe_provider_error_code(exc, "COMWEL")
            errors.append(
                {
                    "document_code": document_code,
                    "period_year": str(period_year or ""),
                    "safe_error_code": safe_error_code,
                    "message": str(exc)[:240],
                }
            )
            if preexisting_ready:
                # Preserve the old downloadable file, but the current refresh
                # did not produce a verified document for this scope.
                failed_count += 1
                report(document_code)
                return current
            failed_count += 1
            try:
                fail_kwargs: dict[str, Any] = {
                    "document_code": document_code,
                    "period_year": period_year,
                    "safe_error_code": safe_error_code,
                }
                if collection_key != CLAIM_DEFAULT_COLLECTION_KEY:
                    fail_kwargs["collection_key"] = collection_key
                    fail_kwargs["facts"] = dict(scope_facts or {})
                repository.fail_document(
                    case_id,
                    **fail_kwargs,
                )
            except ClaimRepositoryError:
                pass
            report(document_code)
            return None

    data_scopes: list[dict[str, str]] = []
    if business_scopes:
        if has_management_document:
            target_count = len(business_scopes)
            for business_index, business_scope in enumerate(
                business_scopes,
                start=1,
            ):
                selected_business_number = str(
                    business_scope.get("business_number", "")
                )
                business_collection_key = str(
                    business_scope.get(
                        "collection_key",
                        CLAIM_DEFAULT_COLLECTION_KEY,
                    )
                )
                business_scope_fingerprint = str(
                    business_scope.get(
                        "collection_scope_fingerprint",
                        "",
                    )
                )
                cached_numbers = [
                    _digits(number)
                    for number in safe_cache.get(
                        business_collection_key,
                        [],
                    )
                    if _digits(number)
                ]
                scope_facts = {
                    "collection_scope": "business",
                    "scope_index": business_index,
                    "scope_count": len(business_scopes),
                    "collection_scope_fingerprint": (
                        business_scope_fingerprint
                    ),
                    "business_name": str(
                        business_scope.get("business_name", "")
                    )[:120],
                    "business_number_masked": _masked_business_no(
                        selected_business_number
                    ),
                    "collection_scope_label": _collection_scope_label(
                        business_name=business_scope.get(
                            "business_name",
                            "",
                        ),
                        business_number=selected_business_number,
                    ),
                }

                def collect_management(
                    *,
                    selected_scope: dict[str, str] = business_scope,
                    index: int = business_index,
                ) -> CollectedClaimDocument:
                    collected = client.collect_comwel_management_numbers(
                        identity_number=identity_number,
                        user_name=representative,
                        cellphone=cellphone,
                        business_number=str(
                            selected_scope.get(
                                "business_number",
                                "",
                            )
                        ),
                        session=session,
                    )
                    raw_management_numbers = (
                        collected.facts.get("management_numbers")
                        if isinstance(collected.facts, dict)
                        else []
                    )
                    collected_with_count = CollectedClaimDocument(
                        content=collected.content,
                        file_name=collected.file_name,
                        content_type=collected.content_type,
                        provider_reference=(
                            collected.provider_reference
                        ),
                        facts={
                            **dict(collected.facts or {}),
                            "management_number_count": len(
                                raw_management_numbers
                                if isinstance(
                                    raw_management_numbers,
                                    list,
                                )
                                else []
                            ),
                        },
                        transient_facts=dict(
                            collected.transient_facts or {}
                        ),
                    )
                    return _scoped_claim_document(
                        collected_with_count,
                        scope_index=index,
                        scope_count=len(business_scopes),
                        collection_scope="business",
                        business_name=str(
                            selected_scope.get("business_name", "")
                        ),
                        business_number=str(
                            selected_scope.get("business_number", "")
                        ),
                        scope_fingerprint=str(
                            selected_scope.get(
                                "collection_scope_fingerprint",
                                "",
                            )
                        ),
                    )

                management_error_count = len(errors)
                management_document = store_one(
                    "comwel_management_number_list",
                    collect_management,
                    collection_key=business_collection_key,
                    scope_facts=scope_facts,
                    refresh_ready=not bool(cached_numbers),
                )
                management_lookup_failed = bool(
                    len(errors) > management_error_count
                    and not cached_numbers
                )
                management_facts = (
                    management_document.get("facts")
                    if isinstance(management_document, dict)
                    else {}
                )
                numbers = (
                    management_facts.get("management_numbers")
                    if isinstance(management_facts, dict)
                    else []
                )
                discovered_numbers = [
                    _digits(number)
                    for number in (
                        numbers if isinstance(numbers, list) else []
                    )
                    if _digits(number)
                ]
                selected_numbers = sorted(
                    set(cached_numbers + discovered_numbers)
                )
                management_lookup_failed = bool(
                    management_lookup_failed and not selected_numbers
                )
                safe_cache[business_collection_key] = selected_numbers
                management_number_count += len(selected_numbers)
                for management_number in selected_numbers or [""]:
                    data_scopes.append(
                        {
                            **business_scope,
                            "management_number": management_number,
                            "management_lookup_failed": (
                                management_lookup_failed
                            ),
                        }
                    )
        else:
            data_scopes = [
                {**business_scope, "management_number": ""}
                for business_scope in business_scopes
            ]
    else:
        data_scopes = []

    if business_scopes:
        target_count += len(data_scopes) * len(remuneration_years)
        for scope_index, data_scope in enumerate(
            data_scopes,
            start=1,
        ):
            selected_business_number = str(
                data_scope.get("business_number", "")
            )
            management_number = str(
                data_scope.get("management_number", "")
            )
            collection_key = _claim_collection_variant_key(
                case_id,
                "management",
                selected_business_number,
                management_number or "0",
            )
            scope_fingerprint = _claim_collection_scope_fingerprint(
                case_id,
                "management",
                selected_business_number,
                management_number or "0",
            )
            scope_facts = {
                "collection_scope": (
                    "management" if management_number else "business"
                ),
                "scope_index": scope_index,
                "scope_count": len(data_scopes),
                "collection_scope_fingerprint": scope_fingerprint,
                "business_scope_fingerprint": str(
                    data_scope.get("collection_scope_fingerprint", "")
                ),
                "business_name": str(
                    data_scope.get("business_name", "")
                )[:120],
                "business_number_masked": _masked_business_no(
                    selected_business_number
                ),
                "management_number_masked": (
                    _masked_management_number(management_number)
                ),
                "collection_scope_label": _collection_scope_label(
                    business_name=data_scope.get("business_name", ""),
                    business_number=selected_business_number,
                    management_number=management_number,
                ),
            }

            for year in remuneration_years:
                store_one(
                    "comwel_total_remuneration",
                    lambda year=year, selected_business_number=selected_business_number, management_number=management_number, scope_index=scope_index, data_scope=data_scope: _scoped_claim_document(
                        client.collect_comwel_total_remuneration(
                            year=year,
                            identity_number=identity_number,
                            user_name=representative,
                            cellphone=cellphone,
                            business_number=selected_business_number,
                            management_number=management_number,
                            session=session,
                        ),
                        scope_index=scope_index,
                        scope_count=len(data_scopes),
                        collection_scope=(
                            "management" if management_number else "business"
                        ),
                        business_name=str(
                            data_scope.get("business_name", "")
                        ),
                        business_number=selected_business_number,
                        management_number=management_number,
                        scope_fingerprint=scope_fingerprint,
                        business_scope_fingerprint=str(
                            data_scope.get(
                                "collection_scope_fingerprint",
                                "",
                            )
                        ),
                    ),
                    period_year=year,
                    collection_key=collection_key,
                    scope_facts=scope_facts,
                )
        # Tilko 공식 계약에서 관리번호는 법인인 경우 필수이고 개인은
        # 선택값입니다. 관리번호를 찾지 못했더라도 개인 간편인증 요청은
        # 실제 기관 API까지 보내며, API 응답이 빈 경우에만 no_data로
        # 확정합니다.
        rate_scopes = list(data_scopes)
        target_count += len(rate_scopes) * len(rate_years)
        for scope_index, data_scope in enumerate(
            rate_scopes,
            start=1,
        ):
            selected_business_number = str(
                data_scope.get("business_number", "")
            )
            management_number = str(
                data_scope.get("management_number", "")
            )
            collection_key = _claim_collection_variant_key(
                case_id,
                "workplace-rate",
                selected_business_number,
                management_number or "none",
            )
            scope_fingerprint = _claim_collection_scope_fingerprint(
                case_id,
                "management",
                selected_business_number,
                management_number or "0",
            )
            scope_facts = {
                "collection_scope": (
                    "management" if management_number else "business"
                ),
                "scope_index": scope_index,
                "scope_count": len(rate_scopes),
                "collection_scope_fingerprint": scope_fingerprint,
                "business_scope_fingerprint": str(
                    data_scope.get("collection_scope_fingerprint", "")
                ),
                "business_name": str(
                    data_scope.get("business_name", "")
                )[:120],
                "business_number_masked": _masked_business_no(
                    selected_business_number
                ),
                "management_number_masked": (
                    _masked_management_number(management_number)
                ),
                "collection_scope_label": _collection_scope_label(
                    business_name=data_scope.get("business_name", ""),
                    business_number=selected_business_number,
                    management_number=management_number,
                ),
            }
            for year in rate_years:
                if bool(data_scope.get("management_lookup_failed")):
                    failed_count += 1
                    try:
                        repository.fail_document(
                            case_id,
                            document_code="comwel_workplace_rate",
                            period_year=year,
                            collection_key=collection_key,
                            facts=scope_facts,
                            safe_error_code=(
                                "COMWEL_MANAGEMENT_NUMBER_LOOKUP_FAILED"
                            ),
                        )
                    except ClaimRepositoryError:
                        pass
                    report("comwel_workplace_rate")
                    continue
                collector = (
                    lambda year=year, management_number=management_number, scope_index=scope_index, data_scope=data_scope, selected_business_number=selected_business_number: _scoped_claim_document(
                        client.collect_comwel_workplace_rate(
                            year=year,
                            identity_number=identity_number,
                            user_name=representative,
                            cellphone=cellphone,
                            management_number=management_number,
                            session=session,
                        ),
                        scope_index=scope_index,
                        scope_count=len(data_scopes),
                        collection_scope=(
                            "management" if management_number else "business"
                        ),
                        business_name=str(
                            data_scope.get("business_name", "")
                        ),
                        business_number=selected_business_number,
                        management_number=management_number,
                        scope_fingerprint=scope_fingerprint,
                        business_scope_fingerprint=str(
                            data_scope.get(
                                "collection_scope_fingerprint",
                                "",
                            )
                        ),
                    )
                )
                store_one(
                    "comwel_workplace_rate",
                    collector,
                    period_year=year,
                    collection_key=collection_key,
                    scope_facts=scope_facts,
                )
    else:
        for year in remuneration_years:
            store_one(
                "comwel_total_remuneration",
                lambda year=year: client.collect_comwel_total_remuneration(
                    year=year,
                    identity_number=identity_number,
                    user_name=representative,
                    cellphone=cellphone,
                    business_number="",
                    management_number="",
                    session=session,
                ),
                period_year=year,
            )

    skipped_codes.append("comwel_worker_status:certificate_required")
    return {
        "target": max(0, target_count),
        "ready": completed_count,
        "failed": failed_count,
        "skipped": skipped_codes,
        "errors": errors,
        "management_numbers": [],
        "management_number_count": management_number_count,
        "selection_required": False,
    }


def _collect_case_documents(
    repository: ClaimRepository,
    client: TilkoClaimClient,
    *,
    case_id: str,
    birth_date: str,
    identity_number: str,
    representative: str,
    cellphone: str,
    transient: dict[str, Any],
    on_progress: Any | None = None,
    should_continue: Any | None = None,
) -> dict[str, Any]:
    _ensure_claim_operation_active(should_continue)
    planned_documents = repository.list_documents(case_id)
    estimated_target = max(
        1,
        sum(
            1
            for document in planned_documents
            if automatic_collection_supported(
                str(document.get("document_code", ""))
            )
        ),
    )

    def discovery_progress(
        processed: int,
        _total: int,
        document_code: str,
    ) -> None:
        if on_progress:
            on_progress(processed, estimated_target, document_code)

    business_discovery = _discover_hometax_business_number(
        repository,
        client,
        case_id=case_id,
        birth_date=birth_date,
        representative=representative,
        cellphone=cellphone,
        session=transient["hometax"],
        transient=transient,
        on_progress=discovery_progress,
        should_continue=should_continue,
    )
    business_number = _digits(business_discovery.get("business_number"))
    business_scopes = _business_collection_scopes(
        case_id,
        list(business_discovery.get("candidates", [])),
        business_number,
    )
    if business_scopes:
        transient["business_candidates"] = business_scopes
        business_discovery["candidates"] = business_scopes
    if business_number:
        transient["business_number"] = business_number
    discovery_target = int(business_discovery.get("target", 0) or 0)
    business_valid = _is_valid_business_no(business_number)
    planned_target = estimated_target

    def hometax_progress(
        processed: int,
        _total: int,
        document_code: str,
    ) -> None:
        if on_progress:
            on_progress(
                discovery_target + processed,
                planned_target,
                document_code,
            )

    hometax_summary = _collect_supported_hometax_documents(
        repository,
        client,
        case_id=case_id,
        birth_date=birth_date,
        representative=representative,
        cellphone=cellphone,
        identity_number=identity_number,
        business_number=business_number,
        session=transient["hometax"],
        businesses=business_scopes,
        force_tax_number_discovery=bool(
            not business_valid
            and not business_discovery.get("candidates")
            and not transient.get(
                "hometax_tax_number_discovery_attempted"
            )
        ),
        on_progress=hometax_progress,
        should_continue=should_continue,
    )
    _ensure_claim_operation_active(should_continue)
    if hometax_summary.get("tax_number_discovery_attempted"):
        transient["hometax_tax_number_discovery_attempted"] = True
        transient["hometax_tax_business_numbers"] = list(
            hometax_summary.get("business_numbers", [])
        )

    # MyBizInfo가 개인 납세자 정보만 돌려주는 경우에도, 이미 수집하는
    # 국세납세증명서의 공식 JsonData에서 사업자번호를 메모리 내에서만
    # 보조 확인합니다. 원문 번호는 문서 facts·감사로그에 저장하지 않습니다.
    if (
        not business_valid
        and not business_discovery.get("candidates")
    ):
        fallback_numbers = [
            _digits(candidate)
            for candidate in (
                list(
                    transient.get("hometax_tax_business_numbers", [])
                )
                + list(hometax_summary.get("business_numbers", []))
            )
            if _is_valid_business_no(candidate)
        ]
        fallback_numbers = list(dict.fromkeys(fallback_numbers))
        fallback_candidates = [
            {
                "business_number": candidate,
                "business_name": "",
                "business_status": "",
            }
            for candidate in fallback_numbers
        ]
        if fallback_candidates:
            business_number = sorted(fallback_numbers)[0]
            business_scopes = _business_collection_scopes(
                case_id,
                fallback_candidates,
                business_number,
            )
            business_discovery["candidates"] = business_scopes
            transient["business_candidates"] = business_scopes
            business_discovery["business_number"] = business_number
            business_discovery["selection_required"] = False
            business_valid = _is_valid_business_no(business_number)
            if business_valid:
                transient["business_number"] = business_number
                # 첫 호출에서 저장한 국세납세증명서는 재사용하고,
                # 사업자번호가 필요한 사업자등록증명원만 이어서 수집합니다.
                hometax_summary = _collect_supported_hometax_documents(
                    repository,
                    client,
                    case_id=case_id,
                    birth_date=birth_date,
                    representative=representative,
                    cellphone=cellphone,
                    identity_number=identity_number,
                    business_number=business_number,
                    session=transient["hometax"],
                    businesses=business_scopes,
                    known_ready_keys={
                        (
                            str(key[0]),
                            int(key[1] or 0),
                            (
                                str(key[2])
                                if len(key) >= 3 and key[2]
                                else CLAIM_DEFAULT_COLLECTION_KEY
                            ),
                        )
                        for key in hometax_summary.get(
                            "ready_keys",
                            [],
                        )
                        if len(key) >= 2 and str(key[0])
                    },
                    on_progress=hometax_progress,
                    should_continue=should_continue,
                )
                _ensure_claim_operation_active(should_continue)

    business_scopes = _business_collection_scopes(
        case_id,
        list(business_discovery.get("candidates", [])),
        business_number,
    )
    business_selection_required = False
    business_valid = bool(business_scopes)
    business_dependent_documents = [
        document
        for document in planned_documents
        if str(document.get("document_code", ""))
        in {
            "hometax_business_registration_certificate",
            "hometax_income_tax_return",
            "hometax_closure_certificate",
            "comwel_management_number_list",
            "comwel_workplace_rate",
        }
        and str(document.get("status", "")) != "ready"
    ]
    comwel_business_dependent_documents = [
        document
        for document in business_dependent_documents
        if str(document.get("source", "")) == "comwel"
    ]
    business_number_missing = bool(
        not business_valid
        and business_dependent_documents
        and not business_selection_required
    )
    business_blocked_count = (
        len(comwel_business_dependent_documents)
        if business_number_missing
        else 0
    )
    blocked_status_errors: list[dict[str, str]] = []
    if business_number_missing:
        for document in business_dependent_documents:
            status = str(document.get("status", ""))
            facts = document.get("facts")
            safe_error_code = (
                str(facts.get("safe_error_code", ""))
                if isinstance(facts, dict)
                else ""
            )
            if (
                status == "failed"
                and safe_error_code == "BUSINESS_NUMBER_NOT_FOUND"
            ):
                continue
            # Preserve an earlier provider failure rather than rewriting its
            # diagnostic. Only pending/planned rows are converted to an
            # explicit compatible failed status.
            if status not in {"auth_pending", "integration_required"}:
                continue
            try:
                failure_kwargs: dict[str, Any] = {
                    "document_code": str(
                        document.get("document_code", "")
                    ),
                    "safe_error_code": "BUSINESS_NUMBER_NOT_FOUND",
                }
                if document.get("period_year"):
                    failure_kwargs["period_year"] = int(
                        document.get("period_year")
                    )
                repository.fail_document(
                    case_id,
                    **failure_kwargs,
                )
            except ClaimRepositoryError:
                blocked_status_errors.append(
                    {
                        "document_code": str(
                            document.get("document_code", "")
                        ),
                        "safe_error_code": (
                            "BUSINESS_BLOCK_STATUS_SAVE_FAILED"
                        ),
                        "message": (
                            "사업자번호 필요 상태를 저장하지 못했습니다."
                        ),
                    }
                )

    hometax_target = int(hometax_summary.get("target", 0) or 0)

    def comwel_progress(
        processed: int,
        _total: int,
        document_code: str,
    ) -> None:
        if on_progress:
            on_progress(
                discovery_target + hometax_target + processed,
                planned_target,
                document_code,
            )

    management_cache = transient.get(
        "comwel_management_numbers_by_business"
    )
    if not isinstance(management_cache, dict):
        management_cache = {}
        transient["comwel_management_numbers_by_business"] = (
            management_cache
        )
    comwel_summary = _collect_supported_comwel_documents(
        repository,
        client,
        case_id=case_id,
        identity_number=identity_number,
        representative=representative,
        cellphone=cellphone,
        business_number=business_number,
        session=transient["comwel"],
        businesses=business_scopes,
        management_cache=management_cache,
        on_progress=comwel_progress,
        should_continue=should_continue,
    )
    _ensure_claim_operation_active(should_continue)
    safe_business_candidates = [
        {
            "business_name": str(scope.get("business_name", ""))[:120],
            "business_status": str(
                scope.get("business_status", "")
            )[:120],
            "business_number_masked": _masked_business_no(
                scope.get("business_number", "")
            ),
            "collection_scope_label": _collection_scope_label(
                business_name=scope.get("business_name", ""),
                business_number=scope.get("business_number", ""),
            ),
        }
        for scope in business_scopes
    ]
    safe_business_discovery = {
        key: value
        for key, value in business_discovery.items()
        if key not in {"business_number", "candidates"}
    }
    safe_business_discovery["candidates"] = safe_business_candidates
    safe_hometax_summary = {
        key: value
        for key, value in hometax_summary.items()
        if key != "business_numbers"
    }
    summary = {
        "target": discovery_target
        + int(hometax_summary["target"])
        + int(comwel_summary["target"]),
        "ready": int(business_discovery.get("ready", 0) or 0)
        + int(hometax_summary["ready"])
        + int(comwel_summary["ready"]),
        "failed": int(business_discovery.get("failed", 0) or 0)
        + int(hometax_summary["failed"])
        + int(comwel_summary["failed"]),
        "skipped": list(hometax_summary["skipped"])
        + list(comwel_summary["skipped"]),
        "errors": list(business_discovery.get("errors", []))
        + list(hometax_summary["errors"])
        + list(comwel_summary["errors"])
        + blocked_status_errors,
        "sources": {
            "hometax_business_discovery": safe_business_discovery,
            "hometax": safe_hometax_summary,
            "comwel": comwel_summary,
        },
        "management_numbers": list(
            comwel_summary.get("management_numbers", [])
        ),
        "selection_required": bool(
            comwel_summary.get("selection_required")
        ),
        "business_candidates": safe_business_candidates,
        "business_selection_required": business_selection_required,
        "business_number_missing": business_number_missing,
        "business_blocked_count": business_blocked_count,
    }
    if business_number_missing:
        summary["errors"].append(
            {
                "document_code": "hometax_business_registration_list",
                "safe_error_code": "BUSINESS_NUMBER_NOT_FOUND",
                "message": "홈택스에서 유효한 사업자등록번호를 확인하지 못했습니다.",
            }
        )
    collection_complete = bool(
        summary["target"] > 0
        and summary["ready"] == summary["target"]
        and summary["failed"] == 0
        and not summary["errors"]
        and not summary["selection_required"]
        and not business_selection_required
        and not business_number_missing
    )
    summary["complete"] = collection_complete
    _ensure_claim_operation_active(should_continue)
    if summary["business_selection_required"]:
        repository.update_case_status(
            case_id,
            overall_status="auth_complete_collection_pending",
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=case_id,
            action="business_number_selection_required",
            source="hometax",
            outcome="pending",
            metadata={
                "business_candidate_count": len(
                    summary["business_candidates"]
                ),
                "ready_document_count": summary["ready"],
            },
        )
        return summary
    if summary["selection_required"]:
        repository.update_case_status(
            case_id,
            overall_status="auth_complete_collection_pending",
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=case_id,
            action="management_number_selection_required",
            source="comwel",
            outcome="pending",
            metadata={
                "management_number_count": len(
                    summary["management_numbers"]
                ),
                "ready_document_count": summary["ready"],
            },
        )
        return summary
    if collection_complete:
        repository.update_case_status(
            case_id,
            overall_status="ready",
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=case_id,
            action="collection_complete",
            source="provider",
            outcome="success",
            metadata={
                "supported_document_count": summary["target"],
                "ready_document_count": summary["ready"],
                "skipped_document_codes": summary["skipped"],
                "scope": "currently_connected_documents",
            },
        )
        return summary

    first_error = (
        summary["errors"][0]["safe_error_code"]
        if summary["errors"]
        else "DOCUMENT_COLLECTION_FAILED"
    )
    repository.update_case_status(
        case_id,
        overall_status="auth_complete_collection_pending",
        last_safe_error_code=first_error,
    )
    repository.append_audit_event(
        case_id=case_id,
        action="collection_partial",
        source="provider",
        outcome="failed",
        metadata={
            "supported_document_count": summary["target"],
            "ready_document_count": summary["ready"],
            "failed_document_count": summary["failed"],
            "safe_error_codes": [
                item["safe_error_code"] for item in summary["errors"]
            ],
        },
    )
    return summary


def _advance_personal_case(
    repository: ClaimRepository,
    client: TilkoClaimClient,
    *,
    case: dict[str, Any],
    transient: dict[str, Any],
    representative: str,
    cellphone: str,
    birth_date: str,
    identity_number: str,
    on_progress: Any | None = None,
    should_continue: Any | None = None,
) -> dict[str, Any]:
    _ensure_claim_operation_active(should_continue)
    case_id = str(case.get("id", ""))
    action, _ = _next_auth_action(case, transient)
    if not action:
        return {"event": "idle", "action": ""}

    if action == "check_hometax":
        hometax_complete = client.check_hometax_kakao(
            birth_date=birth_date,
            user_name=representative,
            cellphone=cellphone,
            session=transient["hometax"],
        )
        _ensure_claim_operation_active(should_continue)
        if not hometax_complete:
            repository.update_case_status(
                case_id,
                hometax_status="auth_pending",
                overall_status="auth_pending",
            )
            repository.append_audit_event(
                case_id=case_id,
                action="auth_check",
                source="hometax",
                outcome="pending",
                metadata={"next_source_requested": False},
            )
            return {"event": "hometax_pending", "action": action}

        repository.update_case_status(
            case_id,
            hometax_status="auth_complete",
            overall_status="auth_pending",
            last_safe_error_code=None,
        )
        try:
            if not transient.get("comwel"):
                _ensure_claim_operation_active(should_continue)
                if COMWEL_DISPATCH_DELAY_SECONDS > 0:
                    time.sleep(COMWEL_DISPATCH_DELAY_SECONDS)
                _ensure_claim_operation_active(should_continue)
                comwel_session = client.request_comwel_kakao(
                    identity_number=identity_number,
                    user_name=representative,
                    cellphone=cellphone,
                )
                _ensure_claim_operation_active(should_continue)
                transient["comwel"] = comwel_session
        except ClaimProviderError as exc:
            if (
                _is_transient_auth_error(exc)
                or _provider_error_code(exc) == "AUTH_SESSION_EXPIRED"
            ):
                raise
            repository.update_case_status(
                case_id,
                hometax_status="auth_complete",
                comwel_status="failed",
                overall_status="auth_partial",
                last_safe_error_code="COMWEL_AUTH_REQUEST_FAILED",
            )
            repository.append_audit_event(
                case_id=case_id,
                action="auth_request",
                source="comwel",
                outcome="failed",
                metadata={"safe_error_code": "COMWEL_AUTH_REQUEST_FAILED"},
            )
            raise
        _set_claim_expiry(transient, AUTH_TTL_SECONDS)
        transient["stage_started_at"] = time.time()
        repository.update_case_status(
            case_id,
            hometax_status="auth_complete",
            comwel_status="auth_requested",
            overall_status="auth_pending",
            auth_requested_at=datetime.now(timezone.utc).isoformat(),
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=case_id,
            action="auth_check",
            source="hometax",
            outcome="success",
            metadata={"next_source_requested": True},
        )
        return {"event": "comwel_requested", "action": action}

    if action == "request_comwel":
        _ensure_claim_operation_active(should_continue)
        comwel_session = client.request_comwel_kakao(
            identity_number=identity_number,
            user_name=representative,
            cellphone=cellphone,
        )
        _ensure_claim_operation_active(should_continue)
        transient["comwel"] = comwel_session
        _set_claim_expiry(transient, AUTH_TTL_SECONDS)
        transient["stage_started_at"] = time.time()
        repository.update_case_status(
            case_id,
            comwel_status="auth_requested",
            overall_status="auth_pending",
            auth_requested_at=datetime.now(timezone.utc).isoformat(),
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=case_id,
            action="auth_request",
            source="comwel",
            outcome="success",
            metadata={"sequential_after_hometax": True},
        )
        return {"event": "comwel_requested", "action": action}

    if action == "check_comwel":
        comwel_complete = client.check_comwel_kakao(
            identity_number=identity_number,
            user_name=representative,
            cellphone=cellphone,
            session=transient["comwel"],
        )
        _ensure_claim_operation_active(should_continue)
        if not comwel_complete:
            repository.update_case_status(
                case_id,
                hometax_status="auth_complete",
                comwel_status="auth_pending",
                overall_status="auth_pending",
            )
            repository.append_audit_event(
                case_id=case_id,
                action="auth_check",
                source="comwel",
                outcome="pending",
                metadata={"collection_started": False},
            )
            return {"event": "comwel_pending", "action": action}
        repository.update_case_status(
            case_id,
            hometax_status="auth_complete",
            comwel_status="auth_complete",
            overall_status="collecting",
            auth_completed_at=datetime.now(timezone.utc).isoformat(),
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=case_id,
            action="auth_check",
            source="comwel",
            outcome="success",
            metadata={"collection_started": True},
        )
        _set_claim_expiry(transient, COLLECTION_TTL_SECONDS)
    elif action == "collect":
        _ensure_claim_operation_active(should_continue)
        repository.update_case_status(
            case_id,
            overall_status="collecting",
            last_safe_error_code=None,
        )
        _set_claim_expiry(transient, COLLECTION_TTL_SECONDS)

    if on_progress:
        on_progress(0, 1, "collection_preparing")
    _ensure_claim_operation_active(should_continue)
    summary = _collect_case_documents(
        repository,
        client,
        case_id=case_id,
        birth_date=birth_date,
        identity_number=identity_number,
        representative=representative,
        cellphone=cellphone,
        transient=transient,
        on_progress=on_progress,
        should_continue=should_continue,
    )
    event = (
        "business_selection_required"
        if summary.get("business_selection_required")
        else "management_selection_required"
        if summary.get("selection_required")
        else "collection_complete"
        if summary["complete"]
        else "collection_partial"
    )
    return {
        "event": event,
        "action": action,
        "summary": summary,
    }


def _claim_job_owner_ref(user_id: str) -> str:
    return hashlib.sha256(
        str(user_id or "").strip().lower().encode("utf-8")
    ).hexdigest()


def _seal_claim_job_payload(payload: dict[str, Any]) -> bytes:
    serialized = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return _CLAIM_JOB_CIPHER.encrypt(serialized)


def _unseal_claim_job_payload(value: bytes) -> dict[str, Any]:
    try:
        decoded = _CLAIM_JOB_CIPHER.decrypt(value)
        payload = json.loads(decoded.decode("utf-8"))
    except (InvalidToken, ValueError, TypeError, json.JSONDecodeError) as exc:
        raise ClaimProviderError(
            "임시 인증정보를 복구하지 못했습니다. 새 인증 요청을 시작해 주세요."
        ) from exc
    if not isinstance(payload, dict):
        raise ClaimProviderError(
            "임시 인증정보 형식을 확인하지 못했습니다."
        )
    return payload


def _sync_interrupted_claim_case(
    user_id: str,
    case_id: str,
    *,
    active_action: str,
    safe_error_code: str,
    outcome: str,
) -> None:
    owner_user_id = str(user_id or "").strip().lower()
    if not owner_user_id:
        return
    try:
        repository = ClaimRepository(owner_user_id)
        case = repository.get_case(case_id)
        if not case:
            return
        hometax_status = str(case.get("hometax_status", "") or "")
        comwel_status = str(case.get("comwel_status", "") or "")
        updates: dict[str, Any] = {
            "last_safe_error_code": safe_error_code,
        }
        if safe_error_code == "AUTH_SESSION_EXPIRED":
            if hometax_status != "auth_complete":
                updates["hometax_status"] = "failed"
            elif comwel_status != "auth_complete":
                updates["comwel_status"] = "failed"
        elif (
            active_action == "check_hometax"
            and hometax_status != "auth_complete"
        ):
            updates["hometax_status"] = "failed"
        elif (
            active_action in {"request_comwel", "check_comwel"}
            and comwel_status != "auth_complete"
        ):
            updates["comwel_status"] = "failed"

        authentication_complete = (
            hometax_status == "auth_complete"
            and comwel_status == "auth_complete"
        )
        collection_started = (
            active_action == "collect"
            or str(case.get("overall_status", "") or "")
            in {
                "collecting",
                "collection_queued",
                "auth_complete_collection_pending",
            }
        )
        updates["overall_status"] = (
            "auth_complete_collection_pending"
            if authentication_complete or collection_started
            else "auth_partial"
        )
        repository.update_case_status(case_id, **updates)
        repository.append_audit_event(
            case_id=case_id,
            action="background_auth_or_collection",
            source="provider",
            outcome=outcome,
            metadata={"safe_error_code": safe_error_code},
        )
    except ClaimRepositoryError:
        return


def _expire_claim_job(
    case_id: str,
    owner_ref: str,
    user_id: str = "",
) -> None:
    owner_user_id = str(user_id or "").strip().lower()
    previous_status = ""
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if not job or job.get("owner_ref") != owner_ref:
            return
        previous_status = str(job.get("status", "") or "")
        owner_user_id = (
            owner_user_id
            or str(job.get("owner_user_id", "") or "").strip().lower()
        )
        job["sealed_payload"] = b""
        if previous_status != "complete":
            job["status"] = "expired"
            job["safe_message"] = (
                (
                    "일부 서류는 저장했지만 재시도 유효시간이 지나 "
                    "임시 인증정보를 삭제했습니다. 새 인증 요청을 시작해 주세요."
                )
                if previous_status == "collection_partial"
                else (
                    "인증 유효시간이 지나 임시 인증정보를 삭제했습니다. "
                    "새 인증 요청을 시작해 주세요."
                )
            )
        wake_event = job.get("wake_event")
        if isinstance(wake_event, threading.Event):
            wake_event.set()
    if previous_status != "complete":
        _sync_interrupted_claim_case(
            owner_user_id,
            case_id,
            active_action=(
                "collect"
                if previous_status
                in {
                    "collection_partial",
                    "awaiting_business_selection",
                    "awaiting_management_selection",
                }
                else ""
            ),
            safe_error_code="AUTH_SESSION_EXPIRED",
            outcome="expired",
        )


def _update_claim_job(
    case_id: str,
    owner_ref: str,
    **updates: Any,
) -> bool:
    expired_owner_user_id = ""
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if not job or job.get("owner_ref") != owner_ref:
            return False
        current_status = str(job.get("status", "") or "")
        current_expires_at = float(job.get("expires_at", 0) or 0)
        if current_status in {"complete", "expired"}:
            return False
        if (
            current_expires_at <= time.time()
            or not job.get("sealed_payload")
        ):
            job["sealed_payload"] = b""
            job["status"] = "expired"
            job["safe_message"] = (
                "인증 유효시간이 지나 임시 인증정보를 삭제했습니다. "
                "새 인증 요청을 시작해 주세요."
            )
            job["updated_at"] = time.time()
            expired_owner_user_id = str(
                job.get("owner_user_id", "") or ""
            ).strip().lower()
            wake_event = job.get("wake_event")
            if isinstance(wake_event, threading.Event):
                wake_event.set()
        else:
            job.update(updates)
            job["updated_at"] = time.time()
            return True
    _sync_interrupted_claim_case(
        expired_owner_user_id,
        case_id,
        active_action="",
        safe_error_code="AUTH_SESSION_EXPIRED",
        outcome="expired",
    )
    return False


def _claim_job_can_continue(case_id: str, owner_ref: str) -> bool:
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        return bool(
            job
            and job.get("owner_ref") == owner_ref
            and str(job.get("status", "") or "")
            not in {"complete", "expired"}
            and float(job.get("expires_at", 0) or 0) > time.time()
            and job.get("sealed_payload")
        )


def _run_background_claim_job(
    user_id: str,
    case_id: str,
    owner_ref: str,
    *,
    initial_delay: float = 0,
) -> None:
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        wake_event = job.get("wake_event") if job else None
    if not isinstance(wake_event, threading.Event):
        return
    if initial_delay > 0 and wake_event.wait(initial_delay):
        wake_event.clear()

    transient: dict[str, Any] | None = None
    context: dict[str, Any] | None = None
    active_action = ""
    try:
        repository = ClaimRepository(user_id)
        client = TilkoClaimClient()
        while True:
            with _CLAIM_JOB_LOCK:
                job = _CLAIM_JOBS.get(case_id)
                if not job or job.get("owner_ref") != owner_ref:
                    return
                expires_at = float(job.get("expires_at", 0) or 0)
                sealed_payload = job.get("sealed_payload")
            if time.time() >= expires_at or not isinstance(
                sealed_payload,
                bytes,
            ) or not sealed_payload:
                _expire_claim_job(case_id, owner_ref, user_id)
                return

            transient = _unseal_claim_job_payload(sealed_payload)
            context = transient.get("auth_context")
            if not isinstance(context, dict):
                raise ClaimProviderError(
                    "순차 인증에 필요한 고객정보를 확인하지 못했습니다."
                )
            case = repository.get_case(case_id)
            if not case:
                raise ClaimRepositoryError(
                    "경정청구 요청 건을 찾지 못했습니다."
                )

            def update_collection_progress(
                _processed: int,
                _total: int,
                document_code: str,
            ) -> None:
                with _CLAIM_JOB_LOCK:
                    current_job = dict(_CLAIM_JOBS.get(case_id) or {})
                current_summary = dict(current_job.get("summary") or {})
                (
                    percentage,
                    _progress_text,
                    ready_count,
                    target_count,
                    progress_verified,
                ) = _claim_collection_progress_from_repository(
                    repository,
                    case_id,
                )
                if (
                    not progress_verified
                    and current_summary.get("progress_verified") is True
                ):
                    percentage = int(current_job.get("progress", 0) or 0)
                    ready_count = int(
                        current_summary.get("ready", 0) or 0
                    )
                    target_count = int(
                        current_summary.get("target", 0) or 0
                    )
                _set_claim_expiry(
                    transient,
                    COLLECTION_TTL_SECONDS,
                )
                if transient["expires_at"] <= time.time():
                    raise ClaimProviderError(
                        "인증 유효시간이 지나 임시 인증정보를 삭제했습니다.",
                        error_code="AUTH_SESSION_EXPIRED",
                    )
                updated = _update_claim_job(
                    case_id,
                    owner_ref,
                    sealed_payload=_seal_claim_job_payload(transient),
                    expires_at=transient["expires_at"],
                    progress=percentage,
                    status="running",
                    safe_message=(
                        f"{_document_label(document_code)} 수집 중 · "
                        f"{ready_count}/{target_count}건 처리 완료"
                        if target_count
                        else f"{_document_label(document_code)} 수집 중"
                    ),
                    summary={
                        "ready": ready_count,
                        "target": target_count,
                        "progress_verified": progress_verified,
                    },
                )
                if not updated:
                    transient.clear()
                    context.clear()
                    raise ClaimProviderError(
                        "인증 유효시간이 지나 임시 인증정보를 삭제했습니다.",
                        error_code="AUTH_SESSION_EXPIRED",
                    )

            try:
                active_action = _next_auth_action(
                    case,
                    transient,
                )[0]
                result = _advance_personal_case(
                    repository,
                    client,
                    case=case,
                    transient=transient,
                    representative=str(
                        context.get("representative", "")
                    ).strip(),
                    cellphone=_digits(context.get("cellphone")),
                    birth_date=_digits(context.get("birth_date")),
                    identity_number=_digits(context.get("identity_number")),
                    on_progress=update_collection_progress,
                    should_continue=lambda: _claim_job_can_continue(
                        case_id,
                        owner_ref,
                    ),
                )
            except ClaimProviderError as exc:
                if not _is_transient_auth_error(exc):
                    raise
                if not _claim_job_can_continue(case_id, owner_ref):
                    _sync_interrupted_claim_case(
                        user_id,
                        case_id,
                        active_action=active_action,
                        safe_error_code="AUTH_SESSION_EXPIRED",
                        outcome="expired",
                    )
                    transient.clear()
                    context.clear()
                    return
                retry_count = int(
                    transient.get("provider_session_retry_count", 0) or 0
                ) + 1
                transient["provider_session_retry_count"] = retry_count
                updated = _update_claim_job(
                    case_id,
                    owner_ref,
                    sealed_payload=_seal_claim_job_payload(transient),
                    expires_at=float(
                        transient.get("expires_at", time.time()) or time.time()
                    ),
                    status="running",
                    progress=int(job.get("progress", 0) or 0),
                    safe_message=(
                        "인증정보가 중계 서버에 반영되는 중입니다. "
                        "자동으로 다시 확인합니다."
                    ),
                )
                if not updated:
                    transient.clear()
                    context.clear()
                    return
                if retry_count == 1:
                    repository.append_audit_event(
                        case_id=case_id,
                        action="auth_check",
                        source="provider",
                        outcome="pending",
                        metadata={
                            "safe_error_code": _provider_error_code(exc),
                            "automatic_retry": True,
                        },
                    )
                wake_event.wait(
                    max(
                        TRANSIENT_AUTH_RETRY_SECONDS,
                        _auth_poll_delay(transient),
                    )
                )
                wake_event.clear()
                continue
            if not _claim_job_can_continue(case_id, owner_ref):
                _sync_interrupted_claim_case(
                    user_id,
                    case_id,
                    active_action=active_action,
                    safe_error_code="AUTH_SESSION_EXPIRED",
                    outcome="expired",
                )
                transient.clear()
                context.clear()
                return
            event = str(result.get("event", "") or "")
            updated_case = repository.get_case(case_id) or case
            (
                percentage,
                _progress_text,
                ready_count,
                target_count,
                progress_verified,
            ) = _claim_collection_progress_from_repository(
                repository,
                case_id,
            )
            _, auth_stage_message = _claim_auth_stage(updated_case)

            if event == "business_selection_required":
                summary = dict(result.get("summary") or {})
                business_choices = []
                for candidate in summary.get("business_candidates", []):
                    if not isinstance(candidate, dict):
                        continue
                    business_number = _digits(
                        candidate.get("business_number")
                    )
                    if not _is_valid_business_no(business_number):
                        continue
                    business_name = _clean(
                        candidate.get("business_name")
                    )
                    business_status = _clean(
                        candidate.get("business_status")
                    )
                    label_parts = [
                        part
                        for part in (
                            business_name,
                            _masked_business_choice_no(business_number),
                            business_status,
                        )
                        if part
                    ]
                    business_choices.append(
                        {
                            "token": _business_candidate_token(
                                case_id,
                                business_number,
                            ),
                            "label": " · ".join(label_parts),
                        }
                    )
                updated = _update_claim_job(
                    case_id,
                    owner_ref,
                    sealed_payload=_seal_claim_job_payload(transient),
                    expires_at=float(
                        transient.get("expires_at", time.time())
                        or time.time()
                    ),
                    status="awaiting_business_selection",
                    progress=percentage,
                    safe_message=(
                        "홈택스에서 여러 사업자가 확인됐습니다. "
                        "자료를 수집할 사업자를 선택해 주세요."
                    ),
                    summary={
                        "ready": ready_count,
                        "target": target_count,
                        "progress_verified": progress_verified,
                        "failed": int(summary.get("failed", 0) or 0),
                        "business_choices": business_choices,
                    },
                )
                if not updated:
                    transient.clear()
                    context.clear()
                return
            if event == "management_selection_required":
                summary = dict(result.get("summary") or {})
                management_numbers = [
                    str(number)
                    for number in summary.get("management_numbers", [])
                    if str(number).strip()
                ]
                updated = _update_claim_job(
                    case_id,
                    owner_ref,
                    sealed_payload=_seal_claim_job_payload(transient),
                    expires_at=float(
                        transient.get("expires_at", time.time()) or time.time()
                    ),
                    status="awaiting_management_selection",
                    progress=percentage,
                    safe_message=(
                        "사업장관리번호를 선택하면 사업장요율 수집을 "
                        "계속합니다."
                    ),
                    summary={
                        "ready": ready_count,
                        "target": target_count,
                        "progress_verified": progress_verified,
                        "failed": int(summary.get("failed", 0) or 0),
                        "management_numbers": management_numbers,
                    },
                )
                if not updated:
                    transient.clear()
                    context.clear()
                return
            if event == "collection_complete":
                summary = dict(result.get("summary") or {})
                (
                    percentage,
                    _progress_text,
                    ready_count,
                    target_count,
                    progress_verified,
                ) = _claim_collection_progress_from_repository(
                    repository,
                    case_id,
                )
                if (
                    not progress_verified
                    or target_count <= 0
                    or percentage < 100
                ):
                    verification_error_code = (
                        "COLLECTION_PROGRESS_UNVERIFIED"
                        if not progress_verified
                        else "COLLECTION_PROGRESS_INCOMPLETE"
                    )
                    try:
                        repository.update_case_status(
                            case_id,
                            overall_status=(
                                "auth_complete_collection_pending"
                            ),
                            last_safe_error_code=verification_error_code,
                        )
                        repository.append_audit_event(
                            case_id=case_id,
                            action="collection_progress_verification",
                            source="supabase",
                            outcome="pending",
                            metadata={
                                "safe_error_code": (
                                    verification_error_code
                                ),
                                "ready_document_count": ready_count,
                                "target_document_count": target_count,
                            },
                        )
                    except ClaimRepositoryError:
                        pass
                    safe_message = (
                        "Supabase의 실제 수집 상태를 확인하지 못했습니다. "
                        "완료로 처리하지 않고 다시 확인합니다."
                        if not progress_verified
                        else (
                            f"자동수집 대상 {target_count}건 중 "
                            f"{ready_count}건만 확인되어 완료로 처리하지 "
                            "않았습니다."
                        )
                    )
                    _update_claim_job(
                        case_id,
                        owner_ref,
                        sealed_payload=_seal_claim_job_payload(transient),
                        expires_at=float(
                            transient.get("expires_at", time.time())
                            or time.time()
                        ),
                        status="paused",
                        progress=min(99, percentage),
                        safe_message=safe_message,
                        summary={
                            "ready": ready_count,
                            "target": target_count,
                            "progress_verified": progress_verified,
                            "failed": int(
                                summary.get("failed", 0) or 0
                            ),
                        },
                    )
                    return
                _update_claim_job(
                    case_id,
                    owner_ref,
                    sealed_payload=b"",
                    status="complete",
                    progress=percentage,
                    safe_message=(
                        f"자동수집 대상 {target_count}건 처리가 완료되었습니다."
                        + (
                            f", 추가 정보·미지원 "
                            f"{len(summary.get('skipped', []))}건"
                            if summary.get("skipped")
                            else ""
                        )
                    ),
                    summary={
                        "ready": ready_count,
                        "target": target_count,
                        "progress_verified": True,
                        "failed": int(summary.get("failed", 0) or 0),
                        "skipped_count": len(summary.get("skipped", [])),
                    },
                )
                transient.clear()
                context.clear()
                return
            if event == "collection_partial":
                summary = dict(result.get("summary") or {})
                business_number_missing = bool(
                    summary.get("business_number_missing")
                )
                blocked_count = int(
                    summary.get("business_blocked_count", 0) or 0
                )
                (
                    percentage,
                    _progress_text,
                    ready_count,
                    target_count,
                    progress_verified,
                ) = _claim_collection_progress_from_repository(
                    repository,
                    case_id,
                )
                percentage = min(99, percentage)
                safe_message = (
                    "홈택스에서 사업자등록번호를 자동 확인하지 못했습니다. "
                    "사업자번호 없이 조회 가능한 보수총액은 수집을 시도했고, "
                    f"사업자번호가 필요한 관리번호·요율 {blocked_count}건은 "
                    "보류했습니다."
                    if business_number_missing
                    else (
                        f"서류 {int(summary.get('ready', 0) or 0)}건 저장, "
                        f"{int(summary.get('failed', 0) or 0)}건 실패"
                    )
                )
                updated = _update_claim_job(
                    case_id,
                    owner_ref,
                    sealed_payload=_seal_claim_job_payload(transient),
                    expires_at=float(
                        transient.get("expires_at", time.time()) or time.time()
                    ),
                    status="collection_partial",
                    progress=percentage,
                    safe_message=safe_message,
                    summary={
                        "ready": ready_count,
                        "target": target_count,
                        "progress_verified": progress_verified,
                        "failed": int(summary.get("failed", 0) or 0),
                        "skipped_count": len(summary.get("skipped", [])),
                        "blocked_count": blocked_count,
                        "business_number_missing": business_number_missing,
                    },
                )
                if not updated:
                    transient.clear()
                    context.clear()
                return

            updated = _update_claim_job(
                case_id,
                owner_ref,
                sealed_payload=_seal_claim_job_payload(transient),
                expires_at=float(
                    transient.get("expires_at", time.time()) or time.time()
                ),
                status="running",
                progress=percentage,
                safe_message=auth_stage_message,
                summary={
                    "ready": ready_count,
                    "target": target_count,
                    "progress_verified": progress_verified,
                },
            )
            if not updated:
                transient.clear()
                context.clear()
                return
            wake_event.wait(_auth_poll_delay(transient))
            wake_event.clear()
    except (ClaimProviderError, ClaimRepositoryError) as exc:
        if isinstance(transient, dict) and not _claim_job_can_continue(
            case_id,
            owner_ref,
        ):
            _sync_interrupted_claim_case(
                user_id,
                case_id,
                active_action=active_action,
                safe_error_code="AUTH_SESSION_EXPIRED",
                outcome="expired",
            )
            transient.clear()
            if isinstance(context, dict):
                context.clear()
            return
        safe_error_code = (
            _safe_provider_error_code(exc, "BACKGROUND")
            if isinstance(exc, ClaimProviderError)
            else "BACKGROUND_REPOSITORY_FAILED"
        )
        resealed = (
            _seal_claim_job_payload(transient)
            if isinstance(transient, dict)
            else None
        )
        _update_claim_job(
            case_id,
            owner_ref,
            **(
                {
                    "sealed_payload": resealed,
                    "expires_at": float(
                        transient.get("expires_at", time.time()) or time.time()
                    ),
                }
                if resealed is not None
                else {}
            ),
            status="paused",
            safe_message=(
                "인증 또는 자료수집 연결이 중단되었습니다. "
                "잠시 후 다시 시도해 주세요."
            ),
        )
        _sync_interrupted_claim_case(
            user_id,
            case_id,
            active_action=active_action,
            safe_error_code=safe_error_code,
            outcome="failed",
        )
    except Exception:
        if isinstance(transient, dict) and not _claim_job_can_continue(
            case_id,
            owner_ref,
        ):
            _sync_interrupted_claim_case(
                user_id,
                case_id,
                active_action=active_action,
                safe_error_code="AUTH_SESSION_EXPIRED",
                outcome="expired",
            )
            transient.clear()
            if isinstance(context, dict):
                context.clear()
            return
        resealed = (
            _seal_claim_job_payload(transient)
            if isinstance(transient, dict)
            else None
        )
        _update_claim_job(
            case_id,
            owner_ref,
            **(
                {
                    "sealed_payload": resealed,
                    "expires_at": float(
                        transient.get("expires_at", time.time()) or time.time()
                    ),
                }
                if resealed is not None
                else {}
            ),
            status="paused",
            safe_message=(
                "자동 인증 확인 중 오류가 발생했습니다. "
                "잠시 후 다시 시도해 주세요."
            ),
        )
        _sync_interrupted_claim_case(
            user_id,
            case_id,
            active_action=active_action,
            safe_error_code="BACKGROUND_UNEXPECTED_ERROR",
            outcome="failed",
        )


def _claim_job_sweeper() -> None:
    while True:
        time.sleep(15)
        now = time.time()
        with _CLAIM_JOB_LOCK:
            jobs = [
                (
                    case_id,
                    str(job.get("owner_ref", "")),
                    str(job.get("owner_user_id", "") or ""),
                    float(job.get("expires_at", 0) or 0),
                    str(job.get("status", "") or ""),
                    float(job.get("updated_at", 0) or 0),
                )
                for case_id, job in _CLAIM_JOBS.items()
            ]
        for (
            case_id,
            owner_ref,
            owner_user_id,
            expires_at,
            status,
            updated_at,
        ) in jobs:
            if status != "complete" and expires_at <= now:
                _expire_claim_job(
                    case_id,
                    owner_ref,
                    owner_user_id,
                )
            if status in {"complete", "collection_partial", "expired"}:
                if updated_at and now - updated_at > 30 * 60:
                    with _CLAIM_JOB_LOCK:
                        current = _CLAIM_JOBS.get(case_id)
                        if (
                            current
                            and current.get("owner_ref") == owner_ref
                            and current.get("status") == status
                        ):
                            _CLAIM_JOBS.pop(case_id, None)


def _ensure_claim_job_sweeper() -> None:
    global _CLAIM_SWEEPER_STARTED
    with _CLAIM_JOB_LOCK:
        if _CLAIM_SWEEPER_STARTED:
            return
        _CLAIM_SWEEPER_STARTED = True
    sweeper = threading.Thread(
        target=_claim_job_sweeper,
        daemon=True,
        name="claim-auth-sweeper",
    )
    sweeper.start()


def _register_background_claim_job(
    user_id: str,
    case_id: str,
    transient: dict[str, Any],
) -> None:
    _ensure_claim_job_sweeper()
    owner_ref = _claim_job_owner_ref(user_id)
    absolute_expires_at = _claim_absolute_expiry(transient)
    expires_at = min(
        float(transient.get("expires_at", 0) or 0),
        time.time() + AUTH_TTL_SECONDS,
        absolute_expires_at,
    )
    transient["expires_at"] = expires_at
    wake_event = threading.Event()
    with _CLAIM_JOB_LOCK:
        existing = _CLAIM_JOBS.get(case_id)
        if existing and existing.get("status") == "running":
            return
        _CLAIM_JOBS[case_id] = {
            "owner_ref": owner_ref,
            "owner_user_id": str(user_id or "").strip().lower(),
            "sealed_payload": _seal_claim_job_payload(transient),
            "expires_at": expires_at,
            "status": "queued",
            "progress": 0,
            "safe_message": CLAIM_AUTH_STAGE_MESSAGES[0],
            "summary": {
                "ready": 0,
                "target": 0,
            },
            "updated_at": time.time(),
            "wake_event": wake_event,
        }


def _activate_background_claim_job(
    user_id: str,
    case_id: str,
    *,
    initial_delay: float = AUTH_POLL_SECONDS,
) -> bool:
    owner_ref = _claim_job_owner_ref(user_id)
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if (
            not job
            or job.get("owner_ref") != owner_ref
            or float(job.get("expires_at", 0) or 0) <= time.time()
            or not job.get("sealed_payload")
        ):
            return False
        if job.get("status") == "running":
            wake_event = job.get("wake_event")
            if isinstance(wake_event, threading.Event):
                wake_event.set()
            return True
        job["status"] = "running"
        job["safe_message"] = CLAIM_AUTH_STAGE_MESSAGES[0]
        job["updated_at"] = time.time()
    _CLAIM_JOB_EXECUTOR.submit(
        _run_background_claim_job,
        user_id,
        case_id,
        owner_ref,
        initial_delay=initial_delay,
    )
    return True


def _claim_job_snapshot(
    user_id: str,
    case_id: str,
) -> dict[str, Any] | None:
    owner_ref = _claim_job_owner_ref(user_id)
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if not job or job.get("owner_ref") != owner_ref:
            return None
        return {
            "status": str(job.get("status", "") or ""),
            "progress": int(job.get("progress", 0) or 0),
            "safe_message": str(job.get("safe_message", "") or ""),
            "summary": dict(job.get("summary") or {}),
            "expires_at": float(job.get("expires_at", 0) or 0),
        }


def _retry_or_wake_claim_job(user_id: str, case_id: str) -> bool:
    owner_ref = _claim_job_owner_ref(user_id)
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if (
            not job
            or job.get("owner_ref") != owner_ref
            or float(job.get("expires_at", 0) or 0) <= time.time()
            or not job.get("sealed_payload")
        ):
            return False
        wake_event = job.get("wake_event")
        if not isinstance(wake_event, threading.Event):
            return False
        if job.get("status") == "running":
            wake_event.set()
            return True
    return _activate_background_claim_job(
        user_id,
        case_id,
        initial_delay=0,
    )


def _retry_authenticated_claim_collection(
    user_id: str,
    case_id: str,
) -> tuple[bool, str]:
    owner_user_id = str(user_id or "").strip().lower()
    normalized_case_id = str(case_id or "").strip()
    if not owner_user_id or not normalized_case_id:
        return False, "재수집할 요청을 확인하지 못했습니다."
    if not bool(provider_readiness().get("simple_auth_ready")):
        return False, "자료수집 API 설정을 먼저 확인해 주세요."

    try:
        repository = ClaimRepository(owner_user_id)
        case = repository.get_case(normalized_case_id)
    except ClaimRepositoryError:
        return False, "저장된 요청 상태를 확인하지 못했습니다."
    if not case:
        return False, "재수집할 요청을 찾지 못했습니다."
    if (
        str(case.get("hometax_status", "") or "") != "auth_complete"
        or str(case.get("comwel_status", "") or "") != "auth_complete"
    ):
        return False, "홈택스와 근로복지공단 인증을 모두 완료해야 재수집할 수 있습니다."

    owner_ref = _claim_job_owner_ref(owner_user_id)
    previous_status = ""
    should_expire = False
    now = time.time()
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(normalized_case_id)
        if not job or job.get("owner_ref") != owner_ref:
            return (
                False,
                "보안상 임시 인증정보가 남아 있지 않아 새 인증 요청이 필요합니다.",
            )
        previous_status = str(job.get("status", "") or "")
        sealed_payload = job.get("sealed_payload")
        expires_at = float(job.get("expires_at", 0) or 0)
        if previous_status == "complete":
            return False, "현재 연결된 자료수집이 이미 완료되었습니다."
        if (
            previous_status == "expired"
            or expires_at <= now
            or not isinstance(sealed_payload, bytes)
            or not sealed_payload
        ):
            should_expire = True
        elif previous_status in {"running", "queued"}:
            wake_event = job.get("wake_event")
            if isinstance(wake_event, threading.Event):
                wake_event.set()
            return True, "자료수집이 이미 진행 중입니다."
        elif previous_status not in {
            "paused",
            "collection_partial",
            "awaiting_business_selection",
            "awaiting_management_selection",
        }:
            return False, "현재 상태에서는 자료 재수집을 시작할 수 없습니다."
        else:
            transient: dict[str, Any] | None = None
            try:
                transient = _unseal_claim_job_payload(sealed_payload)
                context = transient.get("auth_context")
                if (
                    not isinstance(context, dict)
                    or not isinstance(transient.get("hometax"), dict)
                    or not isinstance(transient.get("comwel"), dict)
                ):
                    should_expire = True
            except ClaimProviderError:
                should_expire = True
            finally:
                if isinstance(transient, dict):
                    transient.clear()
            if not should_expire:
                job["status"] = "queued"
                job["safe_message"] = "인증 완료 자료의 재수집을 준비합니다."
                job["updated_at"] = now

    if should_expire:
        _expire_claim_job(
            normalized_case_id,
            owner_ref,
            owner_user_id,
        )
        return (
            False,
            "임시 인증정보가 만료되어 고객의 새 인증 요청이 필요합니다.",
        )

    try:
        repository.update_case_status(
            normalized_case_id,
            overall_status="collecting",
            last_safe_error_code=None,
        )
        repository.append_audit_event(
            case_id=normalized_case_id,
            action="collection_retry_requested",
            source="provider",
            outcome="requested",
            metadata={
                "previous_job_status": previous_status,
                "reuses_authenticated_session": True,
            },
        )
    except ClaimRepositoryError:
        with _CLAIM_JOB_LOCK:
            job = _CLAIM_JOBS.get(normalized_case_id)
            if (
                job
                and job.get("owner_ref") == owner_ref
                and str(job.get("status", "") or "") == "queued"
            ):
                job["status"] = previous_status
                job["safe_message"] = "일부 자료를 다시 수집할 수 있습니다."
                job["updated_at"] = time.time()
        try:
            repository.update_case_status(
                normalized_case_id,
                overall_status="auth_complete_collection_pending",
                last_safe_error_code="COLLECTION_RETRY_STATE_SAVE_FAILED",
            )
        except ClaimRepositoryError:
            pass
        return False, "재수집 상태를 저장하지 못했습니다. 잠시 후 다시 시도해 주세요."

    try:
        activated = _activate_background_claim_job(
            owner_user_id,
            normalized_case_id,
            initial_delay=0,
        )
    except Exception:
        activated = False
    if activated:
        _update_claim_job(
            normalized_case_id,
            owner_ref,
            safe_message="인증 완료 자료를 다시 수집하고 있습니다.",
        )
        return True, "인증 완료 상태를 유지한 채 자료 재수집을 시작했습니다."

    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(normalized_case_id)
        if (
            job
            and job.get("owner_ref") == owner_ref
            and str(job.get("status", "") or "") in {"queued", "running"}
        ):
            job["status"] = previous_status
            job["safe_message"] = "일부 자료를 다시 수집할 수 있습니다."
            job["updated_at"] = time.time()
    try:
        repository.update_case_status(
            normalized_case_id,
            overall_status="auth_complete_collection_pending",
            last_safe_error_code="COLLECTION_RETRY_START_FAILED",
        )
        repository.append_audit_event(
            case_id=normalized_case_id,
            action="collection_retry_requested",
            source="provider",
            outcome="failed",
            metadata={"safe_error_code": "COLLECTION_RETRY_START_FAILED"},
        )
    except ClaimRepositoryError:
        pass
    return False, "재수집 작업을 시작하지 못했습니다. 잠시 후 다시 시도해 주세요."


def _claim_collection_retry_state(
    case: dict[str, Any],
    job_snapshot: dict[str, Any] | None,
    *,
    provider_ready: bool,
    collection_complete: bool = False,
) -> str:
    authentication_complete = (
        str(case.get("hometax_status", "") or "") == "auth_complete"
        and str(case.get("comwel_status", "") or "") == "auth_complete"
    )
    if not authentication_complete:
        return "hidden"
    if collection_complete:
        return "complete"

    if not job_snapshot:
        overall_status = str(case.get("overall_status", "") or "")
        if overall_status in {"collected", "ready"}:
            # The persisted case may predate a corrected collection strategy
            # or still contain active pending rows. The caller already proved
            # collection_complete=False from the live document rows, so a new
            # authentication is required instead of hiding the retry notice.
            return "reauth_required"
        return (
            "reauth_required"
            if overall_status
            in {
                "auth_complete_collection_pending",
                "collection_queued",
                "collecting",
            }
            else "hidden"
        )

    job_status = str(job_snapshot.get("status", "") or "")
    if job_status in {
        "paused",
        "collection_partial",
        "awaiting_business_selection",
        "awaiting_management_selection",
    }:
        return "retryable" if provider_ready else "provider_unavailable"
    if job_status in {"running", "queued"}:
        return "running"
    if job_status == "complete":
        return "complete"
    if (
        job_status == "expired"
        or float(job_snapshot.get("expires_at", 0) or 0) <= time.time()
    ):
        return "reauth_required"
    return "hidden"


def _render_claim_collection_retry_action(
    user_id: str,
    case: dict[str, Any],
    job_snapshot: dict[str, Any] | None,
    *,
    provider_ready: bool,
    key_prefix: str,
    collection_complete: bool = False,
) -> str:
    state = _claim_collection_retry_state(
        case,
        job_snapshot,
        provider_ready=provider_ready,
        collection_complete=collection_complete,
    )
    case_id = str(case.get("id", "") or "")
    if state == "retryable":
        st.warning(
            "인증은 완료됐지만 일부 자료를 수집하지 못했습니다. "
            "완료된 자료는 그대로 두고 실패한 자료만 다시 수집할 수 있습니다."
        )
        if st.button(
            "실패 자료 재수집",
            type="primary",
            use_container_width=True,
            key=f"{key_prefix}_{case_id}",
        ):
            retried, message = _retry_authenticated_claim_collection(
                user_id,
                case_id,
            )
            if retried:
                st.session_state["_claim_active_case_v1"] = case_id
                st.session_state.pop(
                    f"_claim_collection_notified_{case_id}",
                    None,
                )
                st.toast(message)
                st.rerun(scope="app")
            else:
                st.error(message)
    elif state == "running":
        st.info(
            "자료 수집 또는 재수집이 진행 중입니다. "
            "완료된 자료는 다시 내려받지 않습니다."
        )
    elif state == "provider_unavailable":
        st.error("자료수집 API 설정을 확인한 뒤 재수집할 수 있습니다.")
    elif state == "reauth_required":
        st.warning(
            "재수집에 필요한 임시 인증정보가 만료됐거나 서버 재시작으로 "
            "삭제되었습니다. 개인정보 보호를 위해 복구하지 않으므로 "
            "고객에게 새 인증 요청을 보내 주세요."
        )
    return state


def _select_claim_business_number(
    user_id: str,
    case_id: str,
    selection_token: str,
) -> bool:
    owner_ref = _claim_job_owner_ref(user_id)
    requested_token = str(selection_token or "").strip()
    expired_after_unseal = False
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if (
            not job
            or job.get("owner_ref") != owner_ref
            or job.get("status") != "awaiting_business_selection"
            or float(job.get("expires_at", 0) or 0) <= time.time()
            or not job.get("sealed_payload")
        ):
            return False
        transient = _unseal_claim_job_payload(job["sealed_payload"])
        raw_candidates = transient.get("business_candidates")
        candidates = (
            raw_candidates if isinstance(raw_candidates, list) else []
        )
        selected_number = next(
            (
                _digits(candidate.get("business_number"))
                for candidate in candidates
                if isinstance(candidate, dict)
                and _is_valid_business_no(candidate.get("business_number"))
                and hmac.compare_digest(
                    _business_candidate_token(
                        case_id,
                        _digits(candidate.get("business_number")),
                    ),
                    requested_token,
                )
            ),
            "",
        )
        if not selected_number:
            transient.clear()
            return False
        transient["selected_business_number"] = selected_number
        transient["business_number"] = selected_number
        _set_claim_expiry(transient, COLLECTION_TTL_SECONDS)
        if transient["expires_at"] <= time.time():
            job["sealed_payload"] = b""
            job["status"] = "expired"
            job["safe_message"] = (
                "인증 유효시간이 지나 임시 인증정보를 삭제했습니다. "
                "새 인증 요청을 시작해 주세요."
            )
            job["updated_at"] = time.time()
            transient.clear()
            expired_after_unseal = True
        else:
            previous_summary = dict(job.get("summary") or {})
            job["sealed_payload"] = _seal_claim_job_payload(transient)
            job["expires_at"] = transient["expires_at"]
            job["status"] = "queued"
            job["progress"] = int(job.get("progress", 0) or 0)
            job["safe_message"] = (
                "선택한 사업자번호로 근로복지공단 자료수집을 준비합니다."
            )
            job["summary"] = {
                "ready": int(previous_summary.get("ready", 0) or 0),
                "target": int(previous_summary.get("target", 0) or 0),
            }
            job["updated_at"] = time.time()
    if expired_after_unseal:
        _sync_interrupted_claim_case(
            user_id,
            case_id,
            active_action="collect",
            safe_error_code="AUTH_SESSION_EXPIRED",
            outcome="expired",
        )
        return False
    return _activate_background_claim_job(
        user_id,
        case_id,
        initial_delay=0,
    )


def _select_claim_management_number(
    user_id: str,
    case_id: str,
    management_number: str,
) -> bool:
    owner_ref = _claim_job_owner_ref(user_id)
    selected_digits = _digits(management_number)
    expired_after_unseal = False
    with _CLAIM_JOB_LOCK:
        job = _CLAIM_JOBS.get(case_id)
        if (
            not job
            or job.get("owner_ref") != owner_ref
            or job.get("status") != "awaiting_management_selection"
            or float(job.get("expires_at", 0) or 0) <= time.time()
            or not job.get("sealed_payload")
        ):
            return False
        summary = dict(job.get("summary") or {})
        allowed_numbers = {
            _digits(number)
            for number in summary.get("management_numbers", [])
            if _digits(number)
        }
        if selected_digits not in allowed_numbers:
            return False
        transient = _unseal_claim_job_payload(job["sealed_payload"])
        transient["selected_management_number"] = selected_digits
        _set_claim_expiry(transient, COLLECTION_TTL_SECONDS)
        if transient["expires_at"] <= time.time():
            job["sealed_payload"] = b""
            job["status"] = "expired"
            job["safe_message"] = (
                "인증 유효시간이 지나 임시 인증정보를 삭제했습니다. "
                "새 인증 요청을 시작해 주세요."
            )
            job["updated_at"] = time.time()
            transient.clear()
            expired_after_unseal = True
        else:
            previous_summary = dict(job.get("summary") or {})
            job["sealed_payload"] = _seal_claim_job_payload(transient)
            job["expires_at"] = transient["expires_at"]
            job["status"] = "queued"
            job["progress"] = int(job.get("progress", 0) or 0)
            job["safe_message"] = "선택한 사업장의 요율 수집을 준비합니다."
            job["summary"] = {
                "ready": int(previous_summary.get("ready", 0) or 0),
                "target": int(previous_summary.get("target", 0) or 0),
            }
            job["updated_at"] = time.time()
    if expired_after_unseal:
        _sync_interrupted_claim_case(
            user_id,
            case_id,
            active_action="collect",
            safe_error_code="AUTH_SESSION_EXPIRED",
            outcome="expired",
        )
        return False
    return _activate_background_claim_job(
        user_id,
        case_id,
        initial_delay=0,
    )


def _case_label(row: dict[str, Any]) -> str:
    company = _clean(row.get("company_name")) or "업체명 없음"
    requested = _clean(row.get("requested_at"))
    date_text = requested[:10] if requested else "-"
    return f"{company} · {date_text} · {str(row.get('id', ''))[:8]}"


def _selected_customer(
    user_id: str,
    input_mode: str,
) -> tuple[dict[str, Any], str]:
    if input_mode == "직접입력":
        return {}, "manual"

    customers = load_registered_customers(
        get_user_cumulative_db_path(user_id),
        owner_user_id=user_id,
    )
    if customers.empty:
        st.info("등록된 고객이 없어 직접 입력 방식으로 전환했습니다.")
        return {}, "manual"

    labels, row_map = build_customer_labels(customers)
    selected = st.selectbox(
        "고객 선택",
        labels,
        key="claim_customer_selector_v1",
    )
    row = customers.loc[row_map[selected]].to_dict()
    suffix = hashlib.sha256(selected.encode("utf-8")).hexdigest()[:8]
    return row, suffix


def _render_intro(
    repository_ready: bool,
    repository_message: str,
    readiness: dict[str, object],
) -> None:
    st.markdown(
        """
        <style>
        .claim-hero {
            border: 1px solid #dce5f3;
            border-radius: 18px;
            padding: 1.15rem 1.25rem;
            margin-bottom: 0.9rem;
            background:
                linear-gradient(135deg, #f8fbff 0%, #eef5ff 58%, #f7fbff 100%);
        }
        .claim-hero h2 {
            margin: 0 0 0.38rem 0;
            color: #102d5c;
            font-size: 1.5rem;
            letter-spacing: -0.04em;
        }
        .claim-hero p {
            margin: 0;
            color: #4c607c;
            line-height: 1.65;
        }
        .claim-flow {
            display: grid;
            grid-template-columns: repeat(4, minmax(0, 1fr));
            gap: 0.55rem;
            margin: 0.75rem 0 1rem;
        }
        .claim-step {
            border: 1px solid #dce5f3;
            border-radius: 12px;
            padding: 0.72rem 0.78rem;
            background: #ffffff;
            color: #18345f;
            font-size: 0.91rem;
            font-weight: 700;
        }
        .claim-step span {
            display: block;
            color: #6c7d96;
            font-size: 0.72rem;
            margin-bottom: 0.22rem;
            font-weight: 700;
        }
        @media (max-width: 700px) {
            .claim-flow { grid-template-columns: 1fr 1fr; }
            .claim-hero { padding: 0.95rem; }
            .claim-hero h2 { font-size: 1.28rem; }
        }
        </style>
        <div class="claim-hero">
            <h2>경정청구 자료수집</h2>
            <p>
                개인사업자는 홈택스 인증 완료 후 근로복지공단 인증을
                순서대로 요청하고, 법인사업자는 공동인증서 인증 완료 후
                자료를 수집합니다.
            </p>
        </div>
        <div class="claim-flow">
            <div class="claim-step"><span>01</span>고객정보 입력</div>
            <div class="claim-step"><span>02</span>홈택스 인증</div>
            <div class="claim-step"><span>03</span>근로복지공단 인증</div>
            <div class="claim-step"><span>04</span>수집결과 확인</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if not repository_ready:
        st.warning(repository_message)
    if not bool(readiness.get("simple_auth_ready")):
        st.info(
            "화면과 안전한 저장 구조는 준비됐습니다. 실제 카카오 인증 발송은 "
            "승인된 중계 API 설정 후 활성화됩니다."
        )


def _render_remote_personal_invite(
    *,
    user_id: str,
    user_name: str,
    repository_ready: bool,
    provider_ready: bool,
) -> None:
    st.markdown("#### 고객 카카오톡 인증 링크 발송")
    st.caption(
        "고객에게 안전한 입력 링크를 보냅니다. 고객이 링크에서 본인정보와 "
        "동의 내용을 직접 입력한 뒤 홈택스·근로복지공단 인증을 진행합니다."
    )

    notice = st.session_state.get(_REMOTE_INVITE_NOTICE_KEY)
    if isinstance(notice, dict):
        expiry_text = _format_remote_invite_expiry(notice.get("expires_at"))
        success_message = "카카오톡 인증 링크 발송 요청을 등록했습니다."
        if expiry_text:
            success_message += f" 링크 유효시간: {expiry_text}까지"
        st.success(success_message)

    readiness_messages: list[str] = []
    if not repository_ready:
        readiness_messages.append(
            "원격 인증 저장소 연결이 준비되지 않았습니다. 관리자에게 문의해주세요."
        )
    if not provider_ready:
        readiness_messages.append(
            "홈택스·근로복지공단 인증 연동 설정이 완료되지 않았습니다. "
            "관리자에게 문의해주세요."
        )
    remote_runtime_ready, remote_runtime_message = (
        _remote_invite_runtime_readiness()
    )
    if not remote_runtime_ready:
        readiness_messages.append(remote_runtime_message)
    for message in readiness_messages:
        st.warning(message)

    with st.form("claim_remote_personal_invite_v1", clear_on_submit=True):
        name_col, phone_col = st.columns(2)
        with name_col:
            customer_name = st.text_input(
                "고객 이름",
                placeholder="홍길동",
                key="claim_remote_customer_name_v1",
            )
        with phone_col:
            customer_phone = st.text_input(
                "고객 휴대전화",
                placeholder="010-0000-0000",
                key="claim_remote_customer_phone_v1",
            )
        submitted = st.form_submit_button(
            "카카오톡 인증 링크 발송",
            use_container_width=True,
            type="primary",
            disabled=bool(readiness_messages),
        )

    if not submitted:
        return

    name, phone, errors = _validate_remote_invite_input(
        customer_name,
        customer_phone,
    )
    if errors:
        for error in errors:
            st.error(error)
        return

    try:
        result = _create_remote_claim_invite(
            owner_user_id=user_id,
            requested_by=user_name or user_id,
            customer_name=name,
            customer_phone=phone,
        )
    except RemoteInviteUIError as exc:
        st.error(str(exc))
        return

    notice = {
        "expires_at": result.get("expires_at"),
        "status": result.get("status"),
    }
    st.session_state[_REMOTE_INVITE_NOTICE_KEY] = notice
    expiry_text = _format_remote_invite_expiry(notice.get("expires_at"))
    success_message = "카카오톡 인증 링크 발송 요청을 등록했습니다."
    if expiry_text:
        success_message += f" 링크 유효시간: {expiry_text}까지"
    st.success(success_message)


def _render_personal_request(
    user_id: str,
    user_name: str,
    repository: ClaimRepository | None,
    provider_ready: bool,
) -> None:
    input_mode = st.radio(
        "고객정보 입력 방식",
        ["카카오톡 발송", "직접입력"],
        horizontal=True,
        key="claim_personal_input_mode_v2",
        help=(
            "카카오톡 발송은 고객 이름과 휴대전화만 입력해 안전한 본인정보 "
            "입력 링크를 보내는 방식입니다. 직접입력은 담당자가 고객의 "
            "인증정보를 입력해 바로 기관 인증을 요청하는 방식입니다."
        ),
    )
    remote_input = input_mode == "카카오톡 발송"
    if remote_input:
        _render_remote_personal_invite(
            user_id=user_id,
            user_name=user_name,
            repository_ready=repository is not None,
            provider_ready=provider_ready,
        )
        return

    suffix = "manual"
    company_name = ""
    business_no = ""
    identity_front = ""
    identity_rear = ""

    form_key = f"claim_personal_request_{suffix}"
    with st.form(form_key, clear_on_submit=True):
        st.markdown("#### 개인사업자 카카오 인증 요청")
        st.caption(
            "홈택스 인증을 먼저 발송합니다. 고객이 홈택스 인증을 마치면 "
            "약 1초 후 근로복지공단 인증 발송을 시작합니다. "
            "사업자정보는 홈택스에서 자동으로 확인합니다."
        )

        name_col, phone_col = st.columns(2)
        with name_col:
            representative = st.text_input(
                "대표자 이름",
                key=f"claim_representative_{suffix}",
            )
        with phone_col:
            cellphone = st.text_input(
                "대표자 휴대전화",
                placeholder="010-0000-0000",
                key=f"claim_cellphone_{suffix}",
            )

        st.markdown("**인증·수집 기관**")
        st.info(
            "① 홈택스 인증 발송 → ② 홈택스 승인 자동 확인 → "
            "③ 약 1초 후 근로복지공단 인증 발송 → "
            "④ 두 기관 인증 완료 후 자료수집"
        )
        id_front_col, id_rear_col = st.columns(2)
        with id_front_col:
            identity_front = st.text_input(
                "주민등록번호 앞 6자리",
                max_chars=6,
                placeholder="생년월일 6자리",
                type="password",
                key=f"claim_identity_front_{suffix}",
                help=(
                    "홈택스에는 생년월일로 변환해 전송합니다. "
                    "입력값은 DB와 로그에 저장하지 않습니다."
                ),
            )
        with id_rear_col:
            identity_rear = st.text_input(
                "주민등록번호 뒤 7자리",
                max_chars=7,
                placeholder="근로복지공단 인증에만 사용",
                type="password",
                key=f"claim_identity_rear_{suffix}",
                    help=(
                        "순차 인증 자동 확인을 위해 Railway 서버 메모리에 "
                        "인증 및 자료수집 중 최대 45분만 암호화해 보관하며 "
                        "DB·로그에는 저장하지 않습니다."
                    ),
                )

        consent_confirmed = st.checkbox(
            CONSENT_NOTICE_TEXT,
            key=f"claim_consent_{suffix}",
        )
        legal_basis_confirmed = st.checkbox(
            COLLECTION_AUTHORITY_TEXT,
            key=f"claim_legal_basis_{suffix}",
        )
        submitted = st.form_submit_button(
            "홈택스 카카오 인증 발송",
            use_container_width=True,
            type="primary",
            disabled=not (repository and provider_ready),
        )

    if not submitted:
        return

    sources = ["hometax", "comwel"]
    front_digits = _digits(identity_front)
    rear_digits = _digits(identity_rear)
    phone_digits = _digits(cellphone)
    business_digits = _digits(business_no)
    errors: list[str] = []
    if business_digits and not _is_valid_business_no(business_digits):
        errors.append("유효한 사업자등록번호인지 확인해주세요.")
    if not representative.strip():
        errors.append("대표자 이름을 입력해주세요.")
    if len(phone_digits) != 11 or not phone_digits.startswith("010"):
        errors.append("카카오 인증을 받을 010 휴대전화 번호를 확인해주세요.")
    if "hometax" in sources and not _birth_date_from_identity(
        front_digits,
        rear_digits,
    ):
        errors.append("홈택스 인증용 주민등록번호 앞자리와 구분값을 확인해주세요.")
    if "comwel" in sources and (
        len(front_digits) != 6 or len(rear_digits) != 7
    ):
        errors.append("근로복지공단 인증에는 주민등록번호 13자리가 필요합니다.")
    if not consent_confirmed or not legal_basis_confirmed:
        errors.append("동의와 법적 근거 확인 항목을 모두 확인해주세요.")
    if errors:
        for error in errors:
            st.error(error)
        return

    assert repository is not None
    case: dict[str, Any] | None = None
    source_results: dict[str, str] = {
        "comwel_status": "request_ready",
    }
    request_started_at = time.time()
    transient: dict[str, Any] = {
        "request_started_at": request_started_at,
        "absolute_expires_at": (
            request_started_at + COLLECTION_TTL_SECONDS
        ),
        "expires_at": min(
            request_started_at + AUTH_TTL_SECONDS,
            request_started_at + COLLECTION_TTL_SECONDS,
        ),
        "stage_started_at": request_started_at,
        "expected_sources": list(sources),
        "business_number": business_digits,
        "auth_context": {
            "representative": representative.strip(),
            "cellphone": phone_digits,
            "birth_date": _birth_date_from_identity(front_digits, rear_digits),
            "identity_number": f"{front_digits}{rear_digits}",
        },
    }
    provider_failures: list[tuple[str, str]] = []
    try:
        case = repository.create_case(
            company_name=company_name.strip() or "상호명 미입력",
            business_no=business_digits,
            business_type="individual",
            representative_name=representative,
            cellphone=phone_digits,
            requested_by=user_name or user_id,
            selected_sources=sources,
            consent_version=CONSENT_VERSION,
            consent_text_sha256=CONSENT_TEXT_SHA256,
            consent_channel="staff_attestation",
            retention_policy_version=RETENTION_POLICY_VERSION,
            collection_authority_confirmed=legal_basis_confirmed,
        )
        client = TilkoClaimClient()
        if "hometax" in sources:
            try:
                birth_date = _birth_date_from_identity(front_digits, rear_digits)
                transient["hometax"] = client.request_hometax_kakao(
                    birth_date=birth_date,
                    user_name=representative.strip(),
                    cellphone=phone_digits,
                )
                source_results["hometax_status"] = "auth_requested"
                _register_background_claim_job(
                    user_id,
                    str(case["id"]),
                    transient,
                )
                st.session_state["_claim_active_case_v1"] = case["id"]
            except ClaimProviderError as exc:
                source_results["hometax_status"] = "failed"
                provider_failures.append(("홈택스", str(exc)))
        requested_sources = [
            source
            for source in ("hometax",)
            if source in transient
        ]
        repository.update_case_status(
            case["id"],
            **source_results,
            overall_status=(
                "auth_partial"
                if provider_failures and requested_sources
                else "auth_pending"
                if requested_sources
                else "failed"
            ),
            auth_requested_at=(
                datetime.now(timezone.utc).isoformat()
                if requested_sources
                else None
            ),
            last_safe_error_code=(
                "AUTH_REQUEST_PARTIAL"
                if provider_failures and requested_sources
                else "AUTH_REQUEST_FAILED"
                if provider_failures
                else None
            ),
        )
        repository.append_audit_event(
            case_id=case["id"],
            action="auth_request",
            source="provider",
            outcome=(
                "partial"
                if provider_failures and requested_sources
                else "failed"
                if provider_failures
                else "success"
            ),
            metadata={
                "requested_sources": requested_sources,
                "planned_sources": ["comwel"],
                "failed_source_count": len(provider_failures),
            },
        )
        if requested_sources and not _activate_background_claim_job(
            user_id,
            str(case["id"]),
        ):
            raise ClaimProviderError(
                "자동 인증 확인 작업을 시작하지 못했습니다. "
                "새 인증 요청을 시작해 주세요."
            )
        source_labels = [
            "홈택스"
            for source in requested_sources
        ]
        if source_labels:
            message = (
                f"{'·'.join(source_labels)} 카카오 인증 요청을 발송했습니다. "
                "화면을 이동하거나 닫아도 Railway가 고객 인증을 자동으로 "
                "확인합니다. "
                "홈택스 인증이 확인되는 즉시 근로복지공단 인증을 이어서 "
                "발송합니다."
            )
            if provider_failures:
                failed_labels = "·".join(
                    label for label, _ in provider_failures
                )
                message += f" {failed_labels} 요청은 실패해 다시 요청해야 합니다."
            st.session_state["_claim_flash_v1"] = message
            # Reload the case/job state after the provider request starts.
            st.rerun()
        transient.clear()
        for label, message in provider_failures:
            st.error(f"{label}: {message}")
    except (ClaimProviderError, ClaimRepositoryError) as exc:
        if case is not None:
            _update_claim_job(
                str(case["id"]),
                _claim_job_owner_ref(user_id),
                status="paused",
                safe_message=(
                    "초기 상태 저장에 실패했습니다. "
                    "‘지금 인증 상태 확인’을 눌러 다시 시도해 주세요."
                ),
            )
            try:
                repository.update_case_status(
                    case["id"],
                    overall_status="failed",
                    last_safe_error_code="AUTH_REQUEST_FAILED",
                )
                repository.append_audit_event(
                    case_id=case["id"],
                    action="auth_request",
                    source="provider",
                    outcome="failed",
                    metadata={"safe_error_code": "AUTH_REQUEST_FAILED"},
                )
            except ClaimRepositoryError:
                pass
        transient.clear()
        st.error(str(exc))


def _render_corporate_request(
    user_id: str,
    user_name: str,
    repository: ClaimRepository | None,
    corporate_ready: bool,
) -> None:
    st.markdown("#### 법인사업자 공동인증서 인증")
    st.caption(
        "공동인증서는 고객 PC의 로컬 인증 모듈에서 사용합니다. "
        "인증서 파일·개인키·비밀번호는 OASIS와 Supabase에 저장하지 않습니다."
    )
    input_mode = st.radio(
        "고객정보 입력 방식",
        ["등록 고객 선택", "직접 입력"],
        horizontal=True,
        key="claim_corporate_input_mode_v1",
    )
    row, suffix = _selected_customer(user_id, input_mode)
    with st.form(f"claim_corporate_request_{suffix}"):
        company_col, business_col = st.columns(2)
        with company_col:
            company_name = st.text_input(
                "법인명",
                value=_clean(row.get("업체명")),
                key=f"claim_corp_company_{suffix}",
            )
        with business_col:
            business_no = st.text_input(
                "사업자등록번호",
                value=_format_business_no(row.get("사업자등록번호")),
                key=f"claim_corp_business_no_{suffix}",
            )
        representative = st.text_input(
            "대표자 이름",
            value=_clean(row.get("대표자명")),
            key=f"claim_corp_representative_{suffix}",
        )
        consent_confirmed = st.checkbox(
            "법인의 자료조회 위임과 수집 범위를 확인했습니다.",
            key=f"claim_corp_consent_{suffix}",
        )
        submitted = st.form_submit_button(
            "공동인증서 인증 준비 건 등록",
            use_container_width=True,
            disabled=not (repository and corporate_ready),
        )

    if submitted:
        if not company_name.strip():
            st.error("법인명을 입력해주세요.")
            return
        if not _is_valid_business_no(business_no):
            st.error("유효한 사업자등록번호인지 확인해주세요.")
            return
        if not representative.strip():
            st.error("대표자 이름을 입력해주세요.")
            return
        if not consent_confirmed:
            st.error("법인의 자료조회 위임과 수집 범위를 먼저 확인해주세요.")
            return
        assert repository is not None
        try:
            case = repository.create_case(
                company_name=company_name,
                business_no=business_no,
                business_type="corporation",
                representative_name=representative,
                cellphone="",
                requested_by=user_name or user_id,
                selected_sources=["hometax", "comwel"],
                consent_version=CONSENT_VERSION,
                consent_text_sha256=CONSENT_TEXT_SHA256,
                consent_channel="staff_attestation",
                retention_policy_version=RETENTION_POLICY_VERSION,
                collection_authority_confirmed=consent_confirmed,
            )
            repository.update_case_status(
                case["id"],
                hometax_status="certificate_required",
                comwel_status="certificate_required",
                overall_status="auth_preparing",
            )
            st.success("공동인증서 인증 준비 건을 등록했습니다.")
        except ClaimRepositoryError as exc:
            st.error(str(exc))

    if not corporate_ready:
        st.info(
            "법인 인증은 고객 PC용 공동인증서 모듈 계약, 요청 건별 인증 "
            "연결, 결과 콜백 검증까지 설정한 뒤 활성화됩니다. 인증서 "
            "파일·개인키·비밀번호는 OASIS 서버로 전송하지 않습니다."
        )


def _render_request_tab(
    user_id: str,
    user_name: str,
    repository: ClaimRepository | None,
    readiness: dict[str, object],
) -> None:
    business_type = st.radio(
        "사업자 구분",
        ["개인사업자", "법인사업자"],
        horizontal=True,
        key="claim_business_type_v1",
    )
    if business_type == "개인사업자":
        _render_personal_request(
            user_id,
            user_name,
            repository,
            bool(readiness.get("simple_auth_ready")),
        )
    else:
        _render_corporate_request(
            user_id,
            user_name,
            repository,
            bool(readiness.get("corporate_auth_ready")),
        )


def _cases_dataframe(cases: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for case in cases:
        rows.append(
            {
                "요청일": _clean(case.get("requested_at"))[:16].replace("T", " "),
                "상호명": _clean(case.get("company_name")),
                "사업자번호": _clean(case.get("business_no_masked")),
                "구분": (
                    "개인"
                    if case.get("business_type") == "individual"
                    else "법인"
                ),
                "홈택스": _source_status(case.get("hometax_status")),
                "근로복지공단": _source_status(case.get("comwel_status")),
                "전체상태": _source_status(case.get("overall_status")),
            }
        )
    return pd.DataFrame(rows)


@st.fragment(run_every="3s")
def _render_auto_claim_monitor(
    user_id: str,
    case_id: str,
    repository: ClaimRepository,
    provider_ready: bool,
) -> None:
    job_snapshot = _claim_job_snapshot(user_id, case_id)
    try:
        current_case = repository.get_case(case_id)
    except ClaimRepositoryError as exc:
        st.error(str(exc))
        return
    if not current_case:
        if st.session_state.get("_claim_active_case_v1") == case_id:
            st.session_state.pop("_claim_active_case_v1", None)
        st.caption("경정청구 요청 건을 찾지 못했습니다.")
        return

    st.markdown("#### 인증 진행")
    _render_claim_auth_stage(current_case)
    st.markdown("#### 자료수집 진행률")
    (
        progress_percent,
        progress_text,
        ready_count,
        target_count,
        progress_verified,
    ) = _claim_collection_progress_from_repository(
        repository,
        case_id,
    )
    st.progress(progress_percent, text=progress_text)
    if not progress_verified:
        st.caption(
            "Supabase의 실제 수집 상태가 확인되면 진행률을 표시합니다."
        )

    if job_snapshot:
        job_status = str(job_snapshot.get("status", "") or "")
        summary = dict(job_snapshot.get("summary") or {})
        verified_complete = bool(
            progress_verified
            and target_count > 0
            and progress_percent == 100
        )
        if job_status == "complete" and verified_complete:
            skipped_count = int(summary.get("skipped_count", 0) or 0)
            completion_message = (
                f"자동수집 대상 {target_count}건 처리가 완료되었습니다."
                + (
                    f" 현재 인증방식 또는 필수 식별정보 제한으로 "
                    f"{skipped_count}건은 제외했습니다."
                    if skipped_count
                    else ""
                )
            )
            notified_key = f"_claim_collection_notified_{case_id}"
            if not st.session_state.get(notified_key):
                st.session_state[notified_key] = True
                st.session_state["_claim_collection_completed_v1"] = (
                    completion_message
                )
                st.session_state["_claim_flash_v1"] = completion_message
                if st.session_state.get("_claim_active_case_v1") == case_id:
                    st.session_state.pop("_claim_active_case_v1", None)
                st.rerun(scope="app")
            st.success(
                f"{completion_message} ‘수집결과’에서 확인해 주세요."
            )
            return
        if job_status == "complete" and not verified_complete:
            st.warning(
                "작업 종료 신호는 받았지만 Supabase의 실제 수집 자료가 "
                "모두 확인되지 않아 완료로 표시하지 않았습니다."
            )
            return
        if job_status == "expired":
            if st.session_state.get("_claim_active_case_v1") == case_id:
                st.session_state.pop("_claim_active_case_v1", None)
            st.warning(str(job_snapshot.get("safe_message", "") or "인증정보가 만료되었습니다."))
            return
        retry_state = _render_claim_collection_retry_action(
            user_id,
            current_case,
            job_snapshot,
            provider_ready=provider_ready,
            key_prefix="claim_monitor_collection_retry",
            collection_complete=verified_complete,
        )
        if retry_state in {
            "retryable",
            "running",
            "provider_unavailable",
            "reauth_required",
        }:
            return
        manual_check = st.button(
            "지금 인증 상태 확인",
            use_container_width=True,
            key=f"claim_auto_check_{case_id}",
        )
        if manual_check:
            if _retry_or_wake_claim_job(user_id, case_id):
                st.toast("인증 상태를 바로 다시 확인합니다.", icon="🔄")
            else:
                st.warning("임시 인증정보가 만료되어 새 인증 요청이 필요합니다.")
        if job_status in {"paused", "collection_partial"}:
            st.warning(
                str(job_snapshot.get("safe_message", "") or "")
                or "자동 확인을 잠시 멈췄습니다."
            )
        else:
            remaining = max(
                0,
                int(float(job_snapshot.get("expires_at", 0) or 0) - time.time()),
            )
            st.caption(
                "화면을 이동하거나 닫아도 Railway 서버가 처음 30초는 약 1초마다, "
                "다음 60초는 약 3초마다, 이후에는 약 10초마다 자동으로 "
                f"확인합니다. 임시 인증정보는 {remaining // 60}분 "
                f"{remaining % 60}초 뒤 자동 삭제됩니다."
            )
        return

    if (
        progress_verified
        and target_count > 0
        and progress_percent == 100
        and str(current_case.get("hometax_status", "") or "")
        == "auth_complete"
        and str(current_case.get("comwel_status", "") or "")
        == "auth_complete"
        and str(current_case.get("overall_status", "") or "")
        in {"ready", "collected"}
    ):
        if st.session_state.get("_claim_active_case_v1") == case_id:
            st.session_state.pop("_claim_active_case_v1", None)
        st.success(
            "현재 연결된 서류 수집이 완료되었습니다. "
            "‘수집결과’에서 확인해 주세요."
        )
        return
    if st.session_state.get("_claim_active_case_v1") == case_id:
        st.session_state.pop("_claim_active_case_v1", None)
    st.caption(
        "이 요청의 임시 인증정보가 없거나 Railway가 재시작되었습니다. "
        "개인정보 보호를 위해 복구하지 않으므로 새 인증 요청을 시작해 주세요."
    )


def _render_status_tab(
    user_id: str,
    repository: ClaimRepository | None,
    provider_ready: bool,
) -> None:
    if repository is None:
        st.info("전용 저장소 설치 후 인증 진행상황이 표시됩니다.")
        return
    try:
        cases = repository.list_cases()
    except ClaimRepositoryError as exc:
        st.error(str(exc))
        return
    if not cases:
        st.info("아직 등록된 경정청구 요청이 없습니다.")
        return

    st.dataframe(
        _cases_dataframe(cases),
        use_container_width=True,
        hide_index=True,
    )
    labels = [_case_label(case) for case in cases]
    selected_label = st.selectbox(
        "완료 여부를 확인할 요청",
        labels,
        key="claim_status_case_selector_v1",
    )
    selected_case = cases[labels.index(selected_label)]
    case_id = str(selected_case.get("id", ""))
    job_snapshot = _claim_job_snapshot(user_id, case_id)
    if job_snapshot:
        if st.session_state.get("_claim_active_case_v1") == case_id:
            st.info("이 요청은 화면 상단에서 자동 확인 중입니다.")
        else:
            st.info(
                "이 요청은 서버에서 자동 처리 중입니다. 최신 상태는 "
                "위 목록에서 확인할 수 있습니다."
            )
        return

    if str(selected_case.get("overall_status", "") or "") in {
        "ready",
        "collected",
    }:
        st.success(
            "자료수집이 완료되었습니다. ‘수집결과’에서 서류를 확인해 주세요."
        )
    elif selected_case.get("business_type") == "corporation":
        st.caption("법인 공동인증 완료 상태는 인증 모듈 콜백으로 갱신됩니다.")
    else:
        st.caption(
            f"현재 상태: {_source_status(selected_case.get('overall_status'))}"
        )


def _claim_result_status_group(case: dict[str, Any]) -> str:
    overall_status = str(case.get("overall_status", "") or "")
    if overall_status in {"collected", "ready"}:
        return "수집 완료"
    if (
        overall_status == "auth_complete_collection_pending"
        and bool(case.get("last_safe_error_code"))
    ):
        return "일부 수집 실패"
    if overall_status in {
        "auth_complete",
        "auth_complete_collection_pending",
        "collection_queued",
        "collecting",
    }:
        return "인증 완료"
    if overall_status in {"failed", "auth_partial"}:
        return "실패"
    return "인증 대기"


def _claim_result_customer_name(case: dict[str, Any]) -> str:
    company_name = _clean(case.get("company_name"))
    if company_name and company_name != "상호명 미입력":
        return company_name
    return (
        _clean(case.get("representative_name_masked"))
        or company_name
        or "고객정보 확인 중"
    )


def _claim_downloadable_documents(
    documents: list[dict[str, Any]],
    *,
    now: datetime | None = None,
) -> list[dict[str, Any]]:
    return [
        document
        for document in documents
        if _claim_document_is_downloadable(document, now=now)
    ]


def _claim_download_cache_fingerprint(
    documents: list[dict[str, Any]],
) -> str:
    """Return a stable version fingerprint for downloadable document caches."""
    versions = sorted(
        (
            _clean(document.get("id")),
            _clean(document.get("storage_path")),
            _clean(document.get("content_sha256")).lower(),
            _clean(document.get("retention_until")),
        )
        for document in documents
        if isinstance(document, dict)
    )
    payload = json.dumps(
        versions,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _claim_result_case_view(
    case: dict[str, Any],
    documents: list[dict[str, Any]] | None = None,
    *,
    now: datetime | None = None,
) -> dict[str, Any]:
    return {
        "customer_name": _claim_result_customer_name(case),
        "business_no": _clean(case.get("business_no_masked")) or "-",
        "phone": _clean(case.get("phone_masked")) or "-",
        "business_type": (
            "개인"
            if case.get("business_type") == "individual"
            else "법인"
        ),
        "hometax_status": _source_status(case.get("hometax_status")),
        "comwel_status": _source_status(case.get("comwel_status")),
        "overall_status": _claim_result_status_group(case),
        "requested_at": (
            _clean(case.get("requested_at"))[:16].replace("T", " ") or "-"
        ),
        "requested_by": _clean(case.get("requested_by")) or "-",
        "downloadable_document_count": len(
            _claim_downloadable_documents(documents or [], now=now)
        ),
    }


_CLAIM_RESULTS_EXCEL_NOTE = (
    "수집 서류는 OASIS CRM의 수집결과 화면에서 고객별 ‘서류조회’를 "
    "눌러 확인하고 다운로드하세요."
)
_CLAIM_RESULTS_EXCEL_COLUMNS = (
    "번호",
    "고객 / 상호",
    "구분",
    "사업자번호",
    "휴대전화",
    "홈택스",
    "근로복지공단",
    "수집 상태",
    "등록일",
    "담당자",
)


def _build_claim_results_excel(cases: list[dict[str, Any]]) -> bytes:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    rows: list[dict[str, Any]] = []
    for row_number, case in enumerate(cases, start=1):
        case_view = _claim_result_case_view(case)
        rows.append(
            {
                "번호": row_number,
                "고객 / 상호": case_view["customer_name"],
                "구분": case_view["business_type"],
                "사업자번호": case_view["business_no"],
                "휴대전화": case_view["phone"],
                "홈택스": case_view["hometax_status"],
                "근로복지공단": case_view["comwel_status"],
                "수집 상태": case_view["overall_status"],
                "등록일": case_view["requested_at"],
                "담당자": case_view["requested_by"],
            }
        )

    output = BytesIO()
    dataframe = pd.DataFrame(
        rows,
        columns=list(_CLAIM_RESULTS_EXCEL_COLUMNS),
    )
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(
            writer,
            sheet_name="수집결과",
            index=False,
            startrow=2,
        )
        worksheet = writer.sheets["수집결과"]
        last_column = len(_CLAIM_RESULTS_EXCEL_COLUMNS)
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=last_column,
        )
        note_cell = worksheet.cell(row=1, column=1)
        note_cell.value = _CLAIM_RESULTS_EXCEL_NOTE
        note_cell.font = Font(color="1F4E78", bold=True, size=11)
        note_cell.fill = PatternFill("solid", fgColor="DDEBF7")
        note_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True,
        )
        worksheet.row_dimensions[1].height = 34

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        border_side = Side(style="thin", color="D9E2F3")
        cell_border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=border_side,
        )
        for cell in worksheet[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )
            cell.border = cell_border
        worksheet.row_dimensions[3].height = 26

        alternate_fill = PatternFill("solid", fgColor="F7FAFC")
        for row_index in range(4, 4 + len(dataframe)):
            for column_index in range(1, last_column + 1):
                cell = worksheet.cell(row=row_index, column=column_index)
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal=(
                        "left" if column_index in {2, 10} else "center"
                    ),
                    vertical="center",
                )
                if row_index % 2 == 0:
                    cell.fill = alternate_fill

        column_widths = (8, 24, 10, 17, 17, 16, 18, 18, 19, 18)
        for column_index, width in enumerate(column_widths, start=1):
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width
        worksheet.freeze_panes = "A4"
        worksheet.auto_filter.ref = (
            f"A3:{get_column_letter(last_column)}"
            f"{max(3, 3 + len(dataframe))}"
        )
        worksheet.sheet_view.showGridLines = False

    output.seek(0)
    return output.getvalue()


def _claim_document_source_label(document: dict[str, Any]) -> str:
    return (
        "홈택스"
        if str(document.get("source", "")) in {"hometax", "홈택스"}
        else "근로복지공단"
    )


def _mask_business_numbers_in_text(value: Any) -> str:
    text = _clean(value)
    return re.sub(
        r"(?<!\d)(\d{3})[- ]?(\d{2})[- ]?(\d{3})(\d{2})(?!\d)",
        lambda match: (
            f"{match.group(1)}-**-***{match.group(4)}"
        ),
        text,
    )


def _claim_document_scope_label(document: dict[str, Any]) -> str:
    facts = document.get("facts")
    if not isinstance(facts, dict):
        return ""
    label = _mask_business_numbers_in_text(
        facts.get("collection_scope_label")
    )
    label = re.sub(r"[\x00-\x1f<>]", " ", label)
    return re.sub(r"\s+", " ", label).strip()[:100]


def _claim_document_download_name(document: dict[str, Any]) -> str:
    facts = document.get("facts")
    raw_name = (
        _mask_business_numbers_in_text(facts.get("download_file_name"))
        if isinstance(facts, dict)
        else ""
    )
    content_type = str(document.get("content_type", "") or "").lower()
    extension = CLAIM_DOWNLOAD_EXTENSION_BY_CONTENT_TYPE.get(
        content_type,
        ".bin",
    )
    if not raw_name:
        document_code = _clean(document.get("document_code")) or "수집서류"
        raw_name = f"{document_code}{extension}"
    raw_name = raw_name.replace("\\", "/").split("/")[-1]
    safe_name = re.sub(r'[\x00-\x1f<>:"/\\|?*]', "_", raw_name).strip(" .")
    if not safe_name:
        safe_name = f"수집서류{extension}"
    if extension and not safe_name.lower().endswith(extension):
        safe_name = (
            f"{_clean(document.get('document_code')) or '수집서류'}"
            f"{extension}"
        )
    scope_label = _claim_document_scope_label(document)
    safe_scope = re.sub(
        r'[\x00-\x1f<>:"/\\|?*]',
        "_",
        scope_label,
    ).strip(" .")
    if safe_scope:
        safe_name = f"{safe_scope}_{safe_name}"
    maximum_length = 180
    if len(safe_name) <= maximum_length:
        return safe_name
    preserved_extension = (
        extension
        if extension and safe_name.lower().endswith(extension)
        else ""
    )
    if not preserved_extension:
        return safe_name[:maximum_length]
    stem_limit = maximum_length - len(preserved_extension)
    safe_stem = safe_name[:-len(preserved_extension)]
    safe_stem = safe_stem[:stem_limit].rstrip(" ._") or "수집서류"
    return f"{safe_stem}{preserved_extension}"


_CLAIM_ZIP_PART_MAX_FILES = 75
_CLAIM_ZIP_PART_TARGET_BYTES = 60 * 1024 * 1024
_CLAIM_ZIP_UNKNOWN_DOCUMENT_BYTES = 5 * 1024 * 1024
_CLAIM_ZIP_HARD_MAX_BYTES = 80 * 1024 * 1024


def _claim_document_estimated_zip_bytes(
    document: dict[str, Any],
) -> int:
    try:
        size_bytes = int(document.get("size_bytes") or 0)
    except (TypeError, ValueError):
        size_bytes = 0
    return (
        size_bytes
        if size_bytes > 0
        else _CLAIM_ZIP_UNKNOWN_DOCUMENT_BYTES
    )


def _claim_document_zip_sort_key(
    document: dict[str, Any],
) -> tuple[Any, ...]:
    try:
        period_year = int(document.get("period_year") or 0)
    except (TypeError, ValueError):
        period_year = 0
    return (
        _clean(document.get("source")),
        _clean(document.get("document_code")),
        -period_year,
        _clean(document.get("collection_key")),
        _clean(document.get("id")),
        _clean(document.get("storage_path")),
    )


def _plan_claim_document_zip_parts(
    documents: list[dict[str, Any]],
) -> list[list[dict[str, Any]]]:
    """Split every document into one deterministic, bounded ZIP part."""
    ordered_documents = sorted(
        (
            document
            for document in documents
            if isinstance(document, dict)
        ),
        key=_claim_document_zip_sort_key,
    )
    parts: list[list[dict[str, Any]]] = []
    current_part: list[dict[str, Any]] = []
    current_estimated_bytes = 0
    for document in ordered_documents:
        estimated_bytes = _claim_document_estimated_zip_bytes(document)
        exceeds_file_target = (
            len(current_part) >= _CLAIM_ZIP_PART_MAX_FILES
        )
        exceeds_size_target = bool(
            current_part
            and current_estimated_bytes + estimated_bytes
            > _CLAIM_ZIP_PART_TARGET_BYTES
        )
        if exceeds_file_target or exceeds_size_target:
            parts.append(current_part)
            current_part = []
            current_estimated_bytes = 0
        current_part.append(document)
        current_estimated_bytes += estimated_bytes
    if current_part:
        parts.append(current_part)
    return parts


def _build_claim_documents_zip(
    repository: ClaimRepository,
    case_id: str,
    documents: list[dict[str, Any]],
) -> bytes:
    from io import BytesIO
    from zipfile import ZIP_DEFLATED, ZipFile

    import requests

    archive = BytesIO()
    used_names: set[str] = set()
    total_bytes = 0
    try:
        document_ids = [
            str(document.get("id", "") or "")
            for document in documents
        ]
        download_urls = repository.document_download_urls(
            case_id,
            document_ids,
        )
        if len(download_urls) != len(documents):
            raise ClaimRepositoryError(
                "전체 서류의 다운로드 링크를 준비하지 못했습니다. "
                "잠시 후 다시 시도해 주세요."
            )
        with ZipFile(archive, "w", compression=ZIP_DEFLATED) as zip_file:
            for index, (document, download_url) in enumerate(
                zip(documents, download_urls),
                start=1,
            ):
                response = requests.get(
                    download_url,
                    timeout=(5, 45),
                    stream=True,
                    allow_redirects=True,
                )
                with response:
                    response.raise_for_status()
                    document_content = BytesIO()
                    for chunk in response.iter_content(chunk_size=64 * 1024):
                        if not chunk:
                            continue
                        total_bytes += len(chunk)
                        if total_bytes > _CLAIM_ZIP_HARD_MAX_BYTES:
                            raise ClaimRepositoryError(
                                "선택한 ZIP 묶음의 실제 용량이 80MB를 "
                                "초과했습니다. "
                                "개별 다운로드를 이용해 주세요."
                            )
                        document_content.write(chunk)
                    content = document_content.getvalue()

                archive_name = _claim_document_download_name(document)
                if archive_name in used_names:
                    if "." in archive_name:
                        stem, suffix = archive_name.rsplit(".", 1)
                        archive_name = f"{stem}_{index}.{suffix}"
                    else:
                        archive_name = f"{archive_name}_{index}"
                used_names.add(archive_name)
                zip_file.writestr(archive_name, content)
    except ClaimRepositoryError:
        raise
    except requests.RequestException as exc:
        raise ClaimRepositoryError(
            "전체 서류를 준비하지 못했습니다. 잠시 후 다시 시도해 주세요."
        ) from exc
    except Exception as exc:
        raise ClaimRepositoryError(
            "전체 서류 파일을 만들지 못했습니다. 개별 다운로드를 이용해 주세요."
        ) from exc
    return archive.getvalue()


_CLAIM_NO_DATA_REASON_LABELS = {
    "active_business_no_closure": "계속사업자로 확인되어 폐업사실증명 발급 대상이 아닙니다.",
    "no_income_tax_return": "해당 귀속연도의 종합소득세 신고서가 확인되지 않았습니다.",
    "no_management_number": "사업장관리번호가 확인되지 않아 사업장요율을 조회할 수 없습니다.",
    "no_remuneration_report": "해당 연도의 보수총액 신고내역이 확인되지 않았습니다.",
    "no_workplace_rate": "해당 연도의 사업장요율 자료가 확인되지 않았습니다.",
    "provider_no_records": "기관 조회 결과 해당 자료가 없습니다.",
}


def _claim_no_data_reason_label(document: dict[str, Any]) -> str:
    if _claim_document_needs_recollection(document):
        return "이전 신고일 조회기준으로 확인된 항목입니다. 새 인증 후 다시 조회해 주세요."
    facts = document.get("facts")
    reason = (
        str(facts.get("no_data_reason", "")).strip()
        if isinstance(facts, dict)
        else ""
    )
    if reason in _CLAIM_NO_DATA_REASON_LABELS:
        return _CLAIM_NO_DATA_REASON_LABELS[reason]
    document_code = str(document.get("document_code", "")).strip()
    fallback_by_code = {
        "hometax_closure_certificate": _CLAIM_NO_DATA_REASON_LABELS[
            "active_business_no_closure"
        ],
        "hometax_income_tax_return": _CLAIM_NO_DATA_REASON_LABELS[
            "no_income_tax_return"
        ],
        "comwel_total_remuneration": _CLAIM_NO_DATA_REASON_LABELS[
            "no_remuneration_report"
        ],
        "comwel_management_number_list": _CLAIM_NO_DATA_REASON_LABELS[
            "no_management_number"
        ],
        "comwel_workplace_rate": _CLAIM_NO_DATA_REASON_LABELS[
            "no_workplace_rate"
        ],
    }
    return fallback_by_code.get(
        document_code,
        _CLAIM_NO_DATA_REASON_LABELS["provider_no_records"],
    )


def _claim_pending_reason_label(document: dict[str, Any]) -> str:
    if str(document.get("document_code", "")).strip() == "comwel_workplace_rate":
        return "사업장관리번호 확인 후 조회할 항목입니다."
    return "아직 자동조회가 완료되지 않았습니다."


@st.dialog("서류조회", width="large")
def _show_claim_documents_dialog(
    user_id: str,
    repository: ClaimRepository,
    selected_case: dict[str, Any],
    provider_ready: bool,
) -> None:
    case_id = str(selected_case.get("id", "") or "")
    customer_name = _claim_result_customer_name(selected_case)
    st.markdown(f"### {html.escape(customer_name)} 고객 자료 목록")
    st.caption(
        f"{_clean(selected_case.get('business_no_masked')) or '사업자번호 확인 중'}"
        f" · {_source_status(selected_case.get('overall_status'))}"
    )

    try:
        documents = repository.list_documents(case_id)
    except ClaimRepositoryError as exc:
        st.error(str(exc))
        return
    if not documents:
        st.info("이 고객에게 등록된 수집 항목이 없습니다.")
        return

    progress_percent, _, _, target_count = _claim_collection_progress(
        documents
    )
    selected_job_snapshot = _claim_job_snapshot(user_id, case_id)
    _render_claim_collection_retry_action(
        user_id,
        selected_case,
        selected_job_snapshot,
        provider_ready=provider_ready,
        key_prefix="claim_result_dialog_retry",
        collection_complete=(target_count > 0 and progress_percent == 100),
    )

    active_documents = _claim_active_collection_documents(documents)
    ready_documents = _claim_downloadable_documents(documents)
    failed_count = sum(
        1
        for document in active_documents
        if str(document.get("status", "") or "") == "failed"
        and not _claim_document_is_no_data(document)
    )
    planned_count = sum(
        1
        for document in documents
        if str(document.get("status", "") or "") == "integration_required"
        or not automatic_collection_supported(
            str(document.get("document_code", "") or "")
        )
    )
    stale_documents = [
        document
        for document in active_documents
        if _claim_document_needs_recollection(document)
    ]
    no_data_documents = [
        document
        for document in active_documents
        if _claim_document_is_no_data(document)
        and not _claim_document_needs_recollection(document)
    ]
    pending_documents = [
        document
        for document in active_documents
        if str(document.get("status", "") or "").strip().lower()
        not in {"ready", "failed"}
    ]
    processed_count = sum(
        1
        for document in active_documents
        if str(document.get("status", "") or "").strip().lower() == "ready"
        and not _claim_document_needs_recollection(document)
    )
    expired_count = sum(
        1
        for document in documents
        if str(document.get("status", "") or "") == "ready"
        and bool(document.get("storage_path"))
        and not _claim_document_is_downloadable(document)
        and not _claim_document_is_no_data(document)
        and not _claim_document_needs_recollection(document)
    )
    summary_parts = [
        f"자동조회 {processed_count:,}/{len(active_documents):,}건 처리",
        f"다운로드 가능 {len(ready_documents):,}건",
    ]
    if no_data_documents:
        summary_parts.append(f"조회 내역 없음 {len(no_data_documents):,}건")
    if stale_documents:
        summary_parts.append(f"조회기준 수정 후 재확인 {len(stale_documents):,}건")
    if pending_documents:
        summary_parts.append(f"아직 조회되지 않음 {len(pending_documents):,}건")
    if failed_count:
        summary_parts.append(f"수집 실패 {failed_count:,}건")
    if expired_count:
        summary_parts.append(f"보관 만료 {expired_count:,}건")
    if planned_count:
        summary_parts.append(f"지원 예정 {planned_count:,}건")
    st.caption(" · ".join(summary_parts))

    detail_documents = no_data_documents + stale_documents
    if detail_documents:
        with st.expander(
            f"조회 내역 없음·재확인 상세 {len(detail_documents):,}건",
            expanded=False,
        ):
            st.dataframe(
                pd.DataFrame(
                    [
                        {
                            "자료명": (
                                _clean(document.get("document_name"))
                                or _clean(document.get("document_code"))
                                or "수집자료"
                            ),
                            "기관": _claim_document_source_label(document),
                            "사업자": (
                                _claim_document_scope_label(document) or "공통"
                            ),
                            "연도": document.get("period_year") or "-",
                            "확인 결과": _claim_no_data_reason_label(document),
                        }
                        for document in detail_documents
                    ]
                ),
                use_container_width=True,
                hide_index=True,
                height=min(520, 38 + len(detail_documents) * 35),
            )

    if pending_documents:
        with st.expander(
            f"아직 조회되지 않은 항목 {len(pending_documents):,}건",
            expanded=False,
        ):
            st.dataframe(
                pd.DataFrame(
                    [
                        {
                            "자료명": (
                                _clean(document.get("document_name"))
                                or _clean(document.get("document_code"))
                                or "수집자료"
                            ),
                            "기관": _claim_document_source_label(document),
                            "연도": document.get("period_year") or "-",
                            "상태": _claim_pending_reason_label(document),
                        }
                        for document in pending_documents
                    ]
                ),
                use_container_width=True,
                hide_index=True,
                height=min(420, 38 + len(pending_documents) * 35),
            )
    if not ready_documents:
        st.info(
            "아직 다운로드 가능한 수집 완료 서류가 없습니다. "
            "수집이 끝나면 이 창에 다운로드 버튼이 표시됩니다."
        )
        return

    document_rows = []
    for number, document in enumerate(ready_documents, start=1):
        size_bytes = int(document.get("size_bytes") or 0)
        size_label = (
            f"{size_bytes / (1024 * 1024):.1f}MB"
            if size_bytes >= 1024 * 1024
            else f"{max(1, round(size_bytes / 1024)):,}KB"
            if size_bytes
            else "-"
        )
        document_rows.append(
            {
                "번호": number,
                "자료명": (
                    _clean(document.get("document_name"))
                    or _clean(document.get("document_code"))
                    or "수집자료"
                ),
                "기관": _claim_document_source_label(document),
                "사업자": _claim_document_scope_label(document) or "공통",
                "연도": document.get("period_year") or "-",
                "수집일": _clean(document.get("collected_at"))[:10] or "-",
                "파일": size_label,
            }
        )
    st.dataframe(
        pd.DataFrame(document_rows),
        use_container_width=True,
        hide_index=True,
        height=min(620, 38 + len(document_rows) * 35),
    )

    st.markdown(f"#### 다운로드 가능한 서류 {len(ready_documents):,}건")
    zip_parts = _plan_claim_document_zip_parts(ready_documents)
    part_count = len(zip_parts)
    selected_part_index = 0
    if part_count > 1:
        st.info(
            f"전체 서류 {len(ready_documents):,}건을 빠짐없이 "
            f"{part_count:,}개 ZIP 묶음으로 나눴습니다. "
            "묶음을 하나씩 선택해 내려받아 주세요."
        )
        selected_part_index = int(
            st.selectbox(
                "ZIP 묶음 선택",
                options=list(range(part_count)),
                format_func=lambda index: (
                    f"ZIP {index + 1}/{part_count} · "
                    f"서류 {len(zip_parts[index]):,}건 · "
                    "예상 "
                    f"{sum(_claim_document_estimated_zip_bytes(item) for item in zip_parts[index]) / (1024 * 1024):.1f}MB"
                ),
                key=f"claim_zip_part_selector_{case_id}",
            )
        )
    else:
        st.caption(
            "전체 서류를 ZIP 파일 하나로 받거나, 필요한 서류만 개별로 "
            "받을 수 있습니다."
        )
    selected_part = zip_parts[selected_part_index]
    part_label = f"{selected_part_index + 1}/{part_count}"
    st.caption(
        f"현재 선택: ZIP {part_label} · 서류 {len(selected_part):,}건. "
        "개별 링크는 생성 후 1분간 유효합니다."
    )
    all_documents_fingerprint = _claim_download_cache_fingerprint(
        ready_documents
    )
    selected_documents_fingerprint = _claim_download_cache_fingerprint(
        selected_part
    )
    download_fingerprint = hashlib.sha256(
        "|".join(
            (
                all_documents_fingerprint,
                selected_documents_fingerprint,
                str(selected_part_index),
                str(part_count),
            )
        ).encode("utf-8")
    ).hexdigest()
    archive_key = f"_claim_bulk_download_{case_id}"
    cached_archive = st.session_state.get(archive_key)
    if not (
        isinstance(cached_archive, dict)
        and isinstance(cached_archive.get("data"), bytes)
        and str(cached_archive.get("fingerprint", ""))
        == download_fingerprint
        and float(cached_archive.get("expires_at") or 0) > time.time()
    ):
        st.session_state.pop(archive_key, None)
        cached_archive = None

    bulk_col, guide_col = st.columns([1.35, 3.65])
    with bulk_col:
        if st.button(
            (
                "전체 서류 ZIP 만들기"
                if part_count == 1
                else f"ZIP {part_label} 만들기"
            ),
            type="primary",
            use_container_width=True,
            key=f"claim_bulk_prepare_{case_id}",
        ):
            try:
                with st.spinner(
                    f"ZIP {part_label} 묶음을 안전하게 만들고 있습니다."
                ):
                    archive_data = _build_claim_documents_zip(
                        repository,
                        case_id,
                        selected_part,
                    )
                cached_archive = {
                    "data": archive_data,
                    "fingerprint": download_fingerprint,
                    "expires_at": time.time() + 5 * 60,
                }
                st.session_state[archive_key] = cached_archive
            except ClaimRepositoryError as exc:
                st.error(str(exc))
    with guide_col:
        st.caption(
            f"ZIP {part_label}만 현재 로그인 화면에서 5분간 준비 상태로 "
            "유지됩니다. 다른 묶음을 선택하면 이 캐시는 교체됩니다."
        )
    if cached_archive:
        safe_customer_name = re.sub(
            r'[\x00-\x1f<>:"/\\|?*]',
            "_",
            customer_name,
        ).strip(" .") or "고객"
        st.download_button(
            (
                "전체 ZIP 다운로드"
                if part_count == 1
                else f"ZIP {part_label} 다운로드"
            ),
            data=bytes(cached_archive["data"]),
            file_name=(
                f"{safe_customer_name}_수집서류_"
                + (
                    ""
                    if part_count == 1
                    else f"{selected_part_index + 1}of{part_count}_"
                )
                + f"{datetime.now(_KOREA_TIMEZONE):%Y%m%d}.zip"
            ),
            mime="application/zip",
            use_container_width=True,
            key=f"claim_bulk_download_button_{case_id}",
        )

    st.markdown("##### 개별 서류")
    for document in ready_documents:
        document_id = str(document.get("id", "") or "")
        document_name = (
            _clean(document.get("document_name"))
            or _clean(document.get("document_code"))
            or "수집자료"
        )
        scope_label = _claim_document_scope_label(document)
        year_label = str(document.get("period_year") or "공통")
        size_bytes = int(document.get("size_bytes") or 0)
        size_label = (
            f"{size_bytes / (1024 * 1024):.1f}MB"
            if size_bytes >= 1024 * 1024
            else f"{max(1, round(size_bytes / 1024)):,}KB"
            if size_bytes
            else "크기 미확인"
        )
        state_key = f"_claim_download_link_{case_id}_{document_id}"
        document_fingerprint = _claim_download_cache_fingerprint(
            [document]
        )
        cached_link = st.session_state.get(state_key)
        if not (
            isinstance(cached_link, dict)
            and _clean(cached_link.get("url"))
            and str(cached_link.get("fingerprint", ""))
            == document_fingerprint
            and float(cached_link.get("expires_at") or 0) > time.time()
        ):
            st.session_state.pop(state_key, None)
            cached_link = None

        with st.container(border=True):
            info_col, action_col = st.columns([4.7, 1.3])
            with info_col:
                st.markdown(f"**{html.escape(document_name)}**")
                st.caption(
                    f"{_claim_document_source_label(document)} · "
                    + (f"{scope_label} · " if scope_label else "")
                    + f"{year_label} · {size_label}"
                )
            with action_col:
                if not cached_link and st.button(
                    "다운로드 준비",
                    key=f"claim_download_prepare_{case_id}_{document_id}",
                    use_container_width=True,
                ):
                    try:
                        download_url = repository.document_download_url(
                            case_id,
                            document_id,
                        )
                        cached_link = {
                            "url": download_url,
                            "fingerprint": document_fingerprint,
                            "expires_at": (
                                time.time()
                                + CLAIM_DOWNLOAD_URL_TTL_SECONDS
                                - 5
                            ),
                        }
                        st.session_state[state_key] = cached_link
                    except ClaimRepositoryError as exc:
                        st.error(str(exc))
                if cached_link:
                    st.link_button(
                        "파일 다운로드",
                        str(cached_link["url"]),
                        use_container_width=True,
                    )


def _render_results_tab(
    user_id: str,
    repository: ClaimRepository | None,
    provider_ready: bool,
) -> None:
    if repository is None:
        st.info("전용 저장소 설치 후 수집결과가 표시됩니다.")
        return
    try:
        cases = repository.list_cases()
    except ClaimRepositoryError as exc:
        st.error(str(exc))
        return
    if not cases:
        st.info("수집결과를 확인할 경정청구 요청이 없습니다.")
        return

    st.markdown("### 진행상황 확인")
    filter_cols = st.columns([1.5, 1, 1])
    with filter_cols[0]:
        search_text = st.text_input(
            "고객 검색",
            placeholder="성명 또는 상호명을 입력하세요",
            key="claim_result_search_v2",
        )
    with filter_cols[1]:
        business_filter = st.selectbox(
            "사업자 구분",
            ["전체", "개인사업자", "법인사업자"],
            key="claim_result_business_filter_v2",
        )
    with filter_cols[2]:
        status_filter = st.selectbox(
            "진행 상태",
            [
                "전체",
                "인증 대기",
                "인증 완료",
                "일부 수집 실패",
                "수집 완료",
                "실패",
            ],
            key="claim_result_status_filter_v2",
        )

    filtered_cases = []
    search_key = search_text.strip().lower()
    for case in cases:
        case_view = _claim_result_case_view(case)
        customer_name = str(case_view["customer_name"])
        business_type = str(case.get("business_type", "") or "")
        status_group = str(case_view["overall_status"])
        if search_key and search_key not in customer_name.lower():
            continue
        if (
            business_filter == "개인사업자"
            and business_type != "individual"
        ):
            continue
        if (
            business_filter == "법인사업자"
            and business_type != "corporation"
        ):
            continue
        if status_filter != "전체" and status_group != status_filter:
            continue
        filtered_cases.append(case)

    if not filtered_cases:
        st.caption("TOTAL 0건")
        st.info("조건에 맞는 경정청구 요청이 없습니다.")
        return

    result_summary_col, excel_col = st.columns([4.5, 1.2])
    with result_summary_col:
        st.caption(
            f"TOTAL {len(filtered_cases):,}건 · "
            "수집 서류는 고객별 ‘서류조회’에서 다운로드합니다."
        )
    with excel_col:
        st.download_button(
            "Excel 다운로드",
            data=_build_claim_results_excel(filtered_cases),
            file_name=(
                "OASIS_수집결과_"
                f"{datetime.now(_KOREA_TIMEZONE):%Y%m%d}.xlsx"
            ),
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            key="claim_result_excel_download_v1",
            use_container_width=True,
        )

    page_size = 20
    page_count = max(1, (len(filtered_cases) + page_size - 1) // page_size)
    page_number = 1
    if page_count > 1:
        page_number = st.selectbox(
            "페이지",
            list(range(1, page_count + 1)),
            format_func=lambda page: f"{page:,} / {page_count:,}",
            key="claim_result_page_v2",
        )
    start_index = (int(page_number) - 1) * page_size
    visible_cases = filtered_cases[start_index : start_index + page_size]

    widths = [
        0.4,
        1.45,
        0.75,
        0.65,
        1.0,
        1.05,
        0.9,
        1.05,
        1.0,
        1.15,
        0.9,
    ]
    headers = [
        "번호",
        "고객 / 상호",
        "서류 조회",
        "구분",
        "사업자번호",
        "휴대전화",
        "홈택스",
        "근로복지공단",
        "수집 상태",
        "등록일",
        "담당자",
    ]
    header_cols = st.columns(widths)
    for column, label in zip(header_cols, headers):
        column.markdown(f"**{label}**")

    for offset, case in enumerate(visible_cases, start=start_index + 1):
        case_id = str(case.get("id", "") or "")
        case_view = _claim_result_case_view(case)
        customer_name = str(case_view["customer_name"])
        with st.container(border=True):
            row_cols = st.columns(widths)
            row_cols[0].write(offset)
            row_cols[1].markdown(f"**{html.escape(customer_name)}**")
            if row_cols[2].button(
                "서류조회",
                key=f"claim_result_documents_{case_id}",
                use_container_width=True,
            ):
                _show_claim_documents_dialog(
                    user_id,
                    repository,
                    case,
                    provider_ready,
                )
            row_cols[3].write(
                str(case_view["business_type"])
            )
            row_cols[4].write(str(case_view["business_no"]))
            row_cols[5].write(str(case_view["phone"]))
            row_cols[6].write(str(case_view["hometax_status"]))
            row_cols[7].write(str(case_view["comwel_status"]))
            row_cols[8].write(str(case_view["overall_status"]))
            row_cols[9].write(str(case_view["requested_at"]))
            row_cols[10].write(str(case_view["requested_by"]))


def _render_catalog_tab() -> None:
    rows = [
        {
            "자료명": spec.name,
            "기관": spec.source,
            "기간": spec.period,
            "자동수집": (
                "지원"
                if spec.code in AUTOMATIC_COLLECTION_CODES
                else "연동 예정"
            ),
            "설명": spec.description,
        }
        for spec in DOCUMENT_SPECS
    ]
    st.dataframe(
        pd.DataFrame(rows),
        use_container_width=True,
        hide_index=True,
    )
    st.warning(
        "기관 또는 중계 API가 실제로 제공하는 조회 기간이 7개년보다 짧으면 "
        "가능한 연도만 수집하고, 누락 연도를 결과 화면에 명확히 표시합니다."
    )


@st.dialog("자료수집 완료")
def _show_collection_complete_dialog(message: str) -> None:
    st.success(message)
    st.caption("수집된 문서는 ‘수집결과’ 탭에서 확인할 수 있습니다.")
    if st.button(
        "확인",
        use_container_width=True,
        type="primary",
        key="claim_collection_complete_dialog_close",
    ):
        # Streamlit dialogs close on a rerun; this refresh is intentional.
        st.rerun()


def render_claim_correction_center(
    user_id: str,
    user_name: str = "",
) -> None:
    repository, repository_message = _repository(user_id)
    readiness = provider_readiness()
    _render_intro(
        repository is not None,
        repository_message,
        readiness,
    )
    flash = st.session_state.pop("_claim_flash_v1", "")
    if flash:
        st.success(flash)
    completion_message = st.session_state.pop(
        "_claim_collection_completed_v1",
        "",
    )
    if completion_message:
        _show_collection_complete_dialog(str(completion_message))

    active_case_id = str(
        st.session_state.get("_claim_active_case_v1", "") or ""
    )
    if repository is not None and active_case_id:
        st.markdown("### 인증 및 자료수집 진행")
        _render_auto_claim_monitor(
            user_id,
            active_case_id,
            repository,
            bool(readiness.get("simple_auth_ready")),
        )

    active_section = (
        st.segmented_control(
            "경정청구 화면",
            ["인증 요청", "진행상황", "수집결과", "수집 항목"],
            default="인증 요청",
            key="claim_correction_section_v1032",
            label_visibility="collapsed",
        )
        or "인증 요청"
    )
    if active_section == "인증 요청":
        _render_request_tab(
            user_id,
            user_name,
            repository,
            readiness,
        )
    elif active_section == "진행상황":
        _render_status_tab(
            user_id,
            repository,
            bool(readiness.get("simple_auth_ready")),
        )
    elif active_section == "수집결과":
        _render_results_tab(
            user_id,
            repository,
            bool(readiness.get("simple_auth_ready")),
        )
    else:
        _render_catalog_tab()
