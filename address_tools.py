from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from utils import get_user_cumulative_db_path


REGION_ALIASES = {
    "서울특별시": "서울",
    "서울시": "서울",
    "부산광역시": "부산",
    "부산시": "부산",
    "대구광역시": "대구",
    "대구시": "대구",
    "인천광역시": "인천",
    "인천시": "인천",
    "광주광역시": "광주",
    "광주시": "광주",
    "대전광역시": "대전",
    "대전시": "대전",
    "울산광역시": "울산",
    "울산시": "울산",
    "세종특별자치시": "세종",
    "세종시": "세종",
    "경기도": "경기",
    "강원특별자치도": "강원",
    "강원도": "강원",
    "충청북도": "충북",
    "충청남도": "충남",
    "전북특별자치도": "전북",
    "전라북도": "전북",
    "전라남도": "전남",
    "경상북도": "경북",
    "경상남도": "경남",
    "제주특별자치도": "제주",
    "제주도": "제주",
}

SHORT_REGIONS = {
    "서울", "부산", "대구", "인천", "광주", "대전", "울산",
    "세종", "경기", "강원", "충북", "충남", "전북", "전남",
    "경북", "경남", "제주",
}


def _blank(value: Any) -> bool:
    return value is None or str(value).strip().lower() in {
        "", "nan", "none", "nat"
    }


def clean_address(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^\(\d{5}\)\s*", "", text)
    return re.sub(r"\s+", " ", text).strip()


def split_korean_address(value: Any) -> tuple[str, str]:
    address = clean_address(value)
    if not address:
        return "", ""

    parts = address.split()
    if not parts:
        return "", ""

    first = parts[0]
    sido = REGION_ALIASES.get(first, first if first in SHORT_REGIONS else "")

    sigungu = ""
    if len(parts) >= 2:
        second = parts[1]
        if re.search(r"(시|군|구)$", second):
            sigungu = second
        elif sido == "세종":
            sigungu = "세종시"

    return sido, sigungu


def enrich_address_fields(data: dict[str, Any]) -> dict[str, Any]:
    result = dict(data or {})
    address = (
        result.get("사업장 소재지")
        or result.get("주소")
        or result.get("본점소재지")
        or ""
    )
    sido, sigungu = split_korean_address(address)

    if _blank(result.get("시도")) and sido:
        result["시도"] = sido
    if _blank(result.get("시군구")) and sigungu:
        result["시군구"] = sigungu
    return result


def repair_user_customer_addresses(user_id: str) -> dict[str, Any]:
    """
    기존 고객DB에서 시도·시군구가 비어 있는 행만 보완한다.
    기존 주소와 기존 입력값은 덮어쓰지 않는다.
    """
    path = get_user_cumulative_db_path(user_id)
    result = {
        "updated_rows": 0,
        "path": str(path),
        "message": "",
    }
    if not path.exists():
        result["message"] = "누적 고객DB가 없습니다."
        return result

    backup_dir = path.parent / "_address_backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / (
        f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
    )
    shutil.copy2(path, backup_path)

    workbook = load_workbook(path)
    if "고객DB" not in workbook.sheetnames:
        result["message"] = "고객DB 시트를 찾지 못했습니다."
        return result

    worksheet = workbook["고객DB"]
    header_row = None
    columns: dict[str, int] = {}

    for row_number in range(1, min(worksheet.max_row, 50) + 1):
        row_columns = {}
        for column_number in range(1, worksheet.max_column + 1):
            value = str(
                worksheet.cell(row_number, column_number).value or ""
            ).strip()
            if value:
                row_columns[value] = column_number
        if "업체명" in row_columns:
            header_row = row_number
            columns = row_columns
            break

    if header_row is None:
        result["message"] = "고객DB 헤더를 찾지 못했습니다."
        return result

    for name in ["시도", "시군구", "사업장 소재지"]:
        if name not in columns:
            column_number = worksheet.max_column + 1
            worksheet.cell(header_row, column_number).value = name
            columns[name] = column_number

    updated_rows = 0
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        address = worksheet.cell(
            row_number,
            columns["사업장 소재지"],
        ).value
        if _blank(address):
            continue

        sido, sigungu = split_korean_address(address)
        changed = False

        sido_cell = worksheet.cell(row_number, columns["시도"])
        sigungu_cell = worksheet.cell(row_number, columns["시군구"])

        if _blank(sido_cell.value) and sido:
            sido_cell.value = sido
            changed = True
        if _blank(sigungu_cell.value) and sigungu:
            sigungu_cell.value = sigungu
            changed = True

        if changed:
            updated_rows += 1

    workbook.save(path)
    result["updated_rows"] = updated_rows
    result["message"] = (
        f"주소가 있는 고객 중 {updated_rows}건의 시도·시군구를 보완했습니다."
    )
    return result
