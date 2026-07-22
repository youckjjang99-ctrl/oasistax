from __future__ import annotations

import html
import json
import math
import re
from datetime import datetime
from io import BytesIO
from pathlib import Path
from runtime_error_log import write_runtime_error
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from matching_preferences import get_matching_preferences
from cloud_sync import (
    load_financial_snapshot,
    load_registry_snapshot,
    load_stock_valuations,
)
from consultation_journal import load_company_consultation_context
from employee_status import get_latest_employee_status
from registered_policy_match import (
    build_customer_labels,
    load_registered_customers,
)
from utils import get_user_cumulative_db_path, get_user_dirs
from pdf_report import build_representative_pdf
from tax_diagnosis import build_tax_diagnosis
from comprehensive_financial_diagnosis import (
    build_comprehensive_financial_diagnosis,
)




def _metric_card(label: str, value: str, note: str = "", tone: str = "blue") -> str:
    palette = {
        "blue": ("#EAF2FF", "#1E5BD7"),
        "green": ("#EAF8F2", "#16835F"),
        "purple": ("#F3EDFF", "#7048C8"),
        "orange": ("#FFF4E7", "#B56716"),
        "red": ("#FFF0F0", "#C43D3D"),
    }
    background, accent = palette.get(tone, palette["blue"])
    return f"""
    <div class="oasis-kpi-card" style="--card-bg:{background};--card-accent:{accent};">
      <div class="oasis-kpi-label">{html.escape(label)}</div>
      <div class="oasis-kpi-value">{html.escape(value)}</div>
      <div class="oasis-kpi-note">{html.escape(note or '현재 저장자료 기준')}</div>
    </div>
    """


def _inject_report_css() -> None:
    st.markdown(
        """
        <style>
        .oasis-kpi-card {
          min-height:132px; padding:18px 18px 15px; border-radius:16px;
          border:1px solid color-mix(in srgb, var(--card-accent) 24%, white);
          background:linear-gradient(145deg,#fff,var(--card-bg));
          box-shadow:0 7px 20px rgba(15,42,80,.08); margin-bottom:8px;
          position:relative; overflow:hidden;
        }
        .oasis-kpi-card:before {content:"";position:absolute;left:0;top:0;bottom:0;width:5px;background:var(--card-accent);}
        .oasis-kpi-label {font-size:.87rem;font-weight:750;color:#475467;margin-bottom:13px;}
        .oasis-kpi-value {font-size:1.78rem;line-height:1.1;font-weight:850;color:#172033;letter-spacing:-.04em;word-break:keep-all;}
        .oasis-kpi-note {font-size:.76rem;color:var(--card-accent);font-weight:700;margin-top:12px;}
        .oasis-section-card {border:1px solid #dbe4ef;border-radius:16px;background:#fff;padding:18px 20px;box-shadow:0 5px 16px rgba(15,42,80,.055);height:100%;}
        .oasis-section-card h4 {margin:0 0 12px;color:#0b2b5b;}
        .oasis-section-card ul {padding-left:20px;margin:0;}
        .oasis-section-card li {margin-bottom:7px;color:#344054;}
        </style>
        """,
        unsafe_allow_html=True,
    )


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"", "nan", "none", "nat"}:
        return ""
    return text


def _number(value: Any) -> float | None:
    text = _clean(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _format_money(value: Any) -> str:
    number = _number(value)
    if number is None:
        return "-"
    return f"{int(round(number)):,}원"


def _normalize_business_no(value: Any) -> str:
    digits = re.sub(r"[^0-9]", "", str(value or ""))
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
    return _clean(value)


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def _latest_stock_record(
    user_id: str,
    business_no: str,
    company_name: str = "",
) -> dict[str, Any]:
    dirs = get_user_dirs(user_id)
    local_records = _load_json(
        dirs["base"] / "stock_valuations.json",
        [],
    )
    if not isinstance(local_records, list):
        local_records = []

    try:
        cloud_records = load_stock_valuations(user_id, limit=500)
    except Exception:
        cloud_records = []

    normalized = _normalize_business_no(business_no)
    clean_company = _clean(company_name)
    merged: dict[str, dict[str, Any]] = {}

    for record in [*cloud_records, *local_records]:
        if not isinstance(record, dict):
            continue

        record_no = _normalize_business_no(record.get("business_no", ""))
        record_company = _clean(record.get("company_name", ""))
        number_match = bool(normalized and record_no and normalized == record_no)
        name_match = bool(
            clean_company and record_company and clean_company == record_company
        )
        if not number_match and not name_match:
            continue

        record_id = _clean(record.get("record_id", ""))
        if not record_id:
            record_id = "|".join([
                record_no,
                _clean(record.get("valuation_date", "")),
                _clean(record.get("saved_at", "")),
            ])
        merged[record_id] = dict(record)

    matches = list(merged.values())
    if not matches:
        return {}

    return sorted(
        matches,
        key=lambda record: str(
            record.get("saved_at", "")
            or record.get("valuation_date", "")
        ),
        reverse=True,
    )[0]


def _financial_snapshot(
    user_id: str,
    business_no: str,
) -> dict[str, Any]:
    dirs = get_user_dirs(user_id)
    cache = _load_json(
        dirs["base"] / "stock_financial_cache.json",
        {},
    )
    if not isinstance(cache, dict):
        cache = {}

    normalized = _normalize_business_no(business_no)
    local = cache.get(normalized, {}) or {}
    if local:
        return local

    try:
        cloud = load_financial_snapshot(user_id, normalized)
    except Exception:
        cloud = {}
    return cloud if isinstance(cloud, dict) else {}


def _registry_snapshot(
    user_id: str,
    business_no: str,
) -> dict[str, Any]:
    dirs = get_user_dirs(user_id)
    cache = _load_json(
        dirs["base"] / "registry_cache.json",
        {},
    )
    if not isinstance(cache, dict):
        cache = {}

    normalized = _normalize_business_no(business_no)
    local = cache.get(normalized, {}) or {}
    if local:
        return local

    try:
        cloud = load_registry_snapshot(user_id, normalized)
    except Exception:
        cloud = {}
    return cloud if isinstance(cloud, dict) else {}


def _as_list(value: Any) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        return [
            _clean(item)
            for item in value
            if _clean(item)
        ]
    text = _clean(value)
    return [text] if text else []


def _build_consultation_context(
    user_id: str,
    business_no: str,
    company_name: str,
) -> dict[str, Any]:
    try:
        journals = load_company_consultation_context(
            user_id,
            business_no,
            company_name=company_name,
            limit=50,
        )
    except Exception:
        journals = []

    if not journals:
        return {
            "count": 0,
            "latest_saved_at": "",
            "latest_summary": "",
            "matching_keywords": [],
            "interest_fields": [],
            "client_needs": [],
            "next_actions": [],
            "transcript_count": 0,
            "records": [],
        }

    matching_keywords: list[str] = []
    interest_fields: list[str] = []
    client_needs: list[str] = []
    next_actions: list[str] = []
    transcript_count = 0

    for journal in journals:
        if not isinstance(journal, dict):
            continue
        matching_keywords.extend(
            _as_list(
                journal.get("matching_keywords")
                or journal.get("매칭키워드")
            )
        )
        interest_fields.extend(
            _as_list(
                journal.get("interest_fields")
                or journal.get("관심지원분야")
            )
        )
        client_needs.extend(
            _as_list(
                journal.get("client_needs")
                or journal.get("대표자니즈")
                or journal.get("key_needs")
            )
        )
        next_actions.extend(
            _as_list(
                journal.get("next_actions")
                or journal.get("후속조치")
                or journal.get("next_action")
            )
        )
        if _clean(journal.get("transcript", "")):
            transcript_count += 1

    latest = journals[0] if journals else {}
    latest_summary = _clean(
        latest.get("summary", "")
        or latest.get("consultation_summary", "")
        or latest.get("상담요약", "")
    )

    return {
        "count": len(journals),
        "latest_saved_at": _clean(latest.get("saved_at", "")),
        "latest_summary": latest_summary,
        "matching_keywords": list(dict.fromkeys(matching_keywords)),
        "interest_fields": list(dict.fromkeys(interest_fields)),
        "client_needs": list(dict.fromkeys(client_needs)),
        "next_actions": list(dict.fromkeys(next_actions)),
        "transcript_count": transcript_count,
        "records": journals,
    }


def _value_from_sources(
    customer: pd.Series,
    financial: dict[str, Any],
    *keys: str,
) -> Any:
    for key in keys:
        value = customer.get(key, "")
        if _clean(value):
            return value
    for key in keys:
        value = financial.get(key, "")
        if _clean(value):
            return value
    return ""


def _ratio(numerator: Any, denominator: Any) -> float | None:
    num = _number(numerator)
    den = _number(denominator)
    if num is None or den in (None, 0):
        return None
    return num / den * 100


def build_consulting_analysis(
    customer: pd.Series,
    financial: dict[str, Any],
    registry: dict[str, Any],
    stock_record: dict[str, Any],
    preferences: dict[str, Any],
    consultation_context: dict[str, Any] | None = None,
    articles_review: dict[str, Any] | None = None,
    employee_context: dict[str, Any] | None = None,
) -> dict[str, Any]:
    company_name = _clean(customer.get("업체명", "")) or _clean(
        financial.get("업체명", "")
    )
    business_no = _normalize_business_no(
        customer.get("사업자등록번호", "")
    )
    industry = _clean(customer.get("업종명", ""))
    address = _clean(
        _value_from_sources(
            customer,
            financial,
            "사업장 소재지",
        )
    )
    establishment = _clean(
        _value_from_sources(
            customer,
            financial,
            "설립일",
            "설립년도",
        )
    )
    cretop_employees = _value_from_sources(
        customer,
        financial,
        "종업원수",
        "상시근로자수",
    )

    employee_context = (
        employee_context if isinstance(employee_context, dict) else {}
    )
    employee_summary = (
        employee_context.get("summary", {})
        if isinstance(employee_context.get("summary"), dict)
        else {}
    )
    employee_records = (
        employee_context.get("employees", [])
        if isinstance(employee_context.get("employees"), list)
        else []
    )
    employee_active_count = int(
        employee_summary.get("active_count", 0) or 0
    )
    employee_total_count = int(
        employee_summary.get("total_count", 0) or len(employee_records)
    )
    employee_source_available = bool(
        employee_context
        and (
            employee_active_count
            or employee_total_count
            or employee_records
        )
    )
    employees = (
        employee_active_count
        if employee_source_available
        else cretop_employees
    )
    employees_source = (
        "직원현황"
        if employee_source_available
        else "크레탑·고객DB"
    )

    sales = _value_from_sources(
        customer,
        financial,
        "매출액",
        "연매출",
        "전년도매출",
    )
    operating_profit = _value_from_sources(
        customer,
        financial,
        "영업이익",
    )
    net_income = _value_from_sources(
        customer,
        financial,
        "당기순이익",
    )
    assets = _value_from_sources(
        customer,
        financial,
        "자산총계",
    )
    liabilities = _value_from_sources(
        customer,
        financial,
        "부채총계",
    )
    equity = _value_from_sources(
        customer,
        financial,
        "자본총계",
    )

    operating_margin = _ratio(operating_profit, sales)
    net_margin = _ratio(net_income, sales)
    debt_ratio = _ratio(liabilities, equity)

    strengths: list[str] = []
    cautions: list[str] = []
    strategy: list[str] = []
    questions: list[str] = []

    comprehensive_diagnosis = build_comprehensive_financial_diagnosis(
        customer,
        financial,
    )
    diagnosis_findings = (
        comprehensive_diagnosis.get("findings", [])
        if isinstance(comprehensive_diagnosis, dict)
        else []
    )

    for finding in diagnosis_findings[:5]:
        cautions.append(
            f"[{finding.get('status', '확인 필요')}] "
            f"{finding.get('title', '')} — "
            f"중요도 {finding.get('importance', 0)}점, "
            f"확신도 {finding.get('confidence', 0)}점"
        )
        facts = finding.get("facts", []) or []
        if facts:
            strategy.append(
                f"{finding.get('title', '재무진단')}: "
                + " ".join(str(item) for item in facts[:2])
            )

    questions.extend(
        comprehensive_diagnosis.get("questions", [])[:15]
        if isinstance(comprehensive_diagnosis, dict)
        else []
    )

    consultation_context = (
        consultation_context if isinstance(consultation_context, dict) else {}
    )
    consultation_count = int(consultation_context.get("count", 0) or 0)
    latest_consultation = _clean(consultation_context.get("latest_saved_at", ""))
    latest_summary = _clean(consultation_context.get("latest_summary", ""))
    consultation_keywords = consultation_context.get("matching_keywords", []) or []
    consultation_interests = consultation_context.get("interest_fields", []) or []
    consultation_needs = consultation_context.get("client_needs", []) or []
    consultation_actions = consultation_context.get("next_actions", []) or []

    articles_review = (
        articles_review if isinstance(articles_review, dict) else {}
    )
    articles_score = int(articles_review.get("score", 0) or 0)
    articles_priorities = articles_review.get("priority_items", []) or []
    articles_opportunities = (
        articles_review.get("consulting_opportunities", []) or []
    )

    sales_number = _number(sales)
    operating_number = _number(operating_profit)
    net_number = _number(net_income)
    employees_number = _number(employees)

    cretop_employees_number = _number(cretop_employees)
    youth_count = int(employee_summary.get("youth_count", 0) or 0)
    middle_aged_count = int(
        employee_summary.get("middle_aged_count", 0) or 0
    )
    senior_count = int(employee_summary.get("senior_count", 0) or 0)
    recent_3m_count = int(
        employee_summary.get("recent_3m_count", 0) or 0
    )
    recent_6m_count = int(
        employee_summary.get("recent_6m_count", 0) or 0
    )
    long_tenure_count = int(
        employee_summary.get("long_tenure_count", 0) or 0
    )

    workplace_counts: dict[str, int] = {}
    for employee in employee_records:
        if not isinstance(employee, dict):
            continue
        if not bool(employee.get("active", True)):
            continue
        workplace = _clean(
            employee.get("workplace_name", "")
            or employee.get("workplace_management_no", "")
            or "소속 미확인"
        )
        workplace_counts[workplace] = workplace_counts.get(workplace, 0) + 1

    if sales_number:
        strengths.append(
            f"최근 확인 매출액은 {_format_money(sales)}입니다."
        )
    if operating_number is not None and operating_number > 0:
        strengths.append(
            f"영업이익 {_format_money(operating_profit)}으로 "
            "영업단계 흑자가 확인됩니다."
        )
    if net_number is not None and net_number > 0:
        strengths.append(
            f"당기순이익 {_format_money(net_income)}으로 "
            "최종 손익도 흑자입니다."
        )
    if operating_margin is not None:
        strengths.append(
            f"영업이익률은 약 {operating_margin:.1f}%입니다."
        )
    if employees_number:
        strengths.append(
            f"확인된 종업원수는 {int(employees_number):,}명이며 "
            f"{employees_source} 자료를 기준으로 반영했습니다."
        )

    if employee_source_available:
        if youth_count:
            strengths.append(
                f"직원현황 기준 청년 연령구간 직원은 {youth_count}명입니다."
            )
        if senior_count:
            strengths.append(
                f"직원현황 기준 60세 이상 고령자 추정 직원은 "
                f"{senior_count}명입니다."
            )
        if recent_6m_count:
            strengths.append(
                f"최근 6개월 자격취득 직원이 {recent_6m_count}명 확인됩니다."
            )
        if long_tenure_count:
            strengths.append(
                f"2년 이상 장기근속 직원은 {long_tenure_count}명입니다."
            )

        if (
            cretop_employees_number is not None
            and int(cretop_employees_number) != employee_active_count
        ):
            cautions.append(
                f"크레탑 종업원수 {int(cretop_employees_number):,}명과 "
                f"최신 직원현황 가입중 {employee_active_count:,}명이 다릅니다. "
                "컨설팅과 고용지원금 검토에는 최신 직원현황을 우선 적용했습니다."
            )
    else:
        cautions.append(
            "저장된 직원현황이 없어 종업원수와 고용분석은 "
            "크레탑·고객DB 자료만 반영했습니다."
        )

    if not address:
        cautions.append(
            "사업장 소재지가 비어 있어 지역제한 사업 검토 전 보완이 필요합니다."
        )
    if not establishment:
        cautions.append(
            "설립일이 비어 있어 업력 요건 확인이 필요합니다."
        )
    if net_number is None:
        cautions.append(
            "당기순이익 자료가 없어 수익성 검토가 제한됩니다."
        )
    elif net_number < 0:
        cautions.append(
            "당기순손실 상태이므로 정책자금 심사 전 원인 확인이 필요합니다."
        )
    if debt_ratio is not None:
        if debt_ratio >= 300:
            cautions.append(
                f"부채비율이 약 {debt_ratio:.1f}%로 높아 "
                "보증·대출 심사 전 재무구조 확인이 필요합니다."
            )
        else:
            strengths.append(
                f"부채비율은 약 {debt_ratio:.1f}%입니다."
            )

    interests = preferences.get("관심지원분야", []) or []
    matching_keywords = preferences.get("매칭키워드", []) or []
    exclusions = preferences.get("제외키워드", []) or []
    fund_purpose = _clean(preferences.get("자금사용목적", ""))

    if "운전자금" in interests or "운전자금" in matching_keywords:
        strategy.append(
            "운전자금 수요를 중심으로 정책자금·보증기관 연계 가능성을 우선 검토합니다."
        )
    if any(
        item in interests
        for item in ["시설자금", "기계·설비 구입", "차량 구입"]
    ):
        strategy.append(
            "시설·설비·차량 투자계획과 견적서를 먼저 확보해 시설자금 가능성을 검토합니다."
        )
    if "신규채용" in interests or "고용유지" in interests:
        strategy.append(
            "직원현황의 연령·자격취득일·근속기간을 기준으로 "
            "고용지원금과 고용증대 세액공제를 함께 검토합니다."
        )

    if employee_source_available:
        if recent_6m_count:
            strategy.append(
                f"최근 6개월 취득 직원 {recent_6m_count}명에 대해 "
                "채용일 전 구직등록, 정규직 여부, 근로시간과 임금요건을 확인합니다."
            )
        if youth_count:
            strategy.append(
                f"청년 추정 직원 {youth_count}명의 실제 나이와 입사 당시 "
                "취업애로청년 요건을 확인해 청년일자리도약장려금을 검토합니다."
            )
        if senior_count:
            strategy.append(
                f"고령자 추정 직원 {senior_count}명에 대해 고령자 고용지원금과 "
                "계속고용장려금 적용 가능성을 구분해 검토합니다."
            )
        if workplace_counts:
            strategy.append(
                "사업장별 가입인원: "
                + ", ".join(
                    f"{name} {count}명"
                    for name, count in sorted(workplace_counts.items())
                )
            )
    if "연구개발" in interests or "특허·인증" in interests:
        strategy.append(
            "연구개발 조직과 지식재산 보유현황을 확인해 R&D·인증 지원사업을 검토합니다."
        )
    saved_policy_items = preferences.get("저장정책자금", []) or []
    saved_policy_items = [
        item for item in saved_policy_items if isinstance(item, dict)
    ]
    if saved_policy_items:
        strengths.append(
            f"기업컨설팅에서 확정 저장한 정책자금 추천 "
            f"{len(saved_policy_items)}건이 반영되었습니다."
        )
        for item in saved_policy_items[:8]:
            title = _clean(item.get("title", ""))
            score = item.get("score", "")
            category = _clean(item.get("category", ""))
            agency = _clean(item.get("agency", ""))
            if not title:
                continue
            note_parts = [
                f"{score}점" if str(score).strip() else "",
                category,
                agency,
            ]
            note = " · ".join(
                part for part in note_parts if part
            )
            strategy.append(
                "저장 정책자금 추천: "
                + title
                + (f" ({note})" if note else "")
            )

    if fund_purpose:
        strategy.append(
            f"대표가 입력한 자금사용목적은 '{fund_purpose}'입니다."
        )

    if articles_review:
        strengths.append(f"정관검토 결과 {articles_score}점이 반영되었습니다.")
        if articles_score < 70:
            cautions.append(
                "정관의 핵심 절세·퇴직·유족보상·승계 조항 보완이 필요합니다."
            )
        for item in articles_priorities[:4]:
            title = str(item.get("title", "") or "").strip()
            if title:
                strategy.append(f"정관 우선검토: {title}")
        for opportunity in articles_opportunities[:3]:
            strategy.append(f"연계 컨설팅: {opportunity}")
        questions.append(
            "정관 개정일, 주주총회 의사록, 별도 임원규정의 실제 제정 여부를 확인해 주세요."
        )
    else:
        cautions.append("저장된 정관검토 결과가 없어 정관 리스크는 미반영 상태입니다.")

    if consultation_count:
        strengths.append(
            f"최근 상담이 {consultation_count}건 축적되어 있으며 "
            f"최신 상담일은 {latest_consultation[:10] or '미확인'}입니다."
        )
        if latest_summary:
            strategy.append("최근 상담 요약 반영: " + latest_summary[:220])
        if consultation_keywords:
            strategy.append(
                "상담 기반 정책자금 키워드: "
                + ", ".join(str(item) for item in consultation_keywords[:12])
            )
        if consultation_interests:
            strategy.append(
                "상담 기반 관심지원분야: "
                + ", ".join(str(item) for item in consultation_interests[:8])
            )
        if consultation_needs:
            questions.append(
                "최근 상담에서 확인된 고객 니즈가 현재도 유효한지 확인해 주세요: "
                + ", ".join(str(item) for item in consultation_needs[:5])
            )
        if consultation_actions:
            strategy.append(
                "상담 후속조치: "
                + ", ".join(str(item) for item in consultation_actions[:5])
            )
    else:
        cautions.append(
            "저장된 상담일지가 없어 크레탑·등기·주가자료 중심으로만 진단했습니다."
        )

    if registry:
        strengths.append("등기정보가 종합진단에 반영되었습니다.")
    else:
        cautions.append("등기정보가 없어 지배구조·자본정보 진단이 제한됩니다.")

    if stock_record:
        strengths.append("최신 주가평가 결과가 종합진단에 반영되었습니다.")
    else:
        cautions.append("저장된 주가평가 결과가 없어 기업가치 진단이 제한됩니다.")

    if not strategy:
        strategy.append(
            "운전자금·시설투자·채용계획을 확인한 뒤 지원사업 우선순위를 정하는 것이 좋습니다."
        )

    if employee_source_available:
        questions.extend([
            "최근 자격취득 직원들의 정규직 여부, 주당 근로시간과 월 급여는 어떻게 됩니까?",
            "청년 직원 중 채용 전 4개월 이상 실업, 고졸 이하, 자립준비청년 등 취업애로 요건에 해당하는 직원이 있습니까?",
            "60세 이상 직원 수가 과거 기준기간보다 증가했거나 정년연장·재고용 제도를 운영하고 있습니까?",
            "육아휴직·근로시간 단축·대체인력 또는 유연근무를 사용하는 직원이 있습니까?",
        ])

    questions.extend([
        "최근 1년 이내 기계·시설·차량 투자계획이 있습니까?",
        "향후 6개월 이내 신규채용 또는 고용유지 계획이 있습니까?",
        "현재 정책자금·보증기관 대출 잔액과 만기는 어떻게 됩니까?",
        "국세·지방세 체납이나 최근 연체 이력이 있습니까?",
        "올해 예상 매출과 주요 거래처 변화는 어떻게 됩니까?",
    ])

    if exclusions:
        questions.append(
            "제외한 지원분야가 현재 사업계획과 완전히 무관한지 다시 확인할 필요가 있습니다."
        )

    stock_result = stock_record.get("result", {}) if stock_record else {}
    stock_summary = {
        "평가기준일": stock_record.get("valuation_date", ""),
        "발행주식총수": stock_record.get("current_shares", ""),
        "1주당평가액": stock_result.get("final_value_per_share"),
        "기업전체주식가치": stock_result.get("total_equity_value"),
    }

    # 자료 충족도는 단순 필드 개수가 아니라 실제 컨설팅에 필요한 자료군을
    # 가중 평가합니다. 모든 자료가 갖춰져도 100% 대신 최대 95%로 제한해
    # 추가 확인 가능성이 항상 남아 있음을 표현합니다.
    profile_fields = [business_no, industry, address, establishment, employees]
    finance_fields = [sales, operating_profit, net_income, assets, liabilities, equity]

    profile_ratio = sum(1 for value in profile_fields if _clean(value)) / len(profile_fields)
    finance_ratio = sum(1 for value in finance_fields if _clean(value)) / len(finance_fields)

    preference_values = []
    for key in ("관심지원분야", "매칭키워드", "제외키워드", "자금사용목적"):
        value = preferences.get(key) if isinstance(preferences, dict) else None
        if isinstance(value, (list, tuple, set)):
            preference_values.extend([item for item in value if _clean(item)])
        elif _clean(value):
            preference_values.append(value)
    has_preferences = bool(preference_values)

    completeness_components = [
        ("기업 기본정보", 15.0 * profile_ratio, profile_ratio >= 0.8),
        ("재무자료", 30.0 * finance_ratio, finance_ratio >= 0.8),
        ("등기정보", 12.0 if registry else 0.0, bool(registry)),
        ("주가평가", 12.0 if stock_record else 0.0, bool(stock_record)),
        ("정관검토", 10.0 if articles_review else 0.0, bool(articles_review)),
        ("상담일지", 8.0 if consultation_count else 0.0, bool(consultation_count)),
        ("직원현황", 8.0 if employee_source_available else 0.0, employee_source_available),
        ("매칭설정", 6.0 if has_preferences else 0.0, has_preferences),
    ]
    raw_completeness = sum(score for _, score, _ in completeness_components)
    completeness = min(95, round(raw_completeness / 95.0 * 100))

    if completeness >= 85:
        completeness_status = "매우 충분"
    elif completeness >= 70:
        completeness_status = "충분"
    elif completeness >= 50:
        completeness_status = "보통"
    elif completeness >= 30:
        completeness_status = "보완 필요"
    else:
        completeness_status = "자료 부족"

    missing_sources = [
        name for name, _, is_ready in completeness_components if not is_ready
    ]

    return {
        "company_name": company_name,
        "business_no": business_no,
        "industry": industry,
        "address": address,
        "establishment": establishment,
        "employees": employees,
        "employees_source": employees_source,
        "cretop_employees": cretop_employees,
        "employee_context": employee_context,
        "employee_summary": employee_summary,
        "workplace_counts": workplace_counts,
        "comprehensive_diagnosis": comprehensive_diagnosis,
        "sales": sales,
        "operating_profit": operating_profit,
        "net_income": net_income,
        "assets": assets,
        "liabilities": liabilities,
        "equity": equity,
        "operating_margin": operating_margin,
        "net_margin": net_margin,
        "debt_ratio": debt_ratio,
        "strengths": strengths,
        "cautions": cautions,
        "strategy": strategy,
        "questions": questions,
        "preferences": preferences,
        "registry": registry,
        "stock_summary": stock_summary,
        "consultation_context": consultation_context,
        "data_sources": {
            "cretop": bool(len(customer.index)),
            "financial": bool(financial),
            "comprehensive_financial_diagnosis": bool(diagnosis_findings),
            "registry": bool(registry),
            "stock": bool(stock_record),
            "consultation": bool(consultation_count),
            "employee_status": employee_source_available,
            "articles_review": bool(articles_review),
            "matching_preferences": has_preferences,
        },
        "completeness": completeness,
        "completeness_status": completeness_status,
        "completeness_components": completeness_components,
        "missing_sources": missing_sources,
    }


def _style_sheet(worksheet) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 3, 12),
            55,
        )


def build_consulting_excel_report(analysis: dict[str, Any]) -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary = workbook.create_sheet("기업요약")
    summary.append(["항목", "내용"])
    summary_rows = [
        ["업체명", analysis["company_name"]],
        ["사업자등록번호", analysis["business_no"]],
        ["업종", analysis["industry"]],
        ["사업장 소재지", analysis["address"]],
        ["설립일", analysis["establishment"]],
        ["종업원수", analysis["employees"]],
        ["정보완성도", f"{analysis['completeness']}%"],
    ]
    for row in summary_rows:
        summary.append(row)
    _style_sheet(summary)

    diagnosis_engine = workbook.create_sheet("종합재무진단")
    diagnosis_engine.append(
        [
            "우선순위",
            "진단주제",
            "분류",
            "중요도",
            "확신도",
            "판정",
            "관찰사실",
            "대표질문",
            "요청자료",
            "확인 후 방향",
        ]
    )
    for index, finding in enumerate(
        analysis.get("comprehensive_diagnosis", {}).get(
            "findings",
            [],
        ),
        start=1,
    ):
        diagnosis_engine.append(
            [
                index,
                finding.get("title", ""),
                finding.get("category", ""),
                finding.get("importance", 0),
                finding.get("confidence", 0),
                finding.get("status", ""),
                "\n".join(finding.get("facts", []) or []),
                "\n".join(finding.get("questions", []) or []),
                "\n".join(finding.get("documents", []) or []),
                "\n".join(finding.get("directions", []) or []),
            ]
        )
    _style_sheet(diagnosis_engine)

    financial = workbook.create_sheet("재무분석")
    financial.append(["항목", "금액 또는 비율"])
    financial_rows = [
        ["매출액", _format_money(analysis["sales"])],
        ["영업이익", _format_money(analysis["operating_profit"])],
        ["당기순이익", _format_money(analysis["net_income"])],
        ["자산총계", _format_money(analysis["assets"])],
        ["부채총계", _format_money(analysis["liabilities"])],
        ["자본총계", _format_money(analysis["equity"])],
        [
            "영업이익률",
            (
                f"{analysis['operating_margin']:.1f}%"
                if analysis["operating_margin"] is not None
                else "-"
            ),
        ],
        [
            "순이익률",
            (
                f"{analysis['net_margin']:.1f}%"
                if analysis["net_margin"] is not None
                else "-"
            ),
        ],
        [
            "부채비율",
            (
                f"{analysis['debt_ratio']:.1f}%"
                if analysis["debt_ratio"] is not None
                else "-"
            ),
        ],
    ]
    for row in financial_rows:
        financial.append(row)
    _style_sheet(financial)

    diagnosis = workbook.create_sheet("컨설팅진단")
    diagnosis.append(["구분", "내용"])
    for item in analysis["strengths"]:
        diagnosis.append(["강점", item])
    for item in analysis["cautions"]:
        diagnosis.append(["확인 필요", item])
    for item in analysis["strategy"]:
        diagnosis.append(["추천 전략", item])
    _style_sheet(diagnosis)

    questions = workbook.create_sheet("상담질문")
    questions.append(["번호", "대표 상담 질문"])
    for index, item in enumerate(
        analysis["questions"],
        start=1,
    ):
        questions.append([index, item])
    _style_sheet(questions)

    preferences = workbook.create_sheet("매칭설정")
    preferences.append(["항목", "내용"])
    preference_data = analysis["preferences"]
    preference_rows = [
        [
            "매칭키워드",
            ", ".join(preference_data.get("매칭키워드", []) or []),
        ],
        [
            "관심지원분야",
            ", ".join(preference_data.get("관심지원분야", []) or []),
        ],
        [
            "제외키워드",
            ", ".join(preference_data.get("제외키워드", []) or []),
        ],
        ["자금사용목적", preference_data.get("자금사용목적", "")],
        ["투자예정금액", preference_data.get("투자예정금액", "")],
        ["투자예정시기", preference_data.get("투자예정시기", "")],
    ]
    for row in preference_rows:
        preferences.append(row)
    _style_sheet(preferences)

    stock = workbook.create_sheet("주가평가요약")
    stock.append(["항목", "내용"])
    stock_summary = analysis["stock_summary"]
    stock_rows = [
        ["평가기준일", stock_summary.get("평가기준일", "")],
        ["발행주식총수", stock_summary.get("발행주식총수", "")],
        [
            "1주당평가액",
            _format_money(stock_summary.get("1주당평가액")),
        ],
        [
            "기업전체주식가치",
            _format_money(stock_summary.get("기업전체주식가치")),
        ],
    ]
    for row in stock_rows:
        stock.append(row)
    _style_sheet(stock)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def render_ai_consulting_report_page(
    user_id: str,
    user_name: str = "",
    customer: pd.Series | None = None,
    embedded: bool = False,
    key_prefix: str = "ai_report",
) -> None:
    if not embedded:
        st.markdown("## AI 컨설팅 리포트")
        st.caption(
            "기존 고객DB·크레탑 재무정보·등기정보·주가평가·매칭설정을 "
            "읽기 전용으로 결합해 상담용 리포트를 만듭니다."
        )

    if customer is None:
        customers = load_registered_customers(
            get_user_cumulative_db_path(user_id)
        )
        if customers.empty:
            st.info(
                "등록된 고객이 없습니다. 먼저 크레탑 자동등록으로 고객을 등록해주세요."
            )
            return

        labels, row_map = build_customer_labels(customers)
        selected_label = st.selectbox(
            "컨설팅 리포트 대상 고객",
            labels,
            key=f"{key_prefix}_customer_selector",
        )
        customer = customers.loc[row_map[selected_label]]
    business_no = _normalize_business_no(
        customer.get("사업자등록번호", "")
    )

    company_name = _clean(customer.get("업체명", ""))
    financial = _financial_snapshot(user_id, business_no)
    registry = _registry_snapshot(user_id, business_no)
    stock_record = _latest_stock_record(
        user_id,
        business_no,
        company_name=company_name,
    )
    preferences = get_matching_preferences(
        user_id,
        business_no,
    )
    consultation_context = _build_consultation_context(
        user_id,
        business_no,
        company_name,
    )
    try:
        employee_context = get_latest_employee_status(
            user_id,
            business_no,
            company_name,
        )
    except Exception:
        employee_context = {}

    analysis = build_consulting_analysis(
        customer,
        financial,
        registry,
        stock_record,
        preferences,
        consultation_context=consultation_context,
        employee_context=employee_context,
    )

    st.markdown(
        f"""
        <div style="
            padding:22px 24px;
            border-radius:20px;
            background:linear-gradient(135deg,#123d7a,#2563eb);
            color:white;
            margin:8px 0 18px 0;
            box-shadow:0 12px 28px rgba(37,99,235,.18);
        ">
            <div style="font-size:1.45rem;font-weight:800;">
                {analysis['company_name']} 컨설팅 리포트
            </div>
            <div style="margin-top:7px;opacity:.9;">
                사업자번호 {analysis['business_no']} ·
                작성자 {user_name or '-'} ·
                {datetime.now().strftime('%Y-%m-%d')}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    _inject_report_css()
    kpi_columns = st.columns(4, gap="medium")
    kpi_items = [
        ("자료 충족도", f"{analysis['completeness']}%", analysis.get("completeness_status", "보완 필요"), "blue"),
        ("매출액", _format_money(analysis["sales"]), "최근 결산 기준", "green"),
        ("영업이익", _format_money(analysis["operating_profit"]), "본업 수익성", "purple"),
        ("당기순이익", _format_money(analysis["net_income"]), "세후 최종손익", "orange"),
    ]
    for column, (label, value, note, tone) in zip(kpi_columns, kpi_items):
        with column:
            st.markdown(_metric_card(label, value, note, tone), unsafe_allow_html=True)

    missing_sources = analysis.get("missing_sources", []) or []
    if missing_sources:
        st.info(
            "자료 충족도에 반영되지 않은 항목: " + ", ".join(missing_sources)
            + " · 해당 자료를 등록하면 진단 범위와 정확도가 높아집니다."
        )

    source_columns = st.columns(5, gap="medium")
    source_columns[0].metric(
        "등록 등기정보",
        "1건" if registry else "0건",
    )
    source_columns[1].metric(
        "저장 주가평가",
        "1건" if stock_record else "0건",
    )
    source_columns[2].metric(
        "직원현황",
        (
            f"{analysis.get('employee_summary', {}).get('active_count', 0)}명"
            if analysis.get("data_sources", {}).get("employee_status")
            else "0명"
        ),
    )
    source_columns[3].metric(
        "상담일지",
        f"{consultation_context.get('count', 0)}건",
    )
    source_columns[4].metric(
        "녹취 반영",
        f"{consultation_context.get('transcript_count', 0)}건",
    )

    with st.expander("자료 연결 진단", expanded=False):
        st.write(f"조회 회사명: {company_name or '-'}")
        st.write(f"조회 사업자등록번호: {business_no or '-'}")
        st.write(
            "등기정보 연결: "
            + ("성공" if registry else "실패 — 동일 사업자번호의 저장 등기정보 없음")
        )
        st.write(
            "주가평가 연결: "
            + ("성공" if stock_record else "실패 — 사업자번호·회사명이 일치하는 평가자료 없음")
        )
        st.write(
            f"상담일지 연결: {consultation_context.get('count', 0)}건"
        )
        st.write(
            f"녹취 전문 포함 상담일지: "
            f"{consultation_context.get('transcript_count', 0)}건"
        )
        st.write(
            "직원현황 연결: "
            + (
                f"성공 — 가입중 "
                f"{analysis.get('employee_summary', {}).get('active_count', 0)}명"
                if analysis.get("data_sources", {}).get("employee_status")
                else "실패 — 동일 사업자번호의 저장 직원현황 없음"
            )
        )
        st.write(
            f"종업원수 적용 기준: {analysis.get('employees_source', '-')}"
        )

    st.markdown("### 기업 기본정보")
    base_df = pd.DataFrame(
        [
            ["업종", analysis["industry"] or "-"],
            ["사업장 소재지", analysis["address"] or "-"],
            ["설립일", analysis["establishment"] or "-"],
            [
                "종업원수",
                (
                    f"{_clean(analysis['employees']) or '-'}명 "
                    f"({analysis.get('employees_source', '-')})"
                ),
            ],
            [
                "크레탑 종업원수",
                (
                    f"{_clean(analysis.get('cretop_employees'))}명"
                    if _clean(analysis.get("cretop_employees"))
                    else "-"
                ),
            ],
        ],
        columns=["항목", "내용"],
    )
    st.dataframe(
        base_df,
        hide_index=True,
        use_container_width=True,
    )

    if analysis.get("data_sources", {}).get("employee_status"):
        st.markdown("### 직원·고용 현황")
        employee_summary = analysis.get("employee_summary", {}) or {}
        e1, e2, e3, e4 = st.columns(4)
        e1.metric(
            "가입중",
            f"{employee_summary.get('active_count', 0)}명",
        )
        e2.metric(
            "최근 6개월 취득",
            f"{employee_summary.get('recent_6m_count', 0)}명",
        )
        e3.metric(
            "청년 추정",
            f"{employee_summary.get('youth_count', 0)}명",
        )
        e4.metric(
            "고령자 추정",
            f"{employee_summary.get('senior_count', 0)}명",
        )

        workplace_counts = analysis.get("workplace_counts", {}) or {}
        if workplace_counts:
            st.dataframe(
                pd.DataFrame(
                    [
                        {"사업장": name, "가입중 인원": count}
                        for name, count in sorted(workplace_counts.items())
                    ]
                ),
                hide_index=True,
                use_container_width=True,
            )

    st.markdown("### 종합 재무진단 우선순위")
    comprehensive = analysis.get("comprehensive_diagnosis", {}) or {}
    findings = comprehensive.get("findings", []) or []

    if not findings:
        st.info(
            "현재 크레탑·재무자료에서 진단 가능한 세부 계정과목을 충분히 찾지 못했습니다. "
            "상세 재무제표나 계정별원장을 추가하면 진단 범위가 넓어집니다."
        )
    else:
        priority_df = pd.DataFrame(
            [
                {
                    "순위": index,
                    "진단 주제": finding.get("title", ""),
                    "분류": finding.get("category", ""),
                    "중요도": finding.get("importance", 0),
                    "확신도": finding.get("confidence", 0),
                    "종합우선도": finding.get("priority", 0),
                    "판정": finding.get("status", ""),
                }
                for index, finding in enumerate(findings, start=1)
            ]
        )
        st.dataframe(
            priority_df,
            hide_index=True,
            use_container_width=True,
        )

        for index, finding in enumerate(findings, start=1):
            with st.expander(
                f"{index}. {finding.get('title', '')} · "
                f"중요도 {finding.get('importance', 0)} · "
                f"확신도 {finding.get('confidence', 0)}",
                expanded=index <= 3,
            ):
                st.markdown("**관찰된 사실**")
                for fact in finding.get("facts", []) or []:
                    st.write(f"- {fact}")

                st.markdown("**대표 확인 질문**")
                for question in finding.get("questions", []) or []:
                    st.write(f"- {question}")

                st.markdown("**추가 요청자료**")
                for document in finding.get("documents", []) or []:
                    st.write(f"- {document}")

                st.markdown("**확인 후 컨설팅 방향**")
                for direction in finding.get("directions", []) or []:
                    st.write(f"- {direction}")

                account_hits = finding.get("account_hits", []) or []
                if account_hits:
                    with st.container(border=True):
                        st.caption("탐지된 계정과목 근거")
                        hit_df = pd.DataFrame(
                            [
                                {
                                    "계정 경로": hit.get("account", ""),
                                    "연도": hit.get("year") or "-",
                                    "금액": _format_money(hit.get("value")),
                                }
                                for hit in account_hits[:20]
                            ]
                        )
                        st.dataframe(
                            hit_df,
                            hide_index=True,
                            use_container_width=True,
                        )

        st.warning(
            "이 결과는 계정과목 이상징후를 찾는 사전진단입니다. "
            "특히 가지급금 가능성은 계정별원장과 거래증빙 확인 전에는 확정할 수 없습니다."
        )

    st.markdown("### 재무 진단")
    ratio_items = [
        ("영업이익률", f"{analysis['operating_margin']:.1f}%" if analysis["operating_margin"] is not None else "-", "매출 대비 영업이익", "green"),
        ("순이익률", f"{analysis['net_margin']:.1f}%" if analysis["net_margin"] is not None else "-", "매출 대비 당기순이익", "blue"),
        ("부채비율", f"{analysis['debt_ratio']:.1f}%" if analysis["debt_ratio"] is not None else "-", "자본 대비 부채", "red" if (analysis["debt_ratio"] or 0) >= 200 else "purple"),
    ]
    ratio_columns = st.columns(3, gap="medium")
    for column, (label, value, note, tone) in zip(ratio_columns, ratio_items):
        with column:
            st.markdown(_metric_card(label, value, note, tone), unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)
    with c1:
        with st.container(border=True):
            st.markdown("#### 강점")
            for item in analysis["strengths"] or [
                "추가 재무정보 확인이 필요합니다."
            ]:
                st.markdown(f"- {item}")
    with c2:
        with st.container(border=True):
            st.markdown("#### 확인 필요")
            for item in analysis["cautions"] or [
                "현재 주요 경고사항은 확인되지 않았습니다."
            ]:
                st.markdown(f"- {item}")
    with c3:
        with st.container(border=True):
            st.markdown("#### 추천 전략")
            for item in analysis["strategy"]:
                st.markdown(f"- {item}")

    st.markdown("### 대표 상담 질문")
    for index, question in enumerate(
        analysis["questions"],
        start=1,
    ):
        st.markdown(f"**{index}.** {question}")

    with st.expander("주가평가 요약", expanded=False):
        stock_summary = analysis["stock_summary"]
        if stock_record:
            st.write(
                f"평가기준일: {stock_summary.get('평가기준일') or '-'}"
            )
            st.write(
                "발행주식총수: "
                f"{stock_summary.get('발행주식총수') or '-'}"
            )
            st.write(
                "1주당 평가액: "
                f"{_format_money(stock_summary.get('1주당평가액'))}"
            )
            st.write(
                "기업 전체 주식가치: "
                f"{_format_money(stock_summary.get('기업전체주식가치'))}"
            )
        else:
            st.info(
                "저장된 주가평가 결과가 없습니다. 주가평가 메뉴에서 평가 후 저장해주세요."
            )

    with st.expander("정책자금 매칭설정", expanded=False):
        if not preferences:
            st.info(
                "저장된 고객별 매칭키워드가 없습니다."
            )
        else:
            setting_rows = [
                {
                    "항목": "매칭키워드",
                    "내용": ", ".join(
                        preferences.get("매칭키워드", []) or []
                    ) or "-",
                },
                {
                    "항목": "관심지원분야",
                    "내용": ", ".join(
                        preferences.get("관심지원분야", []) or []
                    ) or "-",
                },
                {
                    "항목": "제외키워드",
                    "내용": ", ".join(
                        preferences.get("제외키워드", []) or []
                    ) or "-",
                },
                {
                    "항목": "자금사용목적",
                    "내용": _clean(
                        preferences.get("자금사용목적", "")
                    ) or "-",
                },
                {
                    "항목": "투자예정금액",
                    "내용": _clean(
                        preferences.get("투자예정금액", "")
                    ) or "-",
                },
                {
                    "항목": "투자예정시기",
                    "내용": _clean(
                        preferences.get("투자예정시기", "")
                    ) or "-",
                },
            ]
            st.dataframe(
                pd.DataFrame(setting_rows),
                hide_index=True,
                use_container_width=True,
            )

            saved_policies = [
                item
                for item in (
                    preferences.get("저장정책자금", []) or []
                )
                if isinstance(item, dict)
            ]
            if saved_policies:
                st.markdown("#### AI 코파일럿 반영 정책자금")
                st.caption(
                    f"{len(saved_policies)}건 · 최소점수 "
                    f"{preferences.get('저장정책자금_최소점수', '-')}점 · "
                    f"저장일시 "
                    f"{preferences.get('저장정책자금_저장일시', '-')}"
                )
                policy_rows = []
                for item in saved_policies[:30]:
                    policy_rows.append(
                        {
                            "점수": item.get("score", ""),
                            "분류": item.get("category", ""),
                            "공고명": item.get("title", ""),
                            "기관": item.get("agency", ""),
                            "신청종료": item.get("end_date", ""),
                        }
                    )
                st.dataframe(
                    pd.DataFrame(policy_rows),
                    hide_index=True,
                    use_container_width=True,
                )
            else:
                st.info(
                    "AI 코파일럿에 반영하도록 저장한 정책자금 추천이 없습니다."
                )

    # 대표님 PDF에 AI 절세진단 요약을 함께 전달한다.
    try:
        analysis["tax_diagnosis"] = build_tax_diagnosis(user_id, customer)
    except Exception:
        analysis["tax_diagnosis"] = {}

    logo_path = Path(__file__).resolve().parent / "assets" / "oasis_logo.png"
    try:
        pdf_bytes = build_representative_pdf(
            analysis,
            consultant_name=user_name,
            logo_path=str(logo_path) if logo_path.exists() else None,
        )
    except Exception as exc:
        pdf_bytes = b""
        log_path = write_runtime_error(
            "representative_pdf_generation", exc,
            {"company_name": analysis.get("company_name", "")},
        )
        suffix = f" (오류로그: {log_path})" if log_path else ""
        st.warning(f"PDF 생성 중 오류가 발생했습니다: {exc}{suffix}")

    excel_bytes = build_consulting_excel_report(analysis)
    safe_company = re.sub(
        r'[\\/:*?"<>|]',
        "_",
        analysis["company_name"] or "고객",
    )
    filename = (
        f"AI컨설팅리포트_{safe_company}_"
        f"{datetime.now().strftime('%Y%m%d')}.xlsx"
    )

    pdf_filename = (
        f"오아시스_AI기업컨설팅보고서_{safe_company}_"
        f"{datetime.now().strftime('%Y%m%d')}.pdf"
    )
    download_columns = st.columns(2, gap="medium")
    with download_columns[0]:
        st.download_button(
            "대표님용 PDF 보고서 다운로드",
            data=pdf_bytes,
            file_name=pdf_filename,
            mime="application/pdf",
            type="primary",
            use_container_width=True,
            disabled=not bool(pdf_bytes),
            key=f"{key_prefix}_pdf_download",
        )
    with download_columns[1]:
        st.download_button(
            "내부 검토용 엑셀 다운로드",
            data=excel_bytes,
            file_name=filename,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key=f"{key_prefix}_excel_download",
        )

    st.caption(
        "화면은 영업사원의 상담 준비용이며, PDF는 대표님 제출용으로 내부 메모·상담질문을 제외해 구성됩니다. "
        "최종 지원 가능 여부는 최신 공고와 증빙자료를 추가 확인해야 합니다."
    )
